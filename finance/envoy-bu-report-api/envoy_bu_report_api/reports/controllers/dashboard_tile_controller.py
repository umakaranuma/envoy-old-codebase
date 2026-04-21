from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import connection
from django.http import HttpResponse, FileResponse, Http404
from django.conf import settings
from django.urls import reverse
import json
import logging
import io
import os
import tempfile
import uuid
import base64
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from mServices import QueryBuilderService
from mServices.ValidatorService import ValidatorService

from services.EntityService import EntityService
from ..services.response_service import ResponseService
from ..services.data_expoter import ExportToPdf, ExportToCsv, ExportToExcel, SQLToExcelExporter
from ..services.sql_helper_service import SqlHelperService
from ..services.s3_service import S3PresignedService

logger = logging.getLogger(__name__)


class DashboardTileController:
    @staticmethod
    @csrf_exempt
    @require_http_methods(["GET", "PUT", "DELETE"])
    def dashboard_tile_one(request, id):
        """Handle GET, PUT, DELETE requests for a single dashboard tile"""
        if request.method == "GET":
            return DashboardTileController.get_one(request, id)
        
        if request.method == "PUT":
            return DashboardTileController.store_or_update(request, id)
        
        if request.method == "DELETE":
            return DashboardTileController.delete(request, id)

    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST", "PUT", "GET"])
    def dashboard_tile_access(request, id=None):
        """Handle GET, POST, PUT requests for dashboard tiles"""
        if request.method == "GET":
            return DashboardTileController.get_all(request)
        
        if request.method == "POST":
            return DashboardTileController.store_or_update(request, id=None)
        
        if request.method == "PUT":
            return DashboardTileController.store_or_update(request, id)

    @staticmethod
    def get_all(request):
        """Get all dashboard tiles with filtering and pagination"""
        try:
            # Get query parameters
            filters = json.loads(request.GET.get('filters', '{}'))
            search = request.GET.get('search', '')
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 10))
            sort_by = request.GET.get('sort_by', 'tile.id')
            sort_dir = request.GET.get('sort_dir', 'asc')
            
            # Build query using QueryBuilderService
            query = (
                QueryBuilderService("rep_report_dashboard_tiles as tile")
                .leftJoin("rep_report_dashboards as dashboard", "dashboard.id", "tile.dashboard_id_id")
                .leftJoin("rep_report_charts as chart", "chart.id", "tile.chart_id_id")
                .leftJoin("rep_reports as report", "report.id", "tile.report_id_id")
                .select(
                    "tile.*",
                    "dashboard.title as dashboard_title",
                    "chart.title as chart_title",
                    "report.title as report_title"
                )
                .whereNull("tile.deleted_at")
            )
            
            # Apply filters
            if filters.get('type'):
                query = query.where("tile.type", filters['type'])
            if filters.get('dashboard_id'):
                query = query.where("tile.dashboard_id_id", filters['dashboard_id'])
            if filters.get('report_id'):
                query = query.where("tile.report_id_id", filters['report_id'])
            
            # Apply search
            if search:
                query = query.where("tile.type", search, "LIKE")
            
            # Apply sorting and pagination
            data = query.paginate(page, limit, ['tile.id', 'tile.type'], sort_by, sort_dir)
            
            return ResponseService.response('SUCCESS', data, "Dashboard tiles retrieved successfully")
            
        except Exception as e:
            logger.error(f"Error in get_all: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))

    @staticmethod
    def get_one(request, id):
        """Get a single dashboard tile by ID"""
        try:
            tile = (
                QueryBuilderService("rep_report_dashboard_tiles as tile")
                .leftJoin("rep_report_dashboards as dashboard", "dashboard.id", "tile.dashboard_id_id")
                .leftJoin("rep_report_charts as chart", "chart.id", "tile.chart_id_id")
                .leftJoin("rep_reports as report", "report.id", "tile.report_id_id")
                .select(
                    "tile.*",
                    "dashboard.title as dashboard_title",
                    "chart.title as chart_title",
                    "report.title as report_title"
                )
                .where("tile.id", id)
                .whereNull("tile.deleted_at")
                .first()
            )
            
            if not tile:
                return ResponseService.response('NO_DATA_FOUND', None, "data_not_found")
            
            return ResponseService.response('SUCCESS', tile, "Dashboard tile retrieved successfully")
            
        except Exception as e:
            logger.error(f"Error in get_one: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))

    @staticmethod
    def store_or_update(request, id=None):
        """Create or update a dashboard tile"""
        try:
            data = json.loads(request.body)
            
            # Validation
            rules = {
                "type": "required|string",
                "dashboard_id": "required|integer",
                "report_id": "required|integer",
            }
            
            # Add chart_id validation if type is "Chart"
            if data.get('type') == "Chart":
                rules["chart_id"] = "required|integer"
            
            errors = ValidatorService.validate(data, rules)
            if errors:
                return ResponseService.response("VALIDATION_ERROR", errors, "Validation Error")
            
            # Validate type value
            if data.get('type') not in ["Table", "Chart"]:
                return ResponseService.response("VALIDATION_ERROR", {"type": ["Type must be 'Table' or 'Chart'"]}, "Invalid type value")
            
            if id:
                # Update existing tile
                tile = QueryBuilderService("rep_report_dashboard_tiles").where("id", id).first()
                if not tile:
                    return ResponseService.response('NO_DATA_FOUND', None, "data_not_found")
                
                QueryBuilderService("rep_report_dashboard_tiles").where("id", id).update(data)
                message = "default_update_success_msg"
            else:
                entity = EntityService.store_entity("report_dashboard_tiles", request)
                data['entity_id'] = entity['id']
                data['created_by_id'] = request.user.id
                data['dashboard_id_id'] = data['dashboard_id']
                data['report_id_id'] = data['report_id']
                if data.get('type') == "Chart":
                    data['chart_id_id'] = data['chart_id']
                else:
                    data["chart_id_id"] = None
                # Create new tile
                tile = QueryBuilderService("rep_report_dashboard_tiles").insert(data)
                message = "default_create_success_msg"
            
            return ResponseService.response('SUCCESS', None, message)
            
        except Exception as e:
            logger.error(f"Error in store_or_update: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))

    @staticmethod
    def delete(request, id):
        """Soft delete a dashboard tile"""
        try:
            tile = QueryBuilderService("rep_report_dashboard_tiles").where("id", id).whereNull("deleted_at").first()
            
            if not tile:
                return ResponseService.response('NO_DATA_FOUND', None, "data_not_found")
            
            # Soft delete
            QueryBuilderService("rep_report_dashboard_tiles").where("id", id).update({
                "deleted_at": timezone.now(),
                "deleted_by_id": request.user.id
            })
            
            return ResponseService.response('SUCCESS', None, "default_delete_success_msg")
            
        except Exception as e:
            logger.error(f"Error in delete: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))

    @staticmethod
    @csrf_exempt
    @require_http_methods(["GET"])
    def get_by_dashboard(request, dashboard_id):
        """Get all tiles for a specific dashboard"""
        try:
            tiles = (
                QueryBuilderService("rep_report_dashboard_tiles as tile")
                .leftJoin("rep_report_charts as chart", "chart.id", "tile.chart_id_id")
                .leftJoin("rep_reports as report", "report.id", "tile.report_id_id")
                .select(
                    "tile.id",
                    "tile.entity_id",
                    "tile.type",
                    "tile.chart_id_id",
                    "chart.title as chart_title",
                    "tile.report_id_id",
                    "report.title as report_title",
                    "tile.created_at"
                )
                .where("tile.dashboard_id_id", dashboard_id)
                .get()
            )
            
            return ResponseService.response('SUCCESS', {"data": tiles}, "Dashboard tiles retrieved successfully")
            
        except Exception as e:
            logger.error(f"Error in get_by_dashboard: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e)) 

    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    def html_export(request):
        """Export report data to PDF, CSV, or Excel based on type and report_id parameters"""
        try:
            data = json.loads(request.body)
            
            # Validation
            rules = {
                "type": "required|string|in:pdf,excel",
                "report_id": "required|integer",
                "search": "optional|string",
                "filters": "optional|dict",
                "sort_by": "optional|string",
                "sort_dir": "optional|string"
            }
            errors = ValidatorService.validate(data, rules)
            if errors:
                return ResponseService.response("VALIDATION_ERROR", errors, "Validation Error")
            
            export_type = data["type"].lower()
            report_id = data['report_id']
            search = data.get('search', '').strip()
            filters = data.get('filters', {})
            sort_by = data.get('sort_by')
            sort_dir = data.get('sort_dir', 'desc')
            
         
            # Get report
            report = QueryBuilderService("rep_reports").where("id", report_id).first()
            if not report:
                return ResponseService.response('NO_DATA_FOUND', None, "data_not_found")
            
            # Parse JSON config
            json_config = report.get('json') or {}
            if isinstance(json_config, str):
                try:
                    json_config = json.loads(json_config)
                except Exception:
                    json_config = {}
            
            # Clean up corrupted field data (same logic as export_report_to_excel)
            fields = json_config.get('fields', [])
            if fields:
                cleaned_fields = []
                for field in fields:
                    if isinstance(field, dict) and 'label' in field:
                        label = field['label']
                        if ' AS ' in label and '"' in label:
                            try:
                                as_part = label.split(' AS ')[-1]
                                if as_part.startswith('"') and as_part.endswith('"'):
                                    clean_label = as_part[1:-1]
                                elif as_part.startswith('\\"') and as_part.endswith('\\"'):
                                    clean_label = as_part[2:-2]
                                else:
                                    clean_label = label
                            except:
                                clean_label = label
                        elif ' AS ' in label:
                            try:
                                as_part = label.split(' AS ')[-1]
                                clean_label = as_part
                            except:
                                clean_label = label
                        elif '\\"' in label:
                            try:
                                start = label.find('\\"') + 2
                                end = label.rfind('\\"')
                                if start > 1 and end > start:
                                    clean_label = label[start:end]
                                else:
                                    clean_label = label
                            except:
                                clean_label = label
                        else:
                            clean_label = label
                        
                        if clean_label and len(clean_label) > 100:
                            if '\\"' in clean_label:
                                parts = clean_label.split('\\"')
                                if len(parts) >= 3:
                                    clean_label = parts[1]
                                else:
                                    clean_label = clean_label[:50] + "..."
                        
                        cleaned_field = field.copy()
                        cleaned_field['label'] = clean_label
                        cleaned_fields.append(cleaned_field)
                    else:
                        cleaned_fields.append(field)
                fields = cleaned_fields
            
            skip_columns = json_config.get('skip_columns', [])
            
            # Check if report has a query
            if not report.get('query'):
                logger.warning(f"Report {report_id} has no query")
                return ResponseService.response('NO_DATA_FOUND', None, "Report has no SQL query")
            
            # Build SQL (same logic as export_report_to_excel - no pagination, get all data)
            sql = report['query'].rstrip("; \n\r\t")
            logger.info(f"Building SQL for report {report_id} {export_type} export: {sql[:200]}...")
            
            # Remove skipped columns
            if skip_columns:
                sql = SqlHelperService.remove_skipped_columns_from_sql(sql, skip_columns)
            
            # Apply search if provided
            if search:
                searchable_fields = [f"{field['code']}" for field in fields if field.get('dataType') == 'text']
                if searchable_fields:
                    search_conditions = [f"{field} LIKE '%{search}%'" for field in searchable_fields]
                    search_condition = ' OR '.join(search_conditions)
                    sql = SqlHelperService.add_where_condition(sql, search_condition)
            
            # Apply sorting if provided
            allowed_sorting_columns = [field['label'] for field in fields if field['code'] not in [col['code'] for col in skip_columns]]
            if sort_by and sort_by in allowed_sorting_columns:
                sql = SqlHelperService.apply_sort(sql, sort_by, sort_dir, allowed_sorting_columns)
            
            # Execute query to get all data (no pagination)
            try:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    columns = [col[0] for col in cursor.description]
                    all_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
                    logger.info(f"Retrieved {len(all_data)} records for {export_type} export")
            except Exception as query_error:
                logger.error(f"Error executing query for report {report_id}: {str(query_error)}")
                return ResponseService.response('INTERNAL_SERVER_ERROR', f"Query execution failed: {str(query_error)}")
            
            # Handle different export types
            if export_type == "excel":
                # Use SQLToExcelExporter for Excel (same as export_report_to_excel)
                sheet_title = (report['title'] or f"Report_{report_id}")[:31]
                queries = [{
                    "query": sql,
                    "title": sheet_title
                }]
                payload = {
                    "queries": queries
                }
                
                # Print payload being sent to external API
                print("=" * 80)
                print("EXTERNAL API REQUEST - html_export (Excel)")
                print("=" * 80)
                print("Payload:", json.dumps(payload, indent=2))
                print("=" * 80)
                
                exporter = SQLToExcelExporter()
                result = exporter.export(payload)
                
                # Print response from external API
                print("=" * 80)
                print("EXTERNAL API RESPONSE - html_export (Excel)")
                print("=" * 80)
                print("Response:", json.dumps(result, indent=2))
                print("=" * 80)
                
                if result["status"] == "SUCCESS" and result.get("data"):
                    download_url = result["data"].get("download_url")
                    return ResponseService.response("SUCCESS", {
                        "download_url": download_url,
                        "type": export_type
                    }, "file generated successfully")
                else:
                    return ResponseService.response("INTERNAL_SERVER_ERROR", None, result.get("message", "Excel export failed"))
            
            elif export_type == "pdf":
                # Generate PDF file using the same internal logic as html_to_doc_export
                report_title = report.get('title', f'Report_{report_id}')
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                elements = []

                # Define styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#366092'),
                    spaceAfter=30,
                    alignment=1  # Center alignment
                )

                # Add title
                if report_title:
                    title = Paragraph(report_title, title_style)
                    elements.append(title)
                    elements.append(Spacer(1, 0.2*inch))

                # Prepare table data
                table_data = []

                # Add header row
                header_row = [str(col) for col in columns]
                table_data.append(header_row)

                # Add data rows
                for row_data in all_data:
                    row = []
                    for col in columns:
                        value = row_data.get(col, '')
                        if value is None:
                            value = ''
                        row.append(str(value))
                    table_data.append(row)

                # Create table
                table = Table(table_data)

                # Style the table
                table.setStyle(TableStyle([
                    # Header row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    # Data rows
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    # Alternating row colors
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))

                elements.append(table)

                # Build PDF
                doc.build(elements)
                buffer.seek(0)
                file_bytes = buffer.getvalue()

                # Generate filename
                clean_title = DashboardTileController._clean_filename(report_title)
                filename = f"{clean_title}_{timestamp}.pdf"

                # Upload to S3 using presigned service (same as html_to_doc_export)
                try:
                    upload_result = S3PresignedService.upload_file_to_s3(
                        file_content=file_bytes,
                        file_name=filename,
                        folder="exports"
                    )

                    file_key = upload_result["file_key"]

                    cdn_base_url = os.getenv("CDN_BASE_URL", "")
                    document_key = os.getenv("DOCUMENT_KEY", "")

                    if cdn_base_url:
                        from urllib.parse import urlencode
                        if document_key:
                            params = {'document_key': document_key}
                            query_string = urlencode(params)
                            download_url = f"{cdn_base_url.rstrip('/')}/{file_key}?{query_string}"
                        else:
                            download_url = f"{cdn_base_url.rstrip('/')}/{file_key}"
                    else:
                        download_url = upload_result["file_url"]

                    return ResponseService.response("SUCCESS", {
                        "download_url": download_url,
                        "type": export_type,
                        "filename": filename,
                        "file_key": file_key
                    }, "file generated successfully")

                except Exception as s3_error:
                    logger.error(f"Failed to upload PDF file to S3 (html_export): {str(s3_error)}")
                    return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Failed to upload file to S3: {str(s3_error)}")
               
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in request body: {str(e)}")
            return ResponseService.response("VALIDATION_ERROR", {"body": ["Invalid JSON format"]}, "Invalid JSON in request body")
        except Exception as e:
            logger.error(f"Error in html_export: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))
    
    @staticmethod
    def _clean_filename(title):
        """Clean filename by removing spaces and special characters"""
        if not title:
            return "Report"
        # Replace multiple spaces with single space, then replace spaces with underscores
        clean_title = re.sub(r'\s+', ' ', title.strip())
        clean_title = clean_title.replace(' ', '_')
        # Remove any remaining special characters except alphanumeric, underscore, hyphen
        clean_title = re.sub(r'[^a-zA-Z0-9_-]', '', clean_title)
        return clean_title if clean_title else "Report"
    
    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    def html_to_doc_export(request):
        """Export report data to PDF or Excel using Django packages (no external APIs)"""
        try:
            data = json.loads(request.body)
            
            # Validation
            rules = {
                "type": "required|string|in:pdf,excel",
                "report_id": "required|integer",
                "search": "optional|string",
                "filters": "optional|dict",
                "sort_by": "optional|string",
                "sort_dir": "optional|string"
            }
            errors = ValidatorService.validate(data, rules)
            if errors:
                return ResponseService.response("VALIDATION_ERROR", errors, "Validation Error")
            
            export_type = data["type"].lower()
            report_id = data['report_id']
            search = data.get('search', '').strip()
            filters = data.get('filters', {})
            sort_by = data.get('sort_by')
            sort_dir = data.get('sort_dir', 'desc')
            
            # Get report
            report = QueryBuilderService("rep_reports").where("id", report_id).first()
            if not report:
                return ResponseService.response('NO_DATA_FOUND', None, "data_not_found")
            
            # Parse JSON config
            json_config = report.get('json') or {}
            if isinstance(json_config, str):
                try:
                    json_config = json.loads(json_config)
                except Exception:
                    json_config = {}
            
            # Clean up corrupted field data (same logic as html_export)
            fields = json_config.get('fields', [])
            if fields:
                cleaned_fields = []
                for field in fields:
                    if isinstance(field, dict) and 'label' in field:
                        label = field['label']
                        if ' AS ' in label and '"' in label:
                            try:
                                as_part = label.split(' AS ')[-1]
                                if as_part.startswith('"') and as_part.endswith('"'):
                                    clean_label = as_part[1:-1]
                                elif as_part.startswith('\\"') and as_part.endswith('\\"'):
                                    clean_label = as_part[2:-2]
                                else:
                                    clean_label = label
                            except:
                                clean_label = label
                        elif ' AS ' in label:
                            try:
                                as_part = label.split(' AS ')[-1]
                                clean_label = as_part
                            except:
                                clean_label = label
                        elif '\\"' in label:
                            try:
                                start = label.find('\\"') + 2
                                end = label.rfind('\\"')
                                if start > 1 and end > start:
                                    clean_label = label[start:end]
                                else:
                                    clean_label = label
                            except:
                                clean_label = label
                        else:
                            clean_label = label
                        
                        if clean_label and len(clean_label) > 100:
                            if '\\"' in clean_label:
                                parts = clean_label.split('\\"')
                                if len(parts) >= 3:
                                    clean_label = parts[1]
                                else:
                                    clean_label = clean_label[:50] + "..."
                        
                        cleaned_field = field.copy()
                        cleaned_field['label'] = clean_label
                        cleaned_fields.append(cleaned_field)
                    else:
                        cleaned_fields.append(field)
                fields = cleaned_fields
            
            skip_columns = json_config.get('skip_columns', [])
            
            # Check if report has a query
            if not report.get('query'):
                logger.warning(f"Report {report_id} has no query")
                return ResponseService.response('NO_DATA_FOUND', None, "Report has no SQL query")
            
            # Build SQL (same logic as html_export - no pagination, get all data)
            sql = report['query'].rstrip("; \n\r\t")
            logger.info(f"Building SQL for report {report_id} {export_type} export: {sql[:200]}...")
            
            # Remove skipped columns
            if skip_columns:
                sql = SqlHelperService.remove_skipped_columns_from_sql(sql, skip_columns)
            
            # Apply search if provided
            if search:
                searchable_fields = [f"{field['code']}" for field in fields if field.get('dataType') == 'text']
                if searchable_fields:
                    search_conditions = [f"{field} LIKE '%{search}%'" for field in searchable_fields]
                    search_condition = ' OR '.join(search_conditions)
                    sql = SqlHelperService.add_where_condition(sql, search_condition)
            
            # Apply sorting if provided
            allowed_sorting_columns = [field['label'] for field in fields if field['code'] not in [col['code'] for col in skip_columns]]
            if sort_by and sort_by in allowed_sorting_columns:
                sql = SqlHelperService.apply_sort(sql, sort_by, sort_dir, allowed_sorting_columns)
            
            # Execute query to get all data (no pagination)
            try:
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    columns = [col[0] for col in cursor.description]
                    all_data = [dict(zip(columns, row)) for row in cursor.fetchall()]
                    logger.info(f"Retrieved {len(all_data)} records for {export_type} export")
            except Exception as query_error:
                logger.error(f"Error executing query for report {report_id}: {str(query_error)}")
                return ResponseService.response('INTERNAL_SERVER_ERROR', f"Query execution failed: {str(query_error)}")
            
            # Generate file based on type
            report_title = report.get('title', f'Report_{report_id}')
            
            # Generate unique filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if export_type == "excel":
                # Generate Excel file using openpyxl
                wb = Workbook()
                ws = wb.active
                ws.title = report_title[:31]  # Excel sheet names are limited to 31 characters
                
                # Write headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_idx, col_name in enumerate(columns, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data rows
                for row_idx, row_data in enumerate(all_data, start=2):
                    for col_idx, col_name in enumerate(columns, start=1):
                        value = row_data.get(col_name, '')
                        if value is None:
                            value = ''
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-adjust column widths
                for col_idx, col_name in enumerate(columns, start=1):
                    max_length = len(str(col_name))
                    for row_idx in range(2, len(all_data) + 2):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
                
                # Save to BytesIO for S3 upload
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                file_bytes = output.getvalue()
                
                # Generate filename
                clean_title = DashboardTileController._clean_filename(report_title)
                filename = f"{clean_title}_{timestamp}.xlsx"
            
            elif export_type == "pdf":
                # Generate PDF file using reportlab
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                elements = []
                
                # Define styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=16,
                    textColor=colors.HexColor('#366092'),
                    spaceAfter=30,
                    alignment=1  # Center alignment
                )
                
                # Add title
                if report_title:
                    title = Paragraph(report_title, title_style)
                    elements.append(title)
                    elements.append(Spacer(1, 0.2*inch))
                
                # Prepare table data
                table_data = []
                
                # Add header row
                header_row = [str(col) for col in columns]
                table_data.append(header_row)
                
                # Add data rows
                for row_data in all_data:
                    row = []
                    for col in columns:
                        value = row_data.get(col, '')
                        if value is None:
                            value = ''
                        row.append(str(value))
                    table_data.append(row)
                
                # Create table
                table = Table(table_data)
                
                # Style the table
                table.setStyle(TableStyle([
                    # Header row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    # Data rows
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    # Alternating row colors
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ]))
                
                elements.append(table)
                
                # Build PDF
                doc.build(elements)
                buffer.seek(0)
                file_bytes = buffer.getvalue()
                
                # Generate filename
                clean_title = DashboardTileController._clean_filename(report_title)
                filename = f"{clean_title}_{timestamp}.pdf"
            
            else:
                return ResponseService.response("VALIDATION_ERROR", {"type": ["Invalid export type. Must be 'pdf' or 'excel'"]}, "Invalid export type")
            
            # Upload to S3 using presigned service
            try:
                upload_result = S3PresignedService.upload_file_to_s3(
                    file_content=file_bytes,
                    file_name=filename,
                    folder="exports"
                )
                
                # Get file key from upload result
                file_key = upload_result["file_key"]
                
                # Generate CDN URL using file_key
                cdn_base_url = os.getenv("CDN_BASE_URL", "")
                document_key = os.getenv("DOCUMENT_KEY", "")
                
                if cdn_base_url:
                    # Use CDN URL with file key
                    if document_key:
                        # Append document_key parameter if provided
                        from urllib.parse import urlencode
                        params = {'document_key': document_key}
                        query_string = urlencode(params)
                        download_url = f"{cdn_base_url.rstrip('/')}/{file_key}?{query_string}"
                    else:
                        download_url = f"{cdn_base_url.rstrip('/')}/{file_key}"
                else:
                    # Fallback to presigned URL if CDN not configured
                    download_url = upload_result["file_url"]
                
                # Return JSON response with download URL and file_key (same format as html_export)
                return ResponseService.response("SUCCESS", {
                    "download_url": download_url,
                    "type": export_type,
                    "filename": filename,
                    "file_key": file_key
                }, "file generated successfully")
                
            except Exception as s3_error:
                logger.error(f"Failed to upload file to S3: {str(s3_error)}")
                return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Failed to upload file to S3: {str(s3_error)}")
               
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in request body: {str(e)}")
            return ResponseService.response("VALIDATION_ERROR", {"body": ["Invalid JSON format"]}, "Invalid JSON in request body")
        except Exception as e:
            logger.error(f"Error in html_to_doc_export: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))
    
    
    @staticmethod
    def _data_to_html_table(data, columns, title=""):
        """Convert data array to HTML table format"""
        html = f"<html><head><title>{title}</title></head><body>"
        if title:
            html += f"<h2>{title}</h2>"
        html += "<table border='1' cellpadding='5' cellspacing='0' style='border-collapse: collapse; width: 100%;'>"
        
        # Header row
        html += "<thead><tr>"
        for col in columns:
            html += f"<th style='background-color: #f0f0f0; font-weight: bold;'>{col}</th>"
        html += "</tr></thead>"
        
        # Data rows
        html += "<tbody>"
        for row in data:
            html += "<tr>"
            for col in columns:
                value = row.get(col, '')
                # Convert None to empty string and escape HTML
                if value is None:
                    value = ''
                else:
                    value = str(value).replace('<', '&lt;').replace('>', '&gt;').replace('&', '&amp;')
                html += f"<td>{value}</td>"
            html += "</tr>"
        html += "</tbody></table></body></html>"
        
        return html

    @staticmethod
    @csrf_exempt
    @require_http_methods(["POST"])
    def json_to_excel_export(request):
        """Export JSON content to Excel with styling support"""
        try:
            data = json.loads(request.body)
            
            # Validation
            rules = {
                "json_data": "required|list",
                "styles": "dict"  # Optional
            }
            errors = ValidatorService.validate(data, rules)
            if errors:
                return ResponseService.response("VALIDATION_ERROR", errors, "Validation Error")
            
            json_data = data["json_data"]
            styles = data.get("styles", {})
            
            # Validate json_data structure
            if not isinstance(json_data, list) or len(json_data) == 0:
                return ResponseService.response("VALIDATION_ERROR", {"json_data": ["json_data must be a non-empty array"]}, "Invalid json_data format")
            
            # Validate each sheet has required fields
            for sheet in json_data:
                if not isinstance(sheet, dict):
                    return ResponseService.response("VALIDATION_ERROR", {"json_data": ["Each sheet must be an object"]}, "Invalid sheet format")
                if "title" not in sheet:
                    return ResponseService.response("VALIDATION_ERROR", {"json_data": ["Each sheet must have a 'title' field"]}, "Missing sheet title")
                if "data" not in sheet:
                    return ResponseService.response("VALIDATION_ERROR", {"json_data": ["Each sheet must have a 'data' field"]}, "Missing sheet data")
                if not isinstance(sheet["data"], list):
                    return ResponseService.response("VALIDATION_ERROR", {"json_data": ["Sheet data must be an array"]}, "Invalid sheet data format")
            
            # Prepare payload for external service
            # The service expects the JSON data directly, not wrapped in a "json" key
            payload = json_data
            
            logger.info(f"JSON data structure: {json_data}")
            logger.info(f"Styles structure: {styles}")
            
            # Add styles only if they exist and are not empty
            if styles and styles != {}:
                payload = {
                    "json_data": json_data,
                    "styles": styles
                }
                logger.info("Added styles to payload with json_data wrapper")
            else:
                logger.info("No styles provided, sending direct JSON data")
            
            logger.info(f"Sending JSON to Excel export payload: {payload}")
            
            # Call external Excel export service
            import requests
            
            application_name = 'envoy'
            base_url = 'https://exporter.utilities.apptimus.lk'
            url = f"{base_url}/api/{application_name}/export/json-to-excel"
            
            try:
                response = requests.post(url, json=payload)
                logger.info(f"Excel export URL: {url}")
                logger.info(f"Excel export payload: {payload}")
                logger.info(f"Excel export response status: {response.status_code}")
                logger.info(f"Excel export response headers: {dict(response.headers)}")
                logger.info(f"Excel export response: {response.text}")
                
                # Don't raise for status yet, let's see the error details
                if response.status_code != 200:
                    logger.error(f"Excel export failed with status {response.status_code}: {response.text}")
                    return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Excel export service returned error: {response.status_code} - {response.text}")
                
                response_data = response.json()
                
                if response_data.get("success") and isinstance(response_data.get("result"), dict):
                    result_data = response_data["result"]
                    download_url = result_data.get("download_url")
                    
                    if download_url:
                        result_data["download_url"] = f"{base_url}{download_url}"
                    
                    return ResponseService.response("SUCCESS", result_data, response_data.get("message", "Excel generated successfully"))
                else:
                    message = response_data.get("message", "Excel export failed")
                    logger.error(f"Excel export failed: {message}")
                    return ResponseService.response("INTERNAL_SERVER_ERROR", None, message)
                    
            except requests.exceptions.RequestException as e:
                logger.error(f"Excel export request error: {str(e)}")
                return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Failed to connect to Excel export service: {str(e)}")
            except ValueError as ve:
                logger.error(f"Excel export JSON error: {str(ve)}")
                return ResponseService.response("INTERNAL_SERVER_ERROR", None, "Invalid JSON response from Excel export service")
            except Exception as ex:
                logger.error(f"Excel export unexpected error: {str(ex)}")
                return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Unexpected error: {str(ex)}")
            
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON in request body: {str(e)}")
            return ResponseService.response("VALIDATION_ERROR", {"body": ["Invalid JSON format"]}, "Invalid JSON in request body")
        except Exception as e:
            logger.error(f"Error in json_to_excel_export: {str(e)}")
            return ResponseService.response('INTERNAL_SERVER_ERROR', str(e))