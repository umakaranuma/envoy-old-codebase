import json
import os
import io
import tempfile
from django import db
from django.conf import settings
import mServices.ResponseService as ResponseService
import mServices.QueryBuilderService as QueryBuilderService
import mServices.ValidatorService as ValidatorService
from rest_framework.decorators import api_view
from envoy_bu_crm_api.quotation.services.NotificationService import NotificationService
from envoy_bu_crm_api.quotation.services.data_expoter import ExportToPdf
from envoy_bu_crm_api.quotation.services.s3_presigned_service import S3PresignedService
from envoy_bu_crm_api.quotation.services.document_cdn_service import DocumentCDNService
from envoy_bu_crm_api.sales.models.core_models import GmailCredential
from messages import Message, Error
from django.views.decorators.csrf import csrf_exempt
from services.ActionService import ActionService
from services.AuthService import AuthService
from services.SettingService import SettingService
from services.CodeService import CodeService
from services.EntityService import EntityService
from services.TaskService import TaskService
from datetime import datetime, date 
import requests
from urllib.parse import urlparse
import re
from envoy_bu_crm_api.quotation.services.send_mail_service import SendMail
from envoy_bu_crm_api.quotation.services.mail_servise import send_email as gmail_send_email
# from bs4 import BeautifulSoup


@csrf_exempt
@api_view(["POST"])
def create_vendor_response(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid JSON format.", Error.VALIDATION_ERROR)

    rules = {
        "quotation_id": "required|exists:crmq_quotations,id",
        "by_user_id": "required|exists:core_users,id",
        "service_provider_id": "required|exists:core_service_providers,id",
        "coverage_details": "required",
        "coverage_details_type": "required",
        "coverage_details_name": "required",
        "received_date": "required",
        "expiry_date": "required|after:received_date",
        "total_amount": "required",
        "re_request": "required",
    }

    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)

    try:
        # Convert quotation_id to int if it's a string
        quotation_id = int(data["quotation_id"]) if isinstance(data["quotation_id"], str) else data["quotation_id"]
        service_provider_id = int(data["service_provider_id"]) if isinstance(data["service_provider_id"], str) else data["service_provider_id"]
        by_user_id = int(data["by_user_id"]) if isinstance(data["by_user_id"], str) else data["by_user_id"]
        
        # Clean total_amount - remove commas and whitespace
        total_amount = str(data["total_amount"]).replace(",", "").strip() if data.get("total_amount") else ""
        
        # Determine version - use decimal format where each service provider gets unique major version
        # Service Provider A: 1.0, 1.1, 1.2...
        # Service Provider B: 2.0, 2.1, 2.2...
        # Service Provider C: 3.0, 3.1, 3.2...
        
        # Check if this service provider already has versions for this quotation
        latest_version_record = QueryBuilderService("crmq_vendor_response")\
            .select("version")\
            .where("quotation_id", quotation_id)\
            .where("service_provider_id", service_provider_id)\
            .orderBy("version", "desc")\
            .first()

        latest_version_str = latest_version_record.get("version") if latest_version_record and latest_version_record.get("version") is not None else None
        
        # Calculate next decimal version
        if latest_version_str is None:
            # First version for this service provider - need to assign new major version
            # Get all existing versions for this quotation to find the highest major version
            all_versions = QueryBuilderService("crmq_vendor_response")\
                .select("version")\
                .where("quotation_id", quotation_id)\
                .get()
            
            max_major = 0
            for version_record in all_versions:
                version_val = version_record.get("version")
                if version_val:
                    try:
                        # Parse version to get major number
                        if isinstance(version_val, str):
                            if '.' in version_val:
                                major = int(version_val.split('.', 1)[0])
                            else:
                                major = int(version_val)
                        else:
                            # Numeric format
                            version_str = str(version_val)
                            if '.' in version_str:
                                major = int(version_str.split('.', 1)[0])
                            else:
                                major = int(version_str)
                        
                        if major > max_major:
                            max_major = major
                    except (ValueError, AttributeError, TypeError):
                        continue
            
            # Assign next major version (max_major + 1) starting at .0
            new_major = max_major + 1
            version = f"{new_major}.0"
        else:
            # Service provider already has versions - increment minor version
            try:
                if isinstance(latest_version_str, str):
                    # Handle string format like "1.0", "1.1", etc.
                    if '.' in latest_version_str:
                        major, minor = latest_version_str.split('.', 1)
                        major = int(major)
                        minor = int(minor)
                    else:
                        # If it's just "1", treat as "1.0"
                        major = int(latest_version_str)
                        minor = 0
                else:
                    # Handle numeric format - convert to string first to avoid floating point issues
                    version_str = str(latest_version_str)
                    if '.' in version_str:
                        major, minor = version_str.split('.', 1)
                        major = int(major)
                        minor = int(minor)
                    else:
                        # If it's just an integer like 1, treat as "1.0"
                        major = int(version_str)
                        minor = 0
                
                # Increment minor version (1.0 -> 1.1, 1.1 -> 1.2, etc.)
                minor += 1
                version = f"{major}.{minor}"
            except (ValueError, AttributeError, TypeError):
                # If parsing fails, get new major version
                all_versions = QueryBuilderService("crmq_vendor_response")\
                    .select("version")\
                    .where("quotation_id", quotation_id)\
                    .get()
                
                max_major = 0
                for version_record in all_versions:
                    version_val = version_record.get("version")
                    if version_val:
                        try:
                            if isinstance(version_val, str):
                                if '.' in version_val:
                                    major = int(version_val.split('.', 1)[0])
                                else:
                                    major = int(version_val)
                            else:
                                version_str = str(version_val)
                                if '.' in version_str:
                                    major = int(version_str.split('.', 1)[0])
                                else:
                                    major = int(version_str)
                            
                            if major > max_major:
                                max_major = major
                        except (ValueError, AttributeError, TypeError):
                            continue
                
                new_major = max_major + 1
                version = f"{new_major}.0"

        # Debug: Print calculated version to verify it's in correct format
        print(f"[create_vendor_response] Calculated version: {version} (type: {type(version)})")
        print(f"[create_vendor_response] quotation_id: {quotation_id}, service_provider_id: {service_provider_id}")

        # Insert into quotation_service_providers
        vendor_quotation = QueryBuilderService("crmq_quotation_service_providers").insert({
            "quotation_id": quotation_id,
            "service_provider_id": service_provider_id,
            "version": version,
            "opportunity_id": data.get("opportunity_id")
        })
        vendor_quotation_id = vendor_quotation.get("id") if vendor_quotation else None

        if not vendor_quotation_id:
            return ResponseService.response("VALIDATION_ERROR", "Failed to create vendor quotation record.", Error.VALIDATION_ERROR)

        # Predict next response ID
        last_record = QueryBuilderService("crmq_vendor_response")\
            .select("id")\
            .orderBy("id", "desc")\
            .first()
        predicted_id = (last_record.get("id", 0) if last_record and last_record.get("id") else 0) + 1
        new_code = f"QRI-{str(predicted_id).zfill(3)}"

        # Get status from backend instead of UI
        status_data = QueryBuilderService("core_status as status")\
                    .select("status.id AS status_id","status.name AS status_name")\
                    .where("status.type", "quotation_pending")\
                    .first()
        
        if not status_data:
            return ResponseService.response("VALIDATION_ERROR", "Status not found for quotation_pending type.", Error.VALIDATION_ERROR)
        
        status_id = status_data.get('status_id')
        status_name = status_data.get('status_name')

        # Insert vendor response with code
        response_record = QueryBuilderService("crmq_vendor_response").insert({
            "quotation_id": quotation_id,
            "service_provider_id": service_provider_id,
            "by_user_id": by_user_id,
            "coverage_details": data["coverage_details"],
            "coverage_details_type": data["coverage_details_type"],
            "coverage_details_name": data["coverage_details_name"],
            "received_date": data["received_date"],
            "expiry_date": data["expiry_date"],
            "total_amount": total_amount,
            "status": status_name,
            "re_request": 1 if str(data["re_request"]).lower() == "yes" else 0,
            "vendor_quotation_id": vendor_quotation_id,
            "code": new_code,
            "version": version
        })

        if not response_record:
            return ResponseService.response("INTERNAL_SERVER_ERROR", "Failed to create vendor response record.", Error.INTERNAL_SERVER_ERROR)

        # Mark as received
        QueryBuilderService("crmq_quotation_service_providers")\
            .where("quotation_id", quotation_id)\
            .where("service_provider_id", service_provider_id)\
            .update({"is_received": True})

        return ResponseService.response("SUCCESS", {
            "version": version,
            "quotation_id": quotation_id,
            "service_provider_id": service_provider_id,
            "response_record": response_record
        }, Message.DATA_CREATED)
    
    except ValueError as e:
        return ResponseService.response("VALIDATION_ERROR", f"Invalid data format: {str(e)}", Error.VALIDATION_ERROR)
    except Exception as e:
        print(f"Error in create_vendor_response: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return ResponseService.response("INTERNAL_SERVER_ERROR", f"An error occurred: {str(e)}", Error.INTERNAL_SERVER_ERROR)






@api_view(["GET"])
def get_vendor_response_columns(request):
    try:
        # Define the columns to exclude
        exclude_columns = [
            "id", "created_at", "updated_at", "re_request", "code",
            "vendor_quotation_id", "by_user_id", "service_provider_id","quotation_id"
        ]

        # Fetch and filter columns
        columns = QueryBuilderService("crmq_vendor_response").get_table_columns()
        filtered_columns = [col for col in columns if col not in exclude_columns]

        # Format column names to Title Case
        formatted = [{"column": col, "title": col.replace('_', ' ').title()} for col in filtered_columns]
        
        # Add service_provider_name with title "Insurer"
        formatted.append({"column": "service_provider_name", "title": "Insurer"})

        # Document extracted details (from coverage PDF/document extraction)
        document_extracted_columns = [
            {"column": "insurance_agency_details", "title": "Insurance Agency Details"},
            {"column": "customer_details", "title": "Customer Details"},
            {"column": "agent_prepared_by", "title": "Agent / Prepared By"},
            {"column": "policy_information", "title": "Policy Information"},
            {"column": "coverage_amounts", "title": "Coverage Amounts"},
            {"column": "deductibles", "title": "Deductibles"},
            {"column": "discounts_mentioned", "title": "Discounts Mentioned"},
            {"column": "endorsements_addons", "title": "Endorsements / Add-ons"},
            {"column": "premium_cost_details", "title": "Premium / Cost Details"},
        ]
        formatted.extend(document_extracted_columns)

        return ResponseService.response(
            "SUCCESS",
            formatted,
           Message.DATA_FETCHED
        )
    except Exception as e:
        return ResponseService.response("INTERNAL_SERVER_ERROR", str(e), Error.INTERNAL_SERVER_ERROR)


@csrf_exempt
@api_view(["GET"])
def get_all_quotation_form(request, id):
    all_columns = [
        'core_service_providers.name as service_provider_name',
        'core_users.display_name as by_user_name',
        'crmq_vendor_response.*',
        'crmq_quotation_service_providers.version as quotation_version',
        'core_status.name as status_name',
        'core_status.color as status_color',
    ]

    # Get all vendor_quotation entries related to the quotation_id
    vendor_quotation = QueryBuilderService("crmq_quotation_service_providers")\
                        .where("quotation_id", id)\
                        .get()

    vendor_quotation_ids = [row['id'] for row in vendor_quotation]
    if not vendor_quotation_ids:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    # Join with crmq_vendor_response and related tables
    query = QueryBuilderService("crmq_vendor_response")\
        .select(*all_columns)\
        .leftJoin("core_users", "core_users.id", "crmq_vendor_response.by_user_id")\
        .leftJoin("crmq_quotation_service_providers", "crmq_quotation_service_providers.id", "crmq_vendor_response.quotation_id")\
        .leftJoin("core_service_providers", "core_service_providers.id", "crmq_vendor_response.service_provider_id")\
        .leftJoin("core_status", "core_status.id", "crmq_quotation_service_providers.status")\
        .whereIn("crmq_vendor_response.quotation_id", vendor_quotation_ids)\
        .get()

    return ResponseService.response("SUCCESS", query, Message.DATA_FETCHED)

@csrf_exempt
@api_view(["GET", "PUT","DELETE"])
def single_vendor_response(request, vendor_response_id):
    if request.method == "GET":
        return get_single_vendor_response(request, vendor_response_id)
    elif request.method == "DELETE":
        return delete_single_vendor_response(request, vendor_response_id)
    elif request.method == "PUT":
        return update_single_vendor_responce(request, vendor_response_id)



def get_single_vendor_response(request, vendor_response_id):
    # Define the columns to select
    all_columns = [
        'crmq_vendor_response.id as vendor_response_id',
        'crmq_vendor_response.quotation_id',
        'crmq_vendor_response.coverage_details',
        'crmq_vendor_response.coverage_details_type',
        'crmq_vendor_response.coverage_details_name',
        'crmq_vendor_response.received_date',
        'crmq_vendor_response.expiry_date',
        'crmq_vendor_response.total_amount',
        'crmq_vendor_response.status',
        'crmq_vendor_response.re_request',
        'crmq_vendor_response.code',
        'crmq_vendor_response.version',
        'crmq_vendor_response.created_at',
        'crmq_vendor_response.updated_at',
        'crmq_quotation_service_providers.id as vendor_quotation_id',
        'crmq_vendor_response.service_provider_id',
        'core_service_providers.name as service_provider_name',
        'core_users.display_name as by_user_name',
        'crmq_vendor_response.by_user_id',
        'crmq_quotations.opportunity_type_id as opportunity_type_id',
        'crmq_quotations.opportunity_id',
        'crm_opportunities.sales_agent_id',
    ]

    # Build the query
    query = QueryBuilderService("crmq_vendor_response")\
        .select(*all_columns)\
        .leftJoin("crmq_quotation_service_providers", "crmq_quotation_service_providers.id", "crmq_vendor_response.quotation_id")\
        .leftJoin("core_service_providers", "core_service_providers.id", "crmq_vendor_response.service_provider_id")\
        .leftJoin("core_users", "core_users.id", "crmq_vendor_response.by_user_id")\
        .leftJoin("crmq_quotations", "crmq_quotations.id", "crmq_vendor_response.quotation_id")\
        .leftJoin("crm_opportunities", "crm_opportunities.id", "crmq_quotations.opportunity_id")\
        .where("crmq_vendor_response.id", vendor_response_id)\
        .first()

    if not query:
        return ResponseService.response("NOT_FOUND", None, Error.NOT_FOUND)

    # Prepare the response data
    result = {
        "vendor_response_id": query.get("vendor_response_id"),
        "vendor_quotation_id": query.get("vendor_quotation_id"),
        "service_provider_id": query.get("service_provider_id"),
        "service_provider_name": query.get("service_provider_name"),
        "by_user_id": query.get("by_user_id"),
        "by_user_name": query.get("by_user_name"),
        "coverage_details": query.get("coverage_details"),
        "coverage_details_type": query.get("coverage_details_type"),
        "coverage_details_name": query.get("coverage_details_name"),
        "received_date": query.get("received_date"),
        "expiry_date": query.get("expiry_date"),
        "total_amount": query.get("total_amount"),
        "status": query.get("status"),
        "re_request": "yes" if query.get("re_request") else "no",
        "code": query.get("code"),
        "version": query.get("version"),
        "sales_agent_id": query.get("sales_agent_id"),
        "created_at": query.get("created_at"),
        "updated_at": query.get("updated_at"),
        "sales_agent_id": query.get("sales_agent_id"),
    }

    # Process opportunity_type_id to get opportunity_type array
    try:
        opportunity_type_id_str = query.get("opportunity_type_id", "[]")
        if opportunity_type_id_str and opportunity_type_id_str != "[]":
            import json
            opportunity_type_ids = json.loads(opportunity_type_id_str)
            if opportunity_type_ids:
                opportunity_types = QueryBuilderService("crm_opportunity_types")\
                    .whereIn("id", opportunity_type_ids)\
                    .select("id", "title")\
                    .get()
                result["opportunity_type"] = [{"id": ot["id"], "title": ot["title"]} for ot in opportunity_types]
            else:
                result["opportunity_type"] = []
        else:
            result["opportunity_type"] = []
    except Exception:
        result["opportunity_type"] = []

    # Get risks information
    try:
        opportunity_id = query.get("opportunity_id")
        if opportunity_id:
            # Get risk details from crm_risks table through crm_risk_submissions
            risk_columns = [
                'r.id as risk_id',
                'r.code as risk_code',
                'r.created_at as risk_created_at',
                'r.updated_at as risk_updated_at',
                'rt.id as risk_type_id',
                'rt.title as risk_type_title',
                'cust.id as customer_id',
                'cust.name as customer_name',
                'rs.submission_id',
                'rs.version'
            ]
            
            risks = QueryBuilderService("crm_risk_submissions as rs")\
                .leftJoin("crm_risks as r", "r.id", "rs.risk_id")\
                .leftJoin("crm_opportunity_types as rt", "rt.id", "r.risk_type_id")\
                .leftJoin("core_customers as cust", "cust.id", "r.customer_id")\
                .select(*risk_columns)\
                .where("rs.lead_id", opportunity_id)\
                .get()
            
            # Format risks data
            result["risks"] = []
            for risk in risks:
                risk_data = {
                    "risk_id": risk.get("risk_id"),
                    "risk_code": risk.get("risk_code"),
                    "risk_type": {
                        "id": risk.get("risk_type_id"),
                        "title": risk.get("risk_type_title")
                    },
                    "customer": {
                        "id": risk.get("customer_id"),
                        "name": risk.get("customer_name")
                    },
                    "submission_id": risk.get("submission_id"),
                    "version": risk.get("version"),
                    "created_at": risk.get("risk_created_at"),
                    "updated_at": risk.get("risk_updated_at")
                }
                result["risks"].append(risk_data)
        else:
            result["risks"] = []
    except Exception as e:
        result["risks"] = []

    return ResponseService.response("SUCCESS", result, Message.DATA_FETCHED)



def update_single_vendor_responce(request, vendor_response_id):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid JSON format.", Error.VALIDATION_ERROR)

    rules = {
        "by_user_id": "required|exists:core_users,id",
        "coverage_details": "required",
        "coverage_details_type": "required",
        "coverage_details_name": "required",
        "received_date": "required",
        "expiry_date": "required|after:received_date",
        "total_amount": "required",
        "status": "required",
        "re_request": "required",
    }

    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)

    # Find vendor response using vendor_quotation_id
    response_record = QueryBuilderService("crmq_vendor_response")\
        .where("id", vendor_response_id)\
        .first()

    if not response_record:
        return ResponseService.response("VALIDATION_ERROR", "Vendor response not found.", Error.VALIDATION_ERROR)

    # Update the record
    update_fields = {
        "by_user_id": data["by_user_id"],
        "coverage_details": data["coverage_details"],
        "coverage_details_type": data["coverage_details_type"],
        "coverage_details_name": data["coverage_details_name"],
        "received_date": data["received_date"],
        "expiry_date": data["expiry_date"],
        "total_amount": data["total_amount"],
        "status": data["status"],
        "re_request": 1 if data["re_request"] == "yes" else 0
    }

    QueryBuilderService("crmq_vendor_response")\
        .where("id", vendor_response_id)\
        .update(update_fields)

    return ResponseService.response("SUCCESS", {
        "vendor_responce_id": vendor_response_id,
        "updated_fields": update_fields,
        "version": response_record.get("version")
    }, Message.DATA_UPDATED)


def delete_single_vendor_response(request, vendor_response_id):
    # Check if the record exists
    response_record = QueryBuilderService("crmq_vendor_response")\
        .where("id", vendor_response_id)\
        .first()

    if not response_record:
        return ResponseService.response("NOT_FOUND", None, Error.NOT_FOUND)

    # Perform the deletion
    QueryBuilderService("crmq_vendor_response")\
        .where("id", vendor_response_id)\
        .delete()

    return ResponseService.response("SUCCESS", {
        "vendor_response_id": vendor_response_id
    }, Message.DATA_DELETED)


@csrf_exempt
@api_view(["GET"])
def get_form_compare(request, vendor_quotation_ids):
    # Parse the comma-separated vendor_quotation_ids into a list of integers
    vendor_quotation_ids = [int(sid) for sid in vendor_quotation_ids.split(",")]

    # Define the columns to retrieve
    columns = [
        'crmq_vendor_response.*',
        'core_users.display_name as by_user_name',
        'core_service_providers.name as service_provider_name'
    ]

    # Query the vendor responses
    responses = QueryBuilderService("crmq_vendor_response")\
        .select(*columns)\
        .leftJoin("core_users", "core_users.id", "crmq_vendor_response.by_user_id")\
        .leftJoin("crmq_quotation_service_providers", "crmq_quotation_service_providers.id", "crmq_vendor_response.vendor_quotation_id")\
        .leftJoin("core_service_providers", "core_service_providers.id", "crmq_quotation_service_providers.service_provider_id")\
        .whereIn("crmq_vendor_response.vendor_quotation_id", vendor_quotation_ids)\
        .get()

    return ResponseService.response("SUCCESS", responses, Message.DATA_FETCHED)


@csrf_exempt
@api_view(["PUT"])
def shortlist_quotation_form(request, vendor_quotation_id):

    data = request.data

    rules = {
        "is_shortlisted" : "required"
    }

    short_list = data["is_shortlisted"]

    if short_list == "yes":
        value = 1
    else:
        value = 0

    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)


    # Check if the vendor quotation exists
    quotation_service = QueryBuilderService("crmq_quotation_service_providers")\
        .where("id", vendor_quotation_id)\
        .first()


    if not quotation_service:
        return ResponseService.response("VALIDATION_ERROR", "Vendor quotation not found.", Error.VALIDATION_ERROR)

    # Update the record to set is_shortlisted to True
    QueryBuilderService("crmq_quotation_service_providers")\
        .where("id", vendor_quotation_id)\
        .update({
            "is_shortlisted": value
        })

    return ResponseService.response("SUCCESS", {"id": vendor_quotation_id}, Message.DATA_UPDATED)


@csrf_exempt
@api_view(["GET"])
def get_vendor_responses(request, quotation_id):
    filter_type = request.GET.get("filter", "received")  # values: received, shortlisted, all
    response_ids_param = request.GET.get("ids", "")
    sort_by = request.GET.get("sort_by")
    sort_dir = request.GET.get("sort_dir", "desc")
    # Normalize empty values to defaults
    sort_by = "id" if sort_by in [None, ""] else sort_by
    sort_dir = "desc" if sort_dir in [None, ""] else sort_dir
    
    try:
        quotation_id = int(quotation_id)
    except ValueError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid quotation_id format.", Error.VALIDATION_ERROR)

    try:
        response_ids = [int(rid) for rid in response_ids_param.split(",") if rid]
    except ValueError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid ids format.", Error.VALIDATION_ERROR)

    base_query = QueryBuilderService("crmq_quotation_service_providers").where("quotation_id", quotation_id)

    if filter_type == "received":
        base_query = base_query.where("is_received", 1)
    elif filter_type == "shortlisted":
        base_query = base_query.where("is_shortlisted", 1)
    elif filter_type != "all":
        return ResponseService.response("VALIDATION_ERROR", "Invalid filter type.", Error.VALIDATION_ERROR)

    vendor_quotations = base_query.get()
    if not vendor_quotations:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    vendor_quotation_ids = [v["id"] for v in vendor_quotations]
    if not vendor_quotation_ids:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    all_columns = [
        'crmq_vendor_response.*',
        'core_service_providers.name as service_provider_name',
        'core_users.display_name as by_user_name',
        'crmq_quotation_service_providers.service_provider_id',
        'crmq_quotation_service_providers.id as vendor_quotation_id',
        'crmq_quotations.code as quotation_code',
        'crmq_quotations.request_type as quotation_request_type',
        'crmq_quotations.opportunity_type_id as opportunity_type_id',
        'core_status.name as status_name',
        'core_status.color as status_color',
        'core_status.type as status_type',
        'core_customers.id as customer_id',
        'core_customers.name as customer_name',
        'customer_contact.email as customer_email',
        'crm_opportunities.product_id as product_id',
        'crm_opportunities.product_group_id as product_group_id',
    ]

    query = QueryBuilderService("crmq_vendor_response")\
        .select(*all_columns)\
        .leftJoin("crmq_quotation_service_providers", "crmq_quotation_service_providers.id", "crmq_vendor_response.vendor_quotation_id")\
        .leftJoin("core_service_providers", "core_service_providers.id", "crmq_vendor_response.service_provider_id")\
        .leftJoin("core_users", "core_users.id", "crmq_vendor_response.by_user_id")\
        .leftJoin("crmq_quotations", "crmq_quotations.id", "crmq_vendor_response.quotation_id")\
        .leftJoin("crm_opportunities", "crm_opportunities.id", "crmq_quotations.opportunity_id")\
        .leftJoin("core_customers", "core_customers.id", "crm_opportunities.customer_id")\
        .leftJoin("core_contacts as customer_contact", "customer_contact.id", "core_customers.primary_contact_id")\
        .leftJoin("core_status", "core_status.name", "crmq_vendor_response.status")\
        .where("core_status.module", "quotation")\
        .whereIn("crmq_vendor_response.vendor_quotation_id", vendor_quotation_ids)

    # Apply filtering by response_ids if provided
    if response_ids:
        query = query.whereIn("crmq_vendor_response.id", response_ids)

    # Apply sorting and fetch
    query = query.orderBy(sort_by, sort_dir)\
        .get()

    results = query

    for row in results:
        try:
            raw_expiry = row.get("expiry_date")
            # Normalize expiry to date
            if raw_expiry is None:
                expire_date = None
            elif isinstance(raw_expiry, datetime):
                expire_date = raw_expiry.date()
            elif isinstance(raw_expiry, date):
                expire_date = raw_expiry
            else:
                # assume string
                try:
                    expire_date = datetime.strptime(str(raw_expiry), "%Y-%m-%d").date()
                except Exception:
                    expire_date = datetime.strptime(str(raw_expiry), "%Y-%m-%d %H:%M:%S").date()

            today = datetime.now().date()
            delta = (expire_date - today).days if expire_date else None
            row["remaining_days"] = max(delta, 0) if delta is not None else 0
            print(f"[get_vendor_responses] id={row.get('id')} expiry_raw={raw_expiry} parsed={expire_date} today={today} delta={delta} remaining_days={row['remaining_days']}")
            # If expired and not already confirmed/rejected, set to EXPIRED
            if delta is not None and delta < 0:
                stype = row.get("status_type")
                if stype not in ["quotation_confirmed", "quotation_rejected"]:
                    try:
                        QueryBuilderService("crmq_vendor_response").where("id", row["id"]).update({"status": "EXPIRED"})
                        # Fetch proper name/color for EXPIRED
                        status_meta = QueryBuilderService("core_status")\
                            .where("module", "quotation")\
                            .where("name", "EXPIRED")\
                            .first()
                        if status_meta:
                            row["status"] = "EXPIRED"
                            row["status_name"] = status_meta.get("name", "EXPIRED")
                            row["status_color"] = status_meta.get("color")
                            
                        print(f"[get_vendor_responses] status updated to EXPIRED for id={row.get('id')}")
                    except Exception as ue:
                        print(f"[get_vendor_responses] failed to set EXPIRED for id={row.get('id')}: {ue}")
        except Exception as e:
            print(f"[get_vendor_responses] remaining_days calc failed for id={row.get('id')}: {e}")
            row["remaining_days"] = 0
        
        # Process opportunity_type_id to get opportunity_type array
        try:
            opportunity_type_id_str = row.get("opportunity_type_id", "[]")
            if opportunity_type_id_str and opportunity_type_id_str != "[]":
                import json
                opportunity_type_ids = json.loads(opportunity_type_id_str)
                if opportunity_type_ids:
                    opportunity_types = QueryBuilderService("crm_opportunity_types")\
                        .whereIn("id", opportunity_type_ids)\
                        .select("id", "title")\
                        .get()
                    row["opportunity_type"] = [{"id": ot["id"], "title": ot["title"]} for ot in opportunity_types]
                else:
                    row["opportunity_type"] = []
            else:
                row["opportunity_type"] = []
        except Exception:
            row["opportunity_type"] = []
        
        # Get insurer product_id from service provider based on opportunity's native product_id
        try:
            # product_id here is the native product id from core_products
            product_id = row.get("product_id")
            service_provider_id = row.get("service_provider_id")
            
            print(f"[get_vendor_responses] Processing row id={row.get('id')}, product_id={product_id}, service_provider_id={service_provider_id}")
            
            if product_id and service_provider_id:
                # Step 1: From native product (core_products.id) go to mapping table core_products_vendor_products
                # to get all related vendor_product_ids
                mappings = QueryBuilderService("core_product_vendor_products")\
                    .where("product_id", product_id)\
                    .select("vendor_product_id")\
                    .get()

                vendor_product_ids = [m.get("vendor_product_id") for m in mappings if m.get("vendor_product_id")]

                insurer_product = None
                if vendor_product_ids:
                    # Step 2: From those vendor_product_ids, find the one in core_vendor_products
                    # that belongs to this insurer (vendor_id = service_provider_id)
                    insurer_product = QueryBuilderService("core_vendor_products")\
                        .whereIn("id", vendor_product_ids)\
                        .where("vendor_id", service_provider_id)\
                        .select("id", "name")\
                        .first()
                
                if insurer_product:
                    row["insurer_product_id"] = insurer_product.get("id")
                    row["insurer_product_name"] = insurer_product.get("name")
                    print(f"[get_vendor_responses] Found insurer product via mapping: insurer_product_id={row['insurer_product_id']}, name={row['insurer_product_name']}")
                else:
                    row["insurer_product_id"] = None
                    row["insurer_product_name"] = None
                    print(f"[get_vendor_responses] No insurer product mapping found for product_id={product_id}, service_provider_id={service_provider_id}")
            else:
                row["insurer_product_id"] = None
                row["insurer_product_name"] = None
                print(f"[get_vendor_responses] Missing product_id or service_provider_id")
        except Exception as e:
            print(f"[get_vendor_responses] Error getting insurer_product_id for row id={row.get('id')}: {e}")
            import traceback
            print(f"[get_vendor_responses] Traceback: {traceback.format_exc()}")
            row["insurer_product_id"] = None
            row["insurer_product_name"] = None
        
        # Get product_group_name if product_group_id exists
        try:
            product_group_id = row.get("product_group_id")
            
            if product_group_id:
                product_group = QueryBuilderService("core_product_groups")\
                    .where("id", product_group_id)\
                    .select("id", "name")\
                    .first()
                
                if product_group:
                    row["product_group_name"] = product_group.get("name")
                    print(f"[get_vendor_responses] Found product_group: id={product_group_id}, name={row['product_group_name']}")
                else:
                    row["product_group_name"] = None
                    print(f"[get_vendor_responses] Product group not found for id={product_group_id}")
            else:
                row["product_group_name"] = None
        except Exception as e:
            print(f"[get_vendor_responses] Error getting product_group_name for row id={row.get('id')}: {e}")
            row["product_group_name"] = None

        # Extract document data from coverage_details PDF/document so response includes data from the document
        if row.get("coverage_details"):
            coverage_path = row.get("coverage_details")
            coverage_name = row.get("coverage_details_name") or ""
            coverage_type = row.get("coverage_details_type") or ""
            insurer_id = row.get("service_provider_id")
            insurer_name = row.get("service_provider_name")
            try:
                doc_result = extract_document_data_from_coverage_path(
                    coverage_path, coverage_name, coverage_type,
                    insurer_id=insurer_id, insurer_name=insurer_name,
                )
                row["extracted_data"] = doc_result.get("extracted_data", {})
                row["quotation_fields"] = doc_result.get("quotation_fields", {})
                row["document_extracted_details"] = doc_result.get("document_extracted_details", {})
                # When we have extracted quotation fields, optionally overlay on response (so received_date, expiry_date, total_amount reflect document)
                qf = doc_result.get("quotation_fields", {})
                if qf:
                    if qf.get("received_date"):
                        row["received_date"] = qf["received_date"]
                    if qf.get("expiry_date"):
                        row["expiry_date"] = qf["expiry_date"]
                    if qf.get("total_amount") is not None and str(qf.get("total_amount", "")).strip() != "":
                        row["total_amount"] = qf["total_amount"]
            except Exception as ex:
                print(f"[get_vendor_responses] Document extraction failed for id={row.get('id')}: {ex}")
                row["extracted_data"] = {}
                row["quotation_fields"] = {}
                row["document_extracted_details"] = {}
        else:
            row["extracted_data"] = {}
            row["quotation_fields"] = {}
            row["document_extracted_details"] = {}

    return ResponseService.response("SUCCESS", results, Message.DATA_FETCHED)


@csrf_exempt
@api_view(["GET"])
def get_shortlist_quotation_form(request, id):
    selected_ids_param = request.GET.get("selected_id", "")

    try:
        selected_ids = [int(sid) for sid in selected_ids_param.split(",")] if selected_ids_param else []
    except ValueError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid selected_id format.", Error.VALIDATION_ERROR)

    all_columns = [
        'crmq_vendor_response.*',
        'core_service_providers.name as service_provider_name',
        'core_users.display_name as by_user_name',
        'crmq_quotation_service_providers.service_provider_id',
        'crmq_quotation_service_providers.id as vendor_quotation_id',
        'crmq_quotations.code as quotation_code',
        'crmq_quotations.request_type as quotation_request_type',
        'crmq_quotations.opportunity_id as opportunity_id',
        'crm_opportunities.title as opportunity_title',
        'core_status.type as status_type',
    ]

    # Step 1: Get shortlisted vendor quotations
    vendor_quotations = QueryBuilderService("crmq_quotation_service_providers")\
        .where("quotation_id", id)\
        .where("is_received", True)\
        .where("is_shortlisted", True)\
        .get()

    if not vendor_quotations:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    # Optional filtering
    if selected_ids:
        vendor_quotations = [v for v in vendor_quotations if v["service_provider_id"] in selected_ids]

    vendor_quotation_ids = [v["id"] for v in vendor_quotations]

    if not vendor_quotation_ids:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    # Step 2: Fetch responses directly from crmq_vendor_response
    query = QueryBuilderService("crmq_vendor_response")\
        .select(*all_columns)\
        .leftJoin("crmq_quotation_service_providers", "crmq_quotation_service_providers.id", "crmq_vendor_response.vendor_quotation_id")\
        .leftJoin("core_service_providers", "core_service_providers.id", "crmq_vendor_response.service_provider_id")\
        .leftJoin("core_users", "core_users.id", "crmq_vendor_response.by_user_id")\
        .leftJoin("crmq_quotations", "crmq_quotations.id", "crmq_vendor_response.quotation_id")\
        .leftJoin("core_status", "core_status.name", "crmq_vendor_response.status")\
        .leftJoin("crm_opportunities", "crm_opportunities.id", "crmq_quotations.opportunity_id")\
        .whereIn("crmq_vendor_response.vendor_quotation_id", vendor_quotation_ids)\
        .get()

    # Step 3: Group data
    grouped_data = {}
    for row in query:
        vendor_id = row["vendor_quotation_id"]
        if vendor_id not in grouped_data:
            grouped_data[vendor_id] = {
                "id": vendor_id,
                "service_provider_id": row["service_provider_id"],
                "service_provider_name": row["service_provider_name"],
                "by_user_name": row.get("by_user_name"),
                "by_user_id": row["by_user_id"],
                "quotation_code": row["quotation_code"],
                "quotation_request_type": row["quotation_request_type"],
                "coverage_details": row["coverage_details"],
                "coverage_details_type": row["coverage_details_type"],
                "coverage_details_name": row["coverage_details_name"],
                "received_date": row["received_date"],
                "expiry_date": row["expiry_date"],
                "total_amount": row["total_amount"],
                "status": row["status"],
                "re_request": row["re_request"],
                "code": row["code"],
                "version": row["version"],
            }

            try:
                raw_expiry = row.get("expiry_date")
                # Normalize to date
                if raw_expiry is None:
                    expire_date = None
                elif isinstance(raw_expiry, datetime):
                    expire_date = raw_expiry.date()
                elif isinstance(raw_expiry, date):
                    expire_date = raw_expiry
                else:
                    try:
                        expire_date = datetime.strptime(str(raw_expiry), "%Y-%m-%d").date()
                    except Exception:
                        expire_date = datetime.strptime(str(raw_expiry), "%Y-%m-%d %H:%M:%S").date()

                today = datetime.now().date()
                days_remaining = (expire_date - today).days if expire_date else None
                grouped_data[vendor_id]["remaining_days"] = max(days_remaining, 0) if days_remaining is not None else 0
                print(f"[get_shortlist_quotation_form] vendor_q_id={vendor_id} expiry_raw={raw_expiry} parsed={expire_date} today={today} delta={days_remaining} remaining_days={grouped_data[vendor_id]['remaining_days']}")
            except Exception as e:
                print(f"[get_shortlist_quotation_form] remaining_days calc failed vendor_q_id={vendor_id}: {e}")
                grouped_data[vendor_id]["remaining_days"] = 0

            # If expired and not already confirmed/rejected, set to EXPIRED
            try:
                if grouped_data[vendor_id]["remaining_days"] == 0:
                    stype = row.get("status_type")
                    if stype not in ["quotation_confirmed", "quotation_rejected"]:
                        QueryBuilderService("crmq_vendor_response").where("id", row["id"]).update({"status": "EXPIRED"})
                        status_meta = QueryBuilderService("core_status")\
                            .where("module", "quotation")\
                            .where("name", "EXPIRED")\
                            .first()
                        if status_meta:
                            grouped_data[vendor_id]["status"] = "EXPIRED"
                            grouped_data[vendor_id]["status_name"] = status_meta.get("name", "EXPIRED")
                            grouped_data[vendor_id]["status_color"] = status_meta.get("color")
                        print(f"[get_shortlist_quotation_form] status updated to EXPIRED for id={row.get('id')}")
            except Exception as ue:
                print(f"[get_shortlist_quotation_form] failed to set EXPIRED for id={row.get('id')}: {ue}")
            
            # Process opportunity_type_id to get opportunity_type array
            try:
                opportunity_type_id_str = row.get("opportunity_type_id", "[]")
                if opportunity_type_id_str and opportunity_type_id_str != "[]":
                    import json
                    opportunity_type_ids = json.loads(opportunity_type_id_str)
                    if opportunity_type_ids:
                        opportunity_types = QueryBuilderService("crm_opportunity_types")\
                            .whereIn("id", opportunity_type_ids)\
                            .select("id", "title")\
                            .get()
                        grouped_data[vendor_id]["opportunity_type"] = [{"id": ot["id"], "title": ot["title"]} for ot in opportunity_types]
                    else:
                        grouped_data[vendor_id]["opportunity_type"] = []
                else:
                    grouped_data[vendor_id]["opportunity_type"] = []
            except Exception:
                grouped_data[vendor_id]["opportunity_type"] = []

    result = list(grouped_data.values())
    return ResponseService.response("SUCCESS", result, Message.DATA_FETCHED)


@csrf_exempt
@api_view(["PUT"])
def draft_quotation_form(request, id, service_provider_id):
    # Check if the incoming service_provider_id is a single or multiple IDs
    if isinstance(service_provider_id, str) and "," in service_provider_id:
        # Split the IDs if they are comma-separated
        service_provider_ids = [int(sp_id.strip()) for sp_id in service_provider_id.split(",")]
    else:
        # Handle single ID case
        service_provider_ids = [int(service_provider_id)]

    # Check if the quotation service providers exist
    quotation_services = QueryBuilderService("crmq_quotation_service_providers")\
        .where("quotation_id", id)\
        .whereIn("service_provider_id", service_provider_ids)\
        .get()  # Fetch all matching records

    if not quotation_services:
        return ResponseService.response("SUCCESS", [], Error.VALIDATION_ERROR)

    # Update the existing records to set is_draft to True
    updated_rows = QueryBuilderService("crmq_quotation_service_providers")\
        .where("quotation_id", id)\
        .whereIn("service_provider_id", service_provider_ids)\
        .update({"is_draft": True})

    if updated_rows == 0:
        return ResponseService.response("SUCCESS", [], Error.VALIDATION_ERROR)

    return ResponseService.response("SUCCESS", "Draft status updated successfully.", Message.DATA_UPDATED)

@csrf_exempt
@api_view(["PUT"])
def send_quotation_form(request, id, service_provider_id):

    # Check if the incoming service_provider_id is a single or multiple IDs
    if isinstance(service_provider_id, str) and "," in service_provider_id:
        # Split the IDs if they are comma-separated
        service_provider_ids = [int(sp_id.strip()) for sp_id in service_provider_id.split(",")]
    else:
        # Handle single ID case
        service_provider_ids = [int(service_provider_id)]

    # Check if the quotation service provider exists
    quotation_service = QueryBuilderService("crmq_quotation_service_providers")\
        .where("quotation_id", id)\
        .whereIn("service_provider_id", service_provider_ids)\
        .get()

    if not quotation_service:
        return ResponseService.response("SUCCESS", [], Error.VALIDATION_ERROR)

    # Update the existing record to set is_sent to True
    send_data = QueryBuilderService("crmq_quotation_service_providers")\
        .where("quotation_id", id)\
        .whereIn("service_provider_id", service_provider_ids)\
        .update({
            "is_sent": True
        })
    
    if send_data == 0:
        return ResponseService.response("SUCCESS", [], Error.VALIDATION_ERROR)


    return ResponseService.response("SUCCESS", send_data, "default_update_success_msg")


# @csrf_exempt
# @api_view(["GET"])
# def get_draft_generate_document_form(request, id):
#     quotation_id = id

#     drafts = QueryBuilderService("crmq_send_quotations")\
#         .select("id", "version", "date as uploaded_date", "uploaded_by", "opportunity_id", "entity_id", "status", "quotation_request_id", "selected_attributes")\
#         .where("quotation_request_id", quotation_id)\
#         .where("status", "draft").get()

#     if not drafts:
#         return ResponseService.response("SUCCESS", [], "No sent quotations found.")

#     # Get uploaded_by user details
#     user_ids = [draft["uploaded_by"] for draft in drafts if draft.get("uploaded_by")]
#     user_map = {}
#     if user_ids:
#         users = QueryBuilderService("core_users")\
#             .select("id", "display_name as uploaded_by_name")\
#             .whereIn("id", list(set(user_ids))).get()
#         user_map = {user["id"]: user["uploaded_by_name"] for user in users}

#     # Get document details
#     entity_ids = [d["entity_id"] for d in drafts if d.get("entity_id")]
#     doc_map = {}
#     if entity_ids:
#         docs = QueryBuilderService("core_entity_docs")\
#             .select("entity_id", "doc", "name", "type")\
#             .whereIn("entity_id", list(set(entity_ids))).get()
#         for d in docs:
#             doc_map[d["entity_id"]] = {
#                 "doc": d["doc"],
#                 "doc_name": d["name"],
#                 "doc_type": d["type"]
#             }

#     result = []

#     for draft in drafts:
#         send_quotation_id = draft["id"]
#         selected_attributes = json.loads(draft.get("selected_attributes", "[]")) if isinstance(draft.get("selected_attributes"), str) else draft.get("selected_attributes", [])

#         vendor_links = QueryBuilderService("crmq_quotation_vendor_quotations")\
#             .where("send_quotation_id", send_quotation_id).get()
#         linked_vq_ids = [v["vendor_quotation_id"] for v in vendor_links]

#         form_submissions = QueryBuilderService("crmq_quotation_form_submissions")\
#             .select("form_submission_id", "vendor_quotation_id")\
#             .whereIn("vendor_quotation_id", linked_vq_ids if linked_vq_ids else [-1]).get()

#         form_submission_ids = []
#         service_provider_ids = set()
#         service_provider_map = {}
#         vendor_quotation_ids_for_this_draft = set()

#         vendor_quotation_map = {}
#         if linked_vq_ids:
#             vendor_quotations = QueryBuilderService("crmq_quotation_service_providers")\
#                 .whereIn("id", linked_vq_ids).get()
#             vendor_quotation_map = {v["id"]: v["service_provider_id"] for v in vendor_quotations if "id" in v and "service_provider_id" in v}

#         for fs in form_submissions:
#             if "form_submission_id" in fs and "vendor_quotation_id" in fs:
#                 vq_id = fs["vendor_quotation_id"]
#                 sp_id = vendor_quotation_map.get(vq_id)
#                 if sp_id:
#                     form_submission_ids.append(fs["form_submission_id"])
#                     service_provider_map[fs["form_submission_id"]] = sp_id
#                     service_provider_ids.add(sp_id)
#                     vendor_quotation_ids_for_this_draft.add(vq_id)

#         form_values = []
#         if form_submission_ids and selected_attributes:
#             form_values = QueryBuilderService("core_form_submission_values")\
#                 .whereIn("form_submission_id", form_submission_ids)\
#                 .whereIn("attribute_id", selected_attributes).get()

#         value_list = []
#         for value in form_values:
#             sp_id = service_provider_map.get(value["form_submission_id"])
#             if not sp_id:
#                 continue
#             value_list.append({
#                 "form_submission_id": value["form_submission_id"],
#                 "attribute_id": value["attribute_id"],
#                 "value": value["value"],
#                 "service_provider_id": sp_id
#             })

#         result.append({
#             "send_quotation_id": send_quotation_id,
#             "version": draft.get("version"),
#             "uploaded_date": draft.get("uploaded_date"),
#             "uploaded_by": draft.get("uploaded_by"),
#             "uploaded_by_name": user_map.get(draft.get("uploaded_by"), "Unknown"),
#             "opportunity_id": draft.get("opportunity_id"),
#             "entity_id": draft.get("entity_id"),
#             "status": draft.get("status"),
#             "quotation_request_id": draft.get("quotation_request_id"),
#             "form_submission_ids": list(set(form_submission_ids)),
#             "attribute_ids": selected_attributes,
#             "service_provider_ids": list(service_provider_ids),
#             "vendor_quotation_ids": list(vendor_quotation_ids_for_this_draft),
#             "values": value_list,
#             **doc_map.get(draft.get("entity_id"), {})
#         })

#     return ResponseService.response("SUCCESS", result, "default_get_all_success_msg")




# @csrf_exempt
# @api_view(["GET"])
# def get_sent_quotation_forms(request,id):
#     all_columns = [
#         'core_service_providers.name as service_provider_name',
#         'core_form_submission_values.value',
#         'core_form_submission_values.attribute_id',
#         'core_form_submission_values.id',
#         'core_users.display_name as by_user_name',
#         'crmq_quotation_form_submissions.by_user_id',
#         'core_form_attributes.attribute_name',
#         'crmq_quotation_form_submissions.service_provider_id',
#         'core_form_submission_values.form_submission_id'
#     ]

#     # Retrieve all received and shortlisted forms for the given quotation_id
#     forms = QueryBuilderService("crmq_quotation_service_providers")\
#         .where("quotation_id", id)\
#         .where("is_received", True)\
#         .where("is_shortlisted", True)\
#         .where("is_sent", True)\
#         .get()

#     if not forms:
#      return ResponseService.response("SUCCESS", [], "No received forms found for the given quotation ID.")

#     # Extract all service_provider_ids
#     service_provider_ids = [entry['service_provider_id'] for entry in forms]

#     # Retrieve all form submissions for the given quotation_id and service_provider_ids
#     form_submissions = QueryBuilderService("crmq_quotation_form_submissions")\
#         .where("quotation_id", id)\
#         .whereIn("service_provider_id", service_provider_ids)\
#         .get()

#     if not form_submissions:
#         return ResponseService.response("SUCCESS", [], Error.VALIDATION_ERROR)

#     # Extract all form_submission_id values
#     form_submission_ids = [submission['form_submission_id'] for submission in form_submissions]

#     # Query the data
#     query = QueryBuilderService("core_form_submission_values")\
#         .select(*all_columns)\
#         .leftJoin("crmq_quotation_form_submissions", "crmq_quotation_form_submissions.form_submission_id", "core_form_submission_values.form_submission_id")\
#         .leftJoin("core_form_attributes", "core_form_attributes.id", "core_form_submission_values.attribute_id")\
#         .leftJoin("core_users", "core_users.id", "crmq_quotation_form_submissions.by_user_id")\
#         .leftJoin("core_service_providers", "core_service_providers.id", "crmq_quotation_form_submissions.service_provider_id")\
#         .whereIn("core_form_submission_values.form_submission_id", form_submission_ids)\
#         .get()
    
   
    
#     # Group data by service_provider_id
#     grouped_data = {}
#     for row in query:
#         service_provider_id = row["service_provider_id"]
#         if service_provider_id not in grouped_data:
#             grouped_data[service_provider_id] = {
#                 "service_provider_id": service_provider_id,
#                 "service_provider_name": row["service_provider_name"],
#                 "by_user_name": row.get("by_user_name", None),
#                 "by_user_id": row["by_user_id"],  
#                 "form_submission_id" : row["form_submission_id"]

#             }
#         # Add attribute_name as key and value as value
#         grouped_data[service_provider_id][row["attribute_name"]] = row["value"]

#     # Convert grouped data to a list
#     result = list(grouped_data.values())

#     return ResponseService.response("SUCCESS", result, "default_get_all_success_msg")
    


@csrf_exempt
@api_view(["GET"])
def preview_document(request, quotation_id):
    # Step 1: Fetch vendor responses for the given quotation_id
    responses = QueryBuilderService("crmq_vendor_response")\
        .where("quotation_id", quotation_id)\
        .get()

    if not responses:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    # Step 2: Fetch service provider names
    service_provider_ids = [resp["service_provider_id"] for resp in responses]
    service_provider_map = {
        sp["id"]: sp["name"]
        for sp in QueryBuilderService("core_service_providers")\
            .whereIn("id", service_provider_ids)\
            .select("id", "name").get()
    }

    # Step 3: Fetch user names
    user_ids = [resp["by_user_id"] for resp in responses]
    user_map = {
        u["id"]: u["display_name"]
        for u in QueryBuilderService("core_users")\
            .whereIn("id", user_ids)\
            .select("id", "display_name").get()
    }

    # Step 4: Fetch notes using quotation's entity_id
    quotation = QueryBuilderService("crmq_quotations")\
        .where("id", quotation_id)\
        .select("entity_id")\
        .first()

    note = None
    if quotation and quotation.get("entity_id"):
        note_data = QueryBuilderService("core_entity_notes")\
            .where("entity_id", quotation["entity_id"])\
            .select("notes")\
            .first()
        note = note_data["notes"] if note_data else None

    # Step 5: Build the response
    result = []
    for resp in responses:
        result.append({
            "service_provider_id": resp["service_provider_id"],
            "service_provider_name": service_provider_map.get(resp["service_provider_id"], ""),
            "by_user_id": resp["by_user_id"],
            "by_user_name": user_map.get(resp["by_user_id"], ""),
            "vendor_quotation_id": resp["vendor_quotation_id"],
            "code": resp["code"],
            "coverage_details": resp["coverage_details"],
            "coverage_details_name": resp["coverage_details_name"],
            "coverage_details_type": resp["coverage_details_type"],
            "received_date": resp["received_date"],
            "expiry_date": resp["expiry_date"],
            "total_amount": resp["total_amount"],
            "status": resp["status"],
            "re_request": resp["re_request"],
            "version": resp["version"],
            "notes": note
        })

    return ResponseService.response("SUCCESS", result,Message.DATA_FETCHED)



@csrf_exempt
@api_view(["GET"])
def quotation_comments(request, id):
    quotation_id = id

    # Step 1: Get the entity_id from the quotation
    quotation = QueryBuilderService("crmq_quotations")\
        .where("id", quotation_id)\
        .select("entity_id")\
        .first()

    if not quotation or not quotation.get("entity_id"):
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    entity_id = quotation["entity_id"]

    # Step 2: Fetch comments from core_entity_notes
    notes = QueryBuilderService("core_entity_notes")\
        .select("entity_id", "notes")\
        .where("entity_id", entity_id)\
        .get()

    result = [{"comment": note["notes"]} for note in notes]

    return ResponseService.response("SUCCESS", result, Message.DATA_FETCHED)



# @csrf_exempt
# @api_view(["GET", "PUT"])
# def manage_generate_document(request, id, form_submission_id):
#     if request.method == "GET":
#         # Call the get_single_form_generate function for GET requests
#         return get_single_form_generate(request, id, form_submission_id)
#     elif request.method == "PUT":
#         # Call the update_generate_document function for PUT requests
#         return update_generate_document(request, id, form_submission_id)
    

@csrf_exempt
@api_view(["GET"])
def get_single_form_generate(request, quotation_id, vendor_quotation_id):
    response = QueryBuilderService("crmq_vendor_response")\
        .where("vendor_quotation_id", vendor_quotation_id)\
        .where("quotation_id", quotation_id)\
        .first()

    if not response:
        return ResponseService.response("SUCCESS", {}, Message.DATA_NOT_FOUND)

    service_provider = QueryBuilderService("core_service_providers")\
        .where("id", response["service_provider_id"]).first()
    user = QueryBuilderService("core_users")\
        .where("id", response["by_user_id"]).first()

    # Get note using entity_id directly from vendor response
    note = None
    entity_id = response.get("entity_id")
    if entity_id:
        note_record = QueryBuilderService("core_entity_notes")\
            .where("entity_id", entity_id).first()
        note = note_record["notes"] if note_record else None

    fixed_fields = [
        "coverage_details",
        "coverage_details_type",
        "coverage_details_name",
        "received_date",
        "expiry_date",
        "total_amount",
        "status",
        "re_request"
    ]

    response_data = {
        "vendor_quotation_id": vendor_quotation_id,
        "service_provider_name": service_provider["name"] if service_provider else None,
        "by_user_id": response["by_user_id"],
        "by_user_name": user["display_name"] if user else None,
        "comment": note,
    }

    for field in fixed_fields:
        response_data[field] = response.get(field)

    return ResponseService.response("SUCCESS", response_data, Message.DATA_FETCHED)


@csrf_exempt
@api_view(["GET"])
def get_all_form_generate(request, quotation_id):
    responses = QueryBuilderService("crmq_vendor_response")\
        .where("quotation_id", quotation_id)\
        .get()

    if not responses:
        return ResponseService.response("SUCCESS", {}, Message.DATA_NOT_FOUND)

    service_provider_ids = [r["service_provider_id"] for r in responses]
    user_ids = [r["by_user_id"] for r in responses]
    entity_ids = [r["entity_id"] for r in responses if r.get("entity_id")]

    sp_map = {
        sp["id"]: sp["name"]
        for sp in QueryBuilderService("core_service_providers")\
            .whereIn("id", service_provider_ids)\
            .select("id", "name").get()
    }

    user_map = {
        u["id"]: u["display_name"]
        for u in QueryBuilderService("core_users")\
            .whereIn("id", user_ids)\
            .select("id", "display_name").get()
    }

    notes = QueryBuilderService("core_entity_notes")\
        .whereIn("entity_id", entity_ids)\
        .select("entity_id", "notes").get()
    notes_map = {n["entity_id"]: n["notes"] for n in notes}

    fixed_fields = [
        "coverage_details",
        "coverage_details_type",
        "coverage_details_name",
        "received_date",
        "expiry_date",
        "total_amount",
        "status",
        "re_request"
    ]

    form_data = []

    for r in responses:
        entity_id = r.get("entity_id")
        row = {
            "vendor_quotation_id": r["vendor_quotation_id"],
            "service_provider_id": r["service_provider_id"],
            "service_provider_name": sp_map.get(r["service_provider_id"]),
            "by_user_id": r["by_user_id"],
            "by_user_name": user_map.get(r["by_user_id"]),
            "comment": notes_map.get(entity_id)
        }

        for field in fixed_fields:
            row[field] = r.get(field)

        form_data.append(row)

    return ResponseService.response("SUCCESS", {
        "form_data": form_data
    }, Message.DATA_FETCHED)




@csrf_exempt
@api_view(["GET"])
def preview_data(request, quotation_id):
    # Extract and validate service_provider_ids from query param
    service_provider_ids_param = request.GET.get("selected_id", "")
    if not service_provider_ids_param:
        return ResponseService.response("VALIDATION_ERROR", "selected_id is required.", Error.VALIDATION_ERROR)

    try:
        service_provider_ids = [int(sid.strip()) for sid in service_provider_ids_param.split(",")]
    except ValueError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid service_provider_ids format.", Error.VALIDATION_ERROR)

    # Columns to select including JOIN fields
    all_columns = [
        "crmq_vendor_response.*",
        "core_service_providers.name as service_provider_name",
        "core_users.display_name as by_user_name"
    ]

    # Query with LEFT JOINs
    responses = QueryBuilderService("crmq_vendor_response")\
        .select(*all_columns)\
        .leftJoin("core_service_providers", "core_service_providers.id", "crmq_vendor_response.service_provider_id")\
        .leftJoin("core_users", "core_users.id", "crmq_vendor_response.by_user_id")\
        .where("crmq_vendor_response.quotation_id", quotation_id)\
        .whereIn("crmq_vendor_response.id", service_provider_ids)\
        .get()

    if not responses:
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    # Format response
    return ResponseService.response("SUCCESS", responses,Message.DATA_FETCHED)


@csrf_exempt
@api_view(["POST", "GET"])
def generate_document(request, quotation_id):
    if request.method == "POST":
        return generate_single_document(request, quotation_id)
    elif request.method == "GET":
        return get_generate_document_forms(request, quotation_id)




def generate_single_document(request, quotation_id):
    try:
        data = request.data
    except Exception as e:
        return ResponseService.response("VALIDATION_ERROR", str(e), Error.VALIDATION_ERROR)

    # Updated Validation Rules
    rules = {
        "vendor_response_ids": "required|array|min:1",
        "vendor_response_ids.*": "exists:crmq_vendor_response,id",
        "columns": "required|array|min:1",
        "expiry_date":"required"
    }

    custom_messages = {
        "vendor_response_ids.required": "Vendor response IDs are required.",
        "vendor_response_ids.array": "Vendor response IDs must be provided as an array.",
        "vendor_response_ids.min": "At least one vendor response ID is required.",
        "vendor_response_ids.*.exists": "Some vendor response IDs are invalid or do not exist.",
        "columns.required": "Columns are required to generate document.",
        "columns.array": "Columns must be in array format.",
        "columns.min": "At least one column must be provided."
    }

    errors = ValidatorService.validate(data, rules, custom_messages)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)
    try:
        datetime.strptime(data.get("expiry_date"), "%Y-%m-%d")  
    except ValueError:
        return ResponseService.response(
            "VALIDATION_ERROR",
            {"expiry_date": ["Invalid date format. Expected YYYY-MM-DD."]},
            Error.VALIDATION_ERROR
        )

    vendor_response_ids = data["vendor_response_ids"]
    columns = data["columns"]
    comment = data.get("comment")
    is_sent = data.get("is_sent", False)
    is_draft = data.get("is_draft", False)
    version = data.get("version")
    date = data.get("date", datetime.now().date())
    user = request.user if request.user.is_authenticated else None

    # Get vendor_quotation_ids from vendor_response_ids
    responses = QueryBuilderService("crmq_vendor_response").whereIn("id", vendor_response_ids).get()
    vendor_quotation_ids = list({r["vendor_quotation_id"] for r in responses if r.get("vendor_quotation_id")})

    if not vendor_quotation_ids:
        return ResponseService.response("VALIDATION_ERROR", "Could not resolve any vendor quotation IDs from response IDs.", Error.VALIDATION_ERROR)

    # Always auto-calculate the next version to ensure consistency
    latest = QueryBuilderService("crmq_send_quotations")\
        .where("quotation_request_id", quotation_id)\
        .orderBy("version", "desc").first()
    version = (latest["version"] + 1) if latest and latest.get("version") else 1

    entity = QueryBuilderService("core_entities").insert({
        "type": "Generate Document",
        "created_at": datetime.now(),
        "created_by_id": user.id if user else None
    })
    if not entity or "id" not in entity:
        return ResponseService.response("INTERNAL_SERVER_ERROR", "Failed to create core entity.", Error.INTERNAL_SERVER_ERROR)
    entity_id = entity["id"]

    send_quotation = QueryBuilderService("crmq_send_quotations").insert({
        "opportunity_id": data.get("opportunity_id"),
        "entity_id": entity_id,
        "status": "sent" if is_sent else "draft",
        "version": version,
        "date": date,
        "uploaded_by": user.id if user else None,
        "quotation_request_id": quotation_id,
        "selected_columns": json.dumps(columns),
        "expiry_date": data["expiry_date"]
    })

    if not send_quotation or "id" not in send_quotation:
        return ResponseService.response("INTERNAL_SERVER_ERROR", "Failed to create send quotation.", Error.INTERNAL_SERVER_ERROR)
    send_quotation_id = send_quotation["id"]

    for vqid in vendor_quotation_ids:
        QueryBuilderService("crmq_quotation_vendor_quotations").insert({
            "send_quotation_id": send_quotation_id,
            "vendor_quotation_id": vqid
        })

    if comment:
        QueryBuilderService("core_entity_notes").insert({
            "is_high_priority": False,
            "notes": comment,
            "entity_id": entity_id,
            "added_at": datetime.now(),
        })

    if data.get("doc") and data.get("doc_name") and data.get("doc_type"):
        QueryBuilderService("core_entity_docs").insert({
            "entity_id": entity_id,
            "doc": data["doc"],
            "name": data["doc_name"],
            "type": data["doc_type"]
        })

    QueryBuilderService("crmq_quotation_service_providers")\
        .where("quotation_id", quotation_id)\
        .whereIn("id", vendor_quotation_ids)\
        .update({"is_sent": is_sent, "is_draft": is_draft})

    user_display = QueryBuilderService("core_users").where("id", user.id).first() if user else None
    service_provider_ids = set()
    values = []

    for vendor_id in vendor_quotation_ids:
        vq_info = QueryBuilderService("crmq_quotation_service_providers").where("id", vendor_id).first()
        if not vq_info:
            continue

        sp_id = vq_info["service_provider_id"]
        service_provider_ids.add(sp_id)

        responses = QueryBuilderService("crmq_vendor_response")\
            .where("vendor_quotation_id", vendor_id)\
            .whereIn("id", vendor_response_ids).get()

        for resp in responses:
            entry = {
                "quotation_id": resp.get("quotation_id"),
                "service_provider_id": sp_id
            }
            for col in columns:
                entry[col] = resp.get(col)
            values.append(entry)

    customer = (
        QueryBuilderService("crmq_quotations")
        .select("crmq_quotations.customer_id", "core_customers.name as display_name")
        .leftJoin("core_customers", "core_customers.id", "crmq_quotations.customer_id")
        .where("crmq_quotations.id", quotation_id)
        .first()
    )

    return ResponseService.response("SUCCESS", {
        "document_id": send_quotation_id,
        "version": version,
        "uploaded_date": str(date),
        "uploaded_by": user.id if user else None,
        "uploaded_by_name": user_display["display_name"] if user_display else "Unknown",
        "opportunity_id": data.get("opportunity_id"),
        "entity_id": entity_id,
        "status": "sent" if is_sent else "draft",
        "quotation_request_id": quotation_id,
        "service_provider_ids": list(service_provider_ids),
        "vendor_quotation_ids": vendor_quotation_ids,
        "values": values,
        "comment": comment,
        "customer": customer
    }, Message.DATA_CREATED)





def get_generate_document_forms(request, quotation_id):
    """
    Retrieves generated documents (draft or sent) for a given quotation.
    URL: /api/quotations/{quotation_id}/generate-documents?status=draft|sent
    """
    # Determine desired status
    status = request.GET.get("status", "draft").lower()
    if status not in ["draft", "sent"]:
        return ResponseService.response(
            "VALIDATION_ERROR",
            "Invalid status value; must be 'draft' or 'sent'.",
            Error.VALIDATION_ERROR
        )

    # Sorting (defaults to desc by id)
    sort_by = request.GET.get("sort_by")
    sort_dir = request.GET.get("sort_dir", "desc")
    # Normalize empty values to defaults
    sort_by = "id" if sort_by in [None, ""] else sort_by
    sort_dir = "desc" if sort_dir in [None, ""] else sort_dir

    # Fetch send_quotations with matching status
    drafts = QueryBuilderService("crmq_send_quotations")\
        .select(
            "id", "version", "date as uploaded_date", "uploaded_by",
            "opportunity_id", "entity_id", "status",
            "quotation_request_id", "generated_pdf","expiry_date"
        )\
        .where("quotation_request_id", quotation_id)\
        .where("status", status)\
        .orderBy(sort_by, sort_dir)\
        .get()

    if not drafts:
        msg = f"No {status} quotations found."
        return ResponseService.response("SUCCESS", [], Message.DATA_NOT_FOUND)

    # Map user IDs to display names
    user_ids = {d["uploaded_by"] for d in drafts if d.get("uploaded_by")}
    user_map = {}
    if user_ids:
        users = QueryBuilderService("core_users")\
            .select("id", "display_name as uploaded_by_name")\
            .whereIn("id", list(user_ids)).get()
        user_map = {u["id"]: u["uploaded_by_name"] for u in users}

    # Map entity IDs to uploaded docs
    entity_ids = {d["entity_id"] for d in drafts if d.get("entity_id")}
    doc_map = {}
    if entity_ids:
        docs = QueryBuilderService("core_entity_docs")\
            .select("entity_id", "doc", "name", "type")\
            .whereIn("entity_id", list(entity_ids)).get()
        for d in docs:
            doc_map[d["entity_id"]] = {
                "coverage_details": d["doc"],
                "coverage_details_name": d["name"],
                "coverage_details_type": d["type"]
            }

    # Get customer info from main quotation
    customer_obj = {}
    main_quote = QueryBuilderService("crmq_quotations")\
        .select("id", "customer_id")\
        .where("id", quotation_id).first()
    if main_quote and main_quote.get("customer_id"):
        cust = QueryBuilderService("core_customers")\
            .select("id", "name")\
            .where("id", main_quote["customer_id"]).first()
        if cust:
            customer_obj = {
                "customer": {"id": cust["id"], "name": cust["name"]}
            }

    result = []
    for draft in drafts:
        send_id = draft["id"]
        vendor_links = QueryBuilderService("crmq_quotation_vendor_quotations")\
            .where("send_quotation_id", send_id).get()
        linked_ids = [v["vendor_quotation_id"] for v in vendor_links]

        values_list, service_provider_ids = [], set()
        if linked_ids:
            vq_map = {
                v["id"]: v["service_provider_id"]
                for v in QueryBuilderService("crmq_quotation_service_providers")\
                    .whereIn("id", linked_ids).get()
            }

            for vqid in linked_ids:
                sp = vq_map.get(vqid)
                if not sp:
                    continue
                service_provider_ids.add(sp)

                responses = QueryBuilderService("crmq_vendor_response")\
                    .where("vendor_quotation_id", vqid).get()
                for resp in responses:
                    values_list.append({
                        "quotation_id": resp.get("quotation_id"),
                        "response_value": resp.get("response_value"),
                        "service_provider_id": sp
                    })

        # Collect document links
        docs = []
        ent = draft.get("entity_id")
        if ent in doc_map:
            docs.append(doc_map[ent])
        # Parse generated PDF JSON
        genpdf = draft.get("generated_pdf")
        if genpdf:
            try:
                pdf_data = json.loads(genpdf)
                docs.append({
                    "coverage_details": pdf_data.get("link"),
                    "coverage_details_name": pdf_data.get("name"),
                    "coverage_details_type": pdf_data.get("type")
                })
            except Exception:
                pass

        result.append({
            "send_quotation_id": send_id,
            "version": draft.get("version"),
            "uploaded_date": draft.get("uploaded_date"),
            "uploaded_by": draft.get("uploaded_by"),
            "uploaded_by_name": user_map.get(draft.get("uploaded_by"), "Unknown"),
            "opportunity_id": draft.get("opportunity_id"),
            "entity_id": draft.get("entity_id"),
            "status": draft.get("status"),
            "quotation_request_id": draft.get("quotation_request_id"),
            "service_provider_ids": list(service_provider_ids),
            "vendor_quotation_ids": list(linked_ids),
            "values": values_list,
            "documents": docs,
            "expiry_date": draft.get("expiry_date"),
            **customer_obj
        })

    return ResponseService.response("SUCCESS", result, Message.DATA_FETCHED)



@csrf_exempt
@api_view(["PUT"])
def update_generate_document(request, doc_id):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return ResponseService.response("VALIDATION_ERROR", "Invalid JSON format.", Error.VALIDATION_ERROR)

    # Validate required fields
    rules = {
        'vendor_quotation_ids': 'required|array|min:1',
        'columns': 'required|array|min:1'
    }
    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)

    vendor_quotation_ids = data['vendor_quotation_ids']
    columns = data['columns']
    comment = data.get("comment")
    is_sent = data.get('is_sent', False)
    is_draft = data.get('is_draft', False)

    # Update send quotation record
    QueryBuilderService("crmq_send_quotations")\
        .where("id", doc_id)\
        .update({
            "status": "sent" if is_sent else "draft",
            "selected_attributes": json.dumps(columns)
        })

    # Remove old vendor links and insert new ones
    QueryBuilderService("crmq_quotation_vendor_quotations")\
        .where("send_quotation_id", doc_id).delete()
    for vqid in vendor_quotation_ids:
        QueryBuilderService("crmq_quotation_vendor_quotations").insert({
            "send_quotation_id": doc_id,
            "vendor_quotation_id": vqid
        })

    # Fetch related record info
    send_record = QueryBuilderService("crmq_send_quotations")\
        .where("id", doc_id).select("entity_id", "quotation_request_id", "opportunity_id", "date", "uploaded_by", "version").first()
    entity_id = send_record["entity_id"] if send_record else None

    # Handle notes and documents
    if entity_id:
        if comment:
            note_record = QueryBuilderService("core_entity_notes").where("entity_id", entity_id).first()
            if note_record:
                QueryBuilderService("core_entity_notes").where("id", note_record["id"]).update({"notes": comment})
            else:
                QueryBuilderService("core_entity_notes").insert({
                    "is_high_priority": False,
                    "notes": comment,
                    "entity_id": entity_id,
                    "added_at": datetime.now()
                })

        if data.get("doc") and data.get("doc_name") and data.get("doc_type"):
            existing_doc = QueryBuilderService("core_entity_docs").where("entity_id", entity_id).first()
            doc_data = {
                "doc": data["doc"],
                "name": data["doc_name"],
                "type": data["doc_type"]
            }
            if existing_doc:
                QueryBuilderService("core_entity_docs").where("entity_id", entity_id).update(doc_data)
            else:
                doc_data["entity_id"] = entity_id
                QueryBuilderService("core_entity_docs").insert(doc_data)

    # Update vendor quotation status
    quotation_id_record = QueryBuilderService("crmq_quotation_service_providers")\
        .whereIn("id", vendor_quotation_ids).select("quotation_id").first()
    if quotation_id_record:
        quotation_id = quotation_id_record["quotation_id"]
        QueryBuilderService("crmq_quotation_service_providers")\
            .where("quotation_id", quotation_id)\
            .whereIn("id", vendor_quotation_ids)\
            .update({"is_sent": is_sent, "is_draft": is_draft})

    # Build values based on selected columns
    values = []
    service_provider_ids = set()
    for vendor_id in vendor_quotation_ids:
        vq_info = QueryBuilderService("crmq_quotation_service_providers").where("id", vendor_id).first()
        vendor_resp = QueryBuilderService("crmq_vendor_response").where("vendor_quotation_id", vendor_id).first()
        if vq_info and vendor_resp:
            sp_id = vq_info["service_provider_id"]
            service_provider_ids.add(sp_id)
            row = {
                "vendor_quotation_id": vendor_id,
                "service_provider_id": sp_id
            }
            for col in columns:
                row[col] = vendor_resp.get(col)
            values.append(row)

    # Final response
    user_display = QueryBuilderService("core_users").where("id", send_record.get("uploaded_by")).first() if send_record.get("uploaded_by") else None
    return ResponseService.response("SUCCESS", {
        "send_quotation_id": doc_id,
        "version": send_record.get("version"),
        "uploaded_date": str(send_record.get("date")) if send_record.get("date") else None,
        "uploaded_by": send_record.get("uploaded_by"),
        "uploaded_by_name": user_display["display_name"] if user_display else "Unknown",
        "opportunity_id": send_record.get("opportunity_id"),
        "entity_id": send_record.get("entity_id"),
        "status": "sent" if is_sent else "draft",
        "quotation_request_id": send_record.get("quotation_request_id"),
        "columns": columns,
        "service_provider_ids": list(service_provider_ids),
        "vendor_quotation_ids": vendor_quotation_ids,
        "values": values,
        "comment": comment
    }, Message.DATA_UPDATED)


@csrf_exempt
@api_view(["POST"])
def upload_docs(request):  
   
    try:
        data = request.data
    except Exception as e:
        return ResponseService.response("VALIDATION_ERROR", str(e), Error.VALIDATION_ERROR)
    
    rules = {
        "quotation_request_id" : "required",
        "doc_link" : "required",
        "doc_type" : "required",
        "doc_name" : "required",
        "uploaded_by" : "required",
        # "expiry_date": "required"
      
    }          

    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)
    # try:
    #     datetime.strptime(data.get("expiry_date"), "%Y-%m-%d")  
    # except ValueError:
    #     return ResponseService.response(
    #         "VALIDATION_ERROR",
    #         {"expiry_date": ["Invalid date format. Expected YYYY-MM-DD."]},
    #         Error.VALIDATION_ERROR
    #     )

    user = request.user if request.user.is_authenticated else None

    entity = QueryBuilderService("core_entities").insert({
        "type": "Generate Document",
        "created_at": datetime.now(),
        "created_by_id": user.id if user else None
    })

    if not entity or "id" not in entity:
        return ResponseService.response("INTERNAL_SERVER_ERROR", "Failed to create core entity.", Error.INTERNAL_SERVER_ERROR)
    entity_id = entity["id"]

    latest = QueryBuilderService("crmq_send_quotations")\
            .where("quotation_request_id", data.get('quotation_request_id'))\
            .orderBy("version", "desc").first()
    version = (latest["version"] + 1) if latest and latest.get("version") else 1


    send_quotation = QueryBuilderService("crmq_send_quotations").insert({
        "opportunity_id": data.get("opportunity_id"),
        "entity_id": entity_id,
        "status": "draft",
        "version": version,
        "date": datetime.now().date(),
        "uploaded_by": user.id if user else None,
        "quotation_request_id": data.get('quotation_request_id'),
        # "expiry_date": data["expiry_date"]
        })

    if not send_quotation or "id" not in send_quotation:
        return ResponseService.response("INTERNAL_SERVER_ERROR", "Failed to create send quotation.", Error.INTERNAL_SERVER_ERROR)
    
    if data.get("doc_link") and data.get("doc_name") and data.get("doc_type"):
        QueryBuilderService("core_entity_docs").insert({
            "entity_id": entity_id,
            "doc": data["doc_link"],
            "name": data["doc_name"],
            "type": data["doc_type"],
        })

    return ResponseService.response("SUCCESS", [send_quotation,data],Message.DATA_CREATED)


# @csrf_exempt
# @api_view(["GET"])
# def generate_doc_single_view(request, id):
#     # Step 1: Get send_quotation record
#     draft = QueryBuilderService("crmq_send_quotations")\
#         .select("id", "version", "date as uploaded_date", "uploaded_by", "opportunity_id", "entity_id",
#                 "status", "quotation_request_id", "selected_attributes")\
#         .where("id", id).first()

#     if not draft:
#         return ResponseService.response("SUCCESS", {}, "Quotation document not found.")

#     selected_attributes = json.loads(draft.get("selected_attributes", "[]")) if isinstance(draft.get("selected_attributes"), str) else draft.get("selected_attributes", [])

#     # Step 2: Get uploaded_by user
#     user_map = {}
#     if draft.get("uploaded_by"):
#         user = QueryBuilderService("core_users")\
#             .select("id", "display_name as uploaded_by_name")\
#             .where("id", draft["uploaded_by"]).first()
#         if user:
#             user_map[draft["uploaded_by"]] = user["uploaded_by_name"]

#     # Step 3: Get doc
#     doc = QueryBuilderService("core_entity_docs")\
#         .select("entity_id", "doc", "name", "type")\
#         .where("entity_id", draft["entity_id"]).first()

#     doc_info = {
#         "doc": doc["doc"],
#         "doc_name": doc["name"],
#         "doc_type": doc["type"]
#     } if doc else {}

#     # Step 4: Vendor links
#     vendor_links = QueryBuilderService("crmq_quotation_vendor_quotations")\
#         .where("send_quotation_id", id).get()
#     linked_vq_ids = [v["vendor_quotation_id"] for v in vendor_links]

#     # Fetch vendor quotations (id + version)
#     vendor_quotation_objs = []
#     if linked_vq_ids:
#         vendor_quotation_records = QueryBuilderService("crmq_quotation_service_providers")\
#             .select("id", "version")\
#             .whereIn("id", linked_vq_ids).get()
#         vendor_quotation_objs = [{"id": v["id"], "code": v.get("version", 0)} for v in vendor_quotation_records]

#     # Fetch form submissions
#     form_submissions = QueryBuilderService("crmq_quotation_form_submissions")\
#         .select("form_submission_id", "vendor_quotation_id")\
#         .whereIn("vendor_quotation_id", linked_vq_ids if linked_vq_ids else [-1]).get()

#     form_submission_ids = []
#     service_provider_ids = set()
#     service_provider_map = {}

#     if form_submissions:
#         vendor_quotations_map = QueryBuilderService("crmq_quotation_service_providers")\
#             .select("id", "service_provider_id")\
#             .whereIn("id", linked_vq_ids).get()
#         vendor_quotation_map = {v["id"]: v["service_provider_id"] for v in vendor_quotations_map if "id" in v and "service_provider_id" in v}

#         for fs in form_submissions:
#             if "form_submission_id" in fs and "vendor_quotation_id" in fs:
#                 vq_id = fs["vendor_quotation_id"]
#                 sp_id = vendor_quotation_map.get(vq_id)
#                 if sp_id:
#                     form_submission_ids.append(fs["form_submission_id"])
#                     service_provider_map[fs["form_submission_id"]] = sp_id
#                     service_provider_ids.add(sp_id)

#     # Step 5: Attributes + Form values
#     attribute_objs = []
#     form_values = []
#     value_list = []

#     if selected_attributes:
#         attribute_records = QueryBuilderService("core_form_attributes")\
#             .select("id", "attribute_name", "title")\
#             .whereIn("id", selected_attributes).get()
#         attribute_objs = [{"id": a["id"], "attribute_name": a.get("attribute_name"), "title": a.get("title")} for a in attribute_records]

#     if form_submission_ids and selected_attributes:
#         form_values = QueryBuilderService("core_form_submission_values")\
#             .whereIn("form_submission_id", form_submission_ids)\
#             .whereIn("attribute_id", selected_attributes).get()

#     for value in form_values:
#         sp_id = service_provider_map.get(value["form_submission_id"])
#         if not sp_id:
#             continue
#         value_list.append({
#             "form_submission_id": value["form_submission_id"],
#             "attribute_id": value["attribute_id"],
#             "value": value["value"],
#             "service_provider_id": sp_id
#         })

#     # Final result
#     result = {
#         "send_quotation_id": draft["id"],
#         "version": draft.get("version"),
#         "uploaded_date": draft.get("uploaded_date"),
#         "uploaded_by": draft.get("uploaded_by"),
#         "uploaded_by_name": user_map.get(draft.get("uploaded_by"), "Unknown"),
#         "opportunity_id": draft.get("opportunity_id"),
#         "entity_id": draft.get("entity_id"),
#         "status": draft.get("status"),
#         "quotation_request_id": draft.get("quotation_request_id"),
#         "form_submission_ids": list(set(form_submission_ids)),
#         "attribute_ids": attribute_objs,
#         "service_provider_ids": list(service_provider_ids),
#         "vendor_quotation_ids": vendor_quotation_objs,
#         "values": value_list,
#         **doc_info
#     }

#     return ResponseService.response("SUCCESS", result, "default_get_success_msg")


@csrf_exempt
@api_view(["GET"])
def generate_doc_single_view(request, id):
    # Step 1: Get send_quotation record
    draft = QueryBuilderService("crmq_send_quotations")\
        .select("id", "version", "date as uploaded_date", "uploaded_by", "opportunity_id", "entity_id",
                "status", "quotation_request_id", "selected_columns")\
        .where("id", id).first()

    if not draft:
        return ResponseService.response("SUCCESS", {}, Message.DATA_NOT_FOUND)

    # Step 2: Get uploaded_by user
    uploaded_by_name = "Unknown"
    if draft.get("uploaded_by"):
        user = QueryBuilderService("core_users")\
            .select("id", "display_name as uploaded_by_name")\
            .where("id", draft["uploaded_by"]).first()
        if user:
            uploaded_by_name = user.get("uploaded_by_name", "Unknown")

    # Step 3: Get document
    doc = QueryBuilderService("core_entity_docs")\
        .select("entity_id", "doc", "name", "type")\
        .where("entity_id", draft["entity_id"]).first()

    doc_info = {
        "doc": doc["doc"],
        "doc_name": doc["name"],
        "doc_type": doc["type"]
    } if doc else {}

    # Step 4: Get comment (from core_entity_notes)
    comment_text = None
    if draft.get("entity_id"):
        note = QueryBuilderService("core_entity_notes")\
            .select("id", "notes")\
            .where("entity_id", draft["entity_id"]).first()
        if note:
            comment_text = note.get("notes")

    # Step 5: Vendor Quotation links
    vendor_links = QueryBuilderService("crmq_quotation_vendor_quotations")\
        .where("send_quotation_id", id).get()
    linked_vq_ids = [v["vendor_quotation_id"] for v in vendor_links]

    # Fetch vendor quotations basic details
    vendor_response_objs = []
    service_provider_ids = set()
    values = []

    for vq_id in linked_vq_ids:
        vq_info = QueryBuilderService("crmq_quotation_service_providers")\
            .select("id", "service_provider_id")\
            .where("id", vq_id).first()
        if not vq_info:
            continue

        sp_id = vq_info["service_provider_id"]
        service_provider_ids.add(sp_id)

        

        responses = QueryBuilderService("crmq_vendor_response")\
            .where("vendor_quotation_id", vq_id).get()
        
        

        for resp in responses:

            vendor_response_objs.append({
            "id": resp.get("id"),
            "code": resp.get("code", ""),
            
            })
            
            values.append({
                "vendor_quotation_id": vq_id,
                "quotation_id": resp.get("quotation_id"),
                "response_value": resp.get("response_value"),
                "service_provider_id": sp_id
            })

    # Step 6: Customer info
    customer_obj = {}
    if draft.get("quotation_request_id"):
        quotation_request = QueryBuilderService("crmq_quotations")\
            .select("id", "customer_id")\
            .where("id", draft["quotation_request_id"]).first()
        if quotation_request and quotation_request.get("customer_id"):
            customer = QueryBuilderService("core_customers")\
                .select("id", "name")\
                .where("id", quotation_request["customer_id"]).first()
            if customer:
                customer_obj = {
                    "customer": {
                        "id": customer["id"],
                        "name": customer["name"]
                    }
                }

    # Step 7: Format selected_columns
    selected_columns_raw = draft.get("selected_columns") or "[]"

    try:
        selected_columns_list = json.loads(selected_columns_raw) if isinstance(selected_columns_raw, str) else selected_columns_raw
    except json.JSONDecodeError:
        selected_columns_list = []

    selected_columns_formatted = [
        {"column": col, "title": col.replace('_', ' ').title()}
        for col in selected_columns_list
    ]

    

    # Step 9: Final Response
    result = {
        "send_quotation_id": draft["id"],
        "version": draft.get("version"),
        "uploaded_date": draft.get("uploaded_date"),
        "uploaded_by": draft.get("uploaded_by"),
        "uploaded_by_name": uploaded_by_name,
        "opportunity_id": draft.get("opportunity_id"),
        "entity_id": draft.get("entity_id"),
        "status": draft.get("status"),
        "quotation_request_id": draft.get("quotation_request_id"),
        "service_provider_ids": list(service_provider_ids),
        "vendor_response_ids": vendor_response_objs,
        "values": values,
        "comment": comment_text,
        "selected_columns": selected_columns_formatted,
        **doc_info,
        **customer_obj
    }

    return ResponseService.response("SUCCESS", result,Message.DATA_FETCHED)



# -----------------------------------------------Email service-----------------------------------------------

@csrf_exempt
@api_view(["POST"])
def quotation_send_email(request):
    data = request.data

    # Step 1: Basic field validation
    rules = {
        "customer_id": "required",
        "subject": "required",
        "body": "required",
        "send_quotation_id": "required"   
    }
    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)

    customer_id = data["customer_id"]
    subject = data["subject"]
    body = data["body"]
    send_quotation_id = data["send_quotation_id"]

    # Step 2: Try to get primary contact from core_customers
    customer = QueryBuilderService("core_customers")\
        .where("id", customer_id)\
        .select("primary_contact_id")\
        .first()

    contact = None
    email = None

    if customer and customer.get("primary_contact_id"):
        contact = QueryBuilderService("core_contacts")\
            .where("id", customer["primary_contact_id"])\
            .select("email", "contact_email")\
            .first()
        if contact:
            email = contact.get("email") or contact.get("contact_email")

    # If no email from primary_contact_id, try core_customer_contacts
    if not email:
        contact_group = QueryBuilderService("core_customer_contacts")\
            .where("customer_id", customer_id)\
            .select("contact_id")\
            .get()
        for cg in contact_group:
            c = QueryBuilderService("core_contacts")\
                .where("id", cg["contact_id"])\
                .select("email", "contact_email")\
                .first()
            if c:
                email = c.get("email") or c.get("contact_email")
                if email:
                    break

    if not email:
        return ResponseService.response("VALIDATION_ERROR", "No valid email found for this customer.", Error.VALIDATION_ERROR)

    recipient_emails = [email]
    print("recipient_emails",recipient_emails)

    # Step 3: Prepare links
    links = data.get("links", [])
    documents = data.get("documents", [])

    # Process documents array to get CDN URLs from doc field
    document_cdn_links = []
    if isinstance(documents, list):
        print(f"\n=== EMAIL DEBUG: Processing {len(documents)} documents ===")
        document_cdn_links = DocumentCDNService.process_documents_for_email(documents)
        print(f"Generated {len(document_cdn_links)} CDN links: {document_cdn_links}")
        
        # Also handle legacy document_link field for backward compatibility
        for doc in documents:
            if isinstance(doc, dict) and doc.get("document_link"):
                links.append(doc["document_link"])

    portal_id = (
        QueryBuilderService("core_customers")\
        .where("id", customer_id)\
        .select("portal_id")\
        .first()    
    )

    if not portal_id:
        return ResponseService.response("NOT_FOUND", {}, Error.VALIDATION_ERROR)


    # Build customer portal link separately
    customer_portal_link = f"{settings.CUSTOMER_FRONTEND_BASE_URL.rstrip('/')}/{portal_id['portal_id']}/a/my-quotations"
    
    # Keep document CDN links separate for attachments
    fixed_links = document_cdn_links.copy()

    # Validate links as URL
    link_errors = {}
    print(f"\n=== EMAIL DEBUG: Validating {len(fixed_links)} links ===")
    for idx, link in enumerate(fixed_links):
        print(f"Validating link {idx}: {link}")
        single_validation = ValidatorService.validate({"link": link}, {"link": "required|url"})
        if single_validation:
            print(f"Link {idx} validation failed: {single_validation}")
            link_errors[str(idx)] = single_validation["link"]
        else:
            print(f"Link {idx} validation passed")

    if link_errors:
        return ResponseService.response(
            "VALIDATION_ERROR",
            {"email_data": [{"links": link_errors}]},
            Error.VALIDATION_ERROR
        )

    # Step 4: Format email body with customer portal link
    # Replace the placeholder link in the body with the actual customer portal link
    formatted_body = body.replace(
        '<a href="#" rel="noopener noreferrer" target="_blank">Proceed to Enrolment</a>',
        f'<a href="{customer_portal_link}" rel="noopener noreferrer" target="_blank">Proceed to Enrolment</a>'
    )
    
    # Step 5: Compose email payload
    email_payload = [{
        "recipient_email": email,
        "subject": subject,
        "body": formatted_body,
        "priority": "high",
        "links": fixed_links  # These are the document CDN links
    } for email in recipient_emails]
    
    print(f"\n=== EMAIL DEBUG: Final email payload ===")
    print(f"Recipient emails: {recipient_emails}")
    print(f"Subject: {subject}")
    print(f"Links to attach: {fixed_links}")
    print(f"Email payload: {email_payload}")

    # Step 6: Send email
    send_mail = SendMail()
    result = send_mail.send_email(email_payload)

    # Step 7: Interpret email sending result
    if isinstance(result, dict) and result.get("email_data"):
        for item in result["email_data"]:
            if "links" in item:
                return ResponseService.response(
                    "VALIDATION_ERROR",
                    result,
                   Error.VALIDATION_ERROR
                )

    # Step 8: Update quotation status to 'sent'
    QueryBuilderService("crmq_send_quotations")\
        .where("id", send_quotation_id)\
        .update({"status": "sent"})
    
    
    try:
            user = request.user
            print("User:", user.id)

            # Get send quotation details to fetch quotation and product information
            send_quotation = QueryBuilderService("crmq_send_quotations").where("id", send_quotation_id).first()
            
            # Get quotation details from entity_id
            quotation_details = None
            product_names = []
            recommendation_docs = []
            
            if send_quotation and send_quotation.get("entity_id"):
                quotation_details = QueryBuilderService("crmq_quotations")\
                    .select(
                        "crmq_quotations.id",
                        "crmq_quotations.code", 
                        "crmq_quotations.opportunity_type_id",
                        "crmq_quotations.opportunity_id",
                        "crmq_quotations.request_type",
                        "crmq_quotations.status",
                        "crmq_quotations.notes",
                        "crmq_quotations.email_data"
                    )\
                    .where("crmq_quotations.entity_id", send_quotation["entity_id"])\
                    .first()
                
                # Get product names from opportunity_type_id
                if quotation_details and quotation_details.get("opportunity_type_id"):
                    try:
                        opportunity_type_ids = quotation_details["opportunity_type_id"]
                        if isinstance(opportunity_type_ids, str):
                            import json
                            opportunity_type_ids = json.loads(opportunity_type_ids)
                        
                        if opportunity_type_ids and len(opportunity_type_ids) > 0:
                            opp_types = QueryBuilderService("crm_opportunity_types")\
                                .select("id", "title")\
                                .whereIn("id", opportunity_type_ids)\
                                .get()
                            product_names = [opp.get("title") for opp in opp_types if opp.get("title")]
                    except Exception as e:
                        print(f"Error fetching product names: {str(e)}")
                
                # Get recommendation documents from email_data
                if quotation_details and quotation_details.get("email_data"):
                    try:
                        email_data = quotation_details["email_data"]
                        if isinstance(email_data, str):
                            import json
                            email_data = json.loads(email_data)
                        
                        if isinstance(email_data, dict) and email_data.get("documents"):
                            recommendation_docs = email_data["documents"]
                    except Exception as e:
                        print(f"Error fetching recommendation documents: {str(e)}")

            # Build enhanced notification message
            notification_title = "Recommendation document for Quotation"
            notification_message = f"Recommendation documents are sent for Quotation with ID {send_quotation_id}"
            
            # Add product names to message
            if product_names:
                product_list = ", ".join(product_names)
                notification_message += f" for products: {product_list}"
            
            # Add detailed recommendation document information
            doc_count = 0
            if recommendation_docs:
                doc_count = len(recommendation_docs) if isinstance(recommendation_docs, list) else 1
            # Always include document count, even if zero
            # notification_message += f" with {doc_count} recommendation document(s)"
            
            # Add document names and details if documents exist
            if recommendation_docs and isinstance(recommendation_docs, list) and len(recommendation_docs) > 0:
                doc_names = []
                doc_types = []
                for doc in recommendation_docs:
                    if isinstance(doc, dict):
                        if doc.get("name"):
                            doc_names.append(doc["name"])
                        if doc.get("type"):
                            doc_types.append(doc["type"])
                
                if doc_names:
                    notification_message += f": {', '.join(doc_names)}"
                if doc_types and len(set(doc_types)) == 1:  # All same type
                    notification_message += f" ({doc_types[0]} files)"
                elif doc_types:  # Mixed types
                    notification_message += f" ({', '.join(set(doc_types))} files)"
            
            # Enhanced metadata with all required details
            meta_data = {
                "quotation_id": send_quotation_id,
                "recommendation_document": "sent",
                "id": send_quotation_id,
                "quotation_code": quotation_details.get("code") if quotation_details else None,
                "quotation_status": quotation_details.get("status") if quotation_details else None,
                "request_type": quotation_details.get("request_type") if quotation_details else None,
                "product_names": product_names,
                "recommendation_documents_count": len(recommendation_docs) if isinstance(recommendation_docs, list) else (1 if recommendation_docs else 0),
                "recommendation_documents": recommendation_docs if isinstance(recommendation_docs, list) else [],
                "document_names": [doc.get("name") for doc in recommendation_docs if isinstance(doc, dict) and doc.get("name")] if isinstance(recommendation_docs, list) else [],
                "document_types": [doc.get("type") for doc in recommendation_docs if isinstance(doc, dict) and doc.get("type")] if isinstance(recommendation_docs, list) else [],
                "opportunity_id": quotation_details.get("opportunity_id") if quotation_details else None,
                "notes": quotation_details.get("notes") if quotation_details else None
            }

            NotificationService.generate_notification(
                type_code="quotation",  #notification type code
                title=notification_title,
                meta_data=meta_data,
                message=notification_message,
                customer_id=customer_id,
                user_id=user.id if user else None
            )
    except Exception as notify_exc:
            print(f"NotificationService error: {notify_exc}")


    return ResponseService.response("SUCCESS", result, Message.EMAIL_SENT)



#----------------------------------------------------------------------------------------



@csrf_exempt
@api_view(['POST'])
def html_to_pdf_export(request, send_quotation_id):
    data = request.data

    # Step 1: Validate HTML tag input
    rules = {"html_tag": "required|string"}
    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors,Error.VALIDATION_ERROR)

    html_content = data["html_tag"]

    # Step 2: Retrieve version from crmq_send_quotations
    send_quotation = QueryBuilderService("crmq_send_quotations").where("id", send_quotation_id).first()
    if not send_quotation:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "Send quotation not found.")

    version = send_quotation.get("version")
    if version is None:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "Version not found in send quotation.")

    # Step 3: Retrieve vendor_quotation_ids
    vendor_quotation_links = QueryBuilderService("crmq_quotation_vendor_quotations")\
        .where("send_quotation_id", send_quotation_id).get()

    if not vendor_quotation_links:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "No vendor quotations found.")

    vendor_quotation_ids = [v["vendor_quotation_id"] for v in vendor_quotation_links if v.get("vendor_quotation_id")]
    if not vendor_quotation_ids:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "Vendor quotation IDs are missing.")

    # Step 4: Get response data from crmq_vendor_response
    responses = QueryBuilderService("crmq_vendor_response")\
        .whereIn("vendor_quotation_id", vendor_quotation_ids).get()

    if not responses:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "No responses found for vendor quotations.")

    document_details = []
    selected_codes = []

    for res in responses:
        code = res.get("code")
        coverage = res.get("coverage_details")
        coverage_name = res.get("coverage_details_name")
        coverage_type = res.get("coverage_details_type")

        if code:
            selected_codes.append(code)

        if coverage:
            # coverage_name = coverage.split("/")[-1]
            # coverage_type = coverage_name.split(".")[-1]
            document_details.append({
                "selected_quotation": code,
                "coverage_details": coverage,
                "coverage_details_name": coverage_name,
                "coverage_details_type": coverage_type
            })

    if not selected_codes:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "No selected quotations found.")

    # Step 5: Generate PDF
    exporter = ExportToPdf()
    result = exporter.export_html_to_pdf(html_content)
    print("PDF export result:", result)

    if not result or not isinstance(result, dict) or result.get("status") != "SUCCESS" or not result.get("data"):
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, "PDF export failed.")

    download_url = result["data"].get("download_url")
    if not download_url:
        return ResponseService.response("INTERNAL_SERVER_ERROR", None,"Download URL not found in PDF export response.")

    suffixes = [code.replace("QRI-", "") for code in selected_codes]
    generated_name = f"Recommendation_doc_{'&'.join(suffixes)}v{version}.pdf"

    # Step 6: Download document from ExportToPdf URL and upload to S3
    try:
        s3_upload_result = S3PresignedService.upload_stream_from_url(
            file_url=download_url,
            file_name=generated_name,
            folder="exports/quotations"
        )
        s3_key = s3_upload_result["file_key"]
        
        # Construct CDN URL using CDN_BASE_URL
        cdn_base_url = os.getenv("CDN_BASE_URL")
        cdn_url = f"{cdn_base_url}/{s3_key}"
    except Exception as e:
        print(f"Error uploading to S3: {str(e)}")
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Failed to upload document to S3: {str(e)}")

    # Step 7: Save to DB
    pdf_meta = {
        "name": generated_name,
        "type": "pdf",
        "link": cdn_url,
        "s3_key": s3_key
    }

    QueryBuilderService("crmq_send_quotations")\
        .where("id", send_quotation_id)\
        .update({"generated_pdf": json.dumps(pdf_meta)})

    return ResponseService.response("SUCCESS", {
        "pdf_data": {
            "coverage_details": s3_key,
            "coverage_details_name": generated_name,
            "version": version,
            "download_link": cdn_url
        },
        "document_details": document_details
    }, Message.PDF_GENERATED)



# @csrf_exempt
# @api_view(['GET'])
# def get_doc_version(request, id):
#     # Fetch the latest version for the given quotation_request_id
#     last_version_record = QueryBuilderService('crmq_send_quotations')\
#         .where('quotation_request_id', id)\
#         .select('version')\
#         .orderBy('version', 'desc')\
#         .first()  # Fetch the latest record

#     # Extract the version or set it to None if no record is found
#     last_version = last_version_record['version'] if last_version_record else 0

#     print("last_version", last_version)

#     new_version = last_version + 1

#     return ResponseService.response("SUCCESS", {"new_version": new_version}, "default_get_success_msg")
                          

@csrf_exempt
@api_view(['GET'])
def get_doc_version(request, ids):
    try:
        # Step 1: Parse incoming vendor_response_ids
        vendor_response_ids = [int(id.strip()) for id in ids.split(',') if id.strip().isdigit()]
        if not vendor_response_ids:
            return ResponseService.response("VALIDATION_ERROR", "No valid vendor response IDs provided.", Error.VALIDATION_ERROR)

        # Step 2: Fetch the first quotation_id from vendor_response
        vendor_response_record = QueryBuilderService('crmq_vendor_response')\
            .whereIn('id', vendor_response_ids)\
            .select('quotation_id')\
            .first()

        if not vendor_response_record or not vendor_response_record.get('quotation_id'):
            return ResponseService.response("VALIDATION_ERROR", "No quotation found for the provided vendor response IDs.", Error.VALIDATION_ERROR)

        quotation_id = vendor_response_record['quotation_id']

        # Step 3: Fetch latest version for that quotation_id
        last_version_record = QueryBuilderService('crmq_send_quotations')\
            .where('quotation_request_id', quotation_id)\
            .select('version')\
            .orderBy('version', 'desc')\
            .first()

        last_version = last_version_record['version'] if last_version_record else 0
        new_version = last_version + 1

        return ResponseService.response("SUCCESS", {"new_version": new_version},Message.DATA_FETCHED)

    except Exception as e:
        print("Exception:", e)
        return ResponseService.response("INTERNAL_SERVER_ERROR", str(e), "An error occurred while generating version.")





@api_view(["PUT"])
def get_single_generate_document_confirm(request, vendor_quotation_id):
    """
    Confirm a sent quotation by setting its status to 'confirmed', if not already confirmed.
    """
    # Check if the record exists
    record = QueryBuilderService("crmq_quotation_service_providers").where("id", vendor_quotation_id).first()
    if not record:
        return ResponseService.response(
            "NOT_FOUND",
            {"error": "No sent quotation found with this id."},
            "not_found"
        )
    if record.get("status") == "confirmed":
        return ResponseService.response(
            "NOT_FOUND",
            {"error": "This quotation is already confirmed."},
            "already_confirmed"
        )

    # Additional validation: vendor response must not be in confirmed/rejected/expired types
    vendor_response = (
        QueryBuilderService("crmq_vendor_response")
        .leftJoin("core_status", "core_status.name", "crmq_vendor_response.status")
        .select(
            "crmq_vendor_response.id",
            "crmq_vendor_response.status",
            "core_status.type as status_type",
        )
        .where("vendor_quotation_id", vendor_quotation_id)
        .first()
    )
    if vendor_response:
        status_type = vendor_response.get("status_type")
        status_name = vendor_response.get("status")
        forbidden_types = ["quotation_confirmed", "quotation_rejected", "quotation_expired"]
        if (status_type and status_type in forbidden_types) or (status_name in ["CONFIRMED", "REJECTED", "EXPIRED"]):
            return ResponseService.response(
                "VALIDATION_ERROR",
                {"error": "This vendor response cannot be confirmed"},
                "quotation_conflict"
            )

    # Get the quotation_id from the current record to find related quotations
    quotation_id = record.get("quotation_id")
    
   
    # Get the accepted status ID
    accepted_status_id = (
        QueryBuilderService("core_status")
        .where("type", "quotation_confirmed")
        .where("module", "quotation")
        .select("id","name")
        .first()
    )

    if not accepted_status_id:
        return ResponseService.response(
            "NOT_FOUND",
            {"error": "Confirmed status not found in core_status table."},
            "status_not_found"
        )

    # Get the rejected status ID
    rejected_status_id = (
        QueryBuilderService("core_status")
        .where("type", "quotation_rejected")
        .where("module", "quotation")
        .select("id","name")
        .first()
    )

    if not rejected_status_id:
        return ResponseService.response(
            "NOT_FOUND",
            {"error": "Rejected status not found in core_status table."},
            "status_not_found"
        )

    # Update the confirmed quotation in crmq_vendor_response table
    QueryBuilderService("crmq_vendor_response")\
        .where("quotation_id", quotation_id)\
        .where("vendor_quotation_id", vendor_quotation_id)\
        .update({"status": accepted_status_id["name"]})

    # Update all other related quotations to rejected status in crmq_vendor_response table
    QueryBuilderService("crmq_vendor_response")\
        .where("quotation_id", quotation_id)\
        .whereNotIn("vendor_quotation_id", [vendor_quotation_id])\
        .update({"status": rejected_status_id["name"]})

    # Update the confirmed quotation
    QueryBuilderService("crmq_quotation_service_providers").where("id", vendor_quotation_id).update({
        "status": accepted_status_id.get("id")
    })

    # Update all other related quotations to rejected status
    QueryBuilderService("crmq_quotation_service_providers").where("quotation_id", quotation_id).whereNotIn("id", [vendor_quotation_id]).update({
        "status": rejected_status_id.get("id")
    })

    # Update the main quotation record in crmq_quotations table
    QueryBuilderService("crmq_quotations").where("id", quotation_id).update({
        "status_id": accepted_status_id.get("id"),
        "status": accepted_status_id.get("name")
    })

    return ResponseService.response(
        "SUCCESS",
        {"id": vendor_quotation_id, "status": "confirmed"},
        "quotation_confirmed"
    )



@csrf_exempt
@api_view(["POST"])
def send_email_customers_quotation(request):
    data = request.data

    # Step 1: Basic field validation
    rules = {
        "customer_id": "required",
        "subject": "required",
        "body": "required",
        "send_quotation_id": "required"   
    }
    errors = ValidatorService.validate(data, rules)
    if errors:
        return ResponseService.response("VALIDATION_ERROR", errors, Error.VALIDATION_ERROR)

    customer_id = data["customer_id"]
    subject = data["subject"]
    body = data["body"]
    send_quotation_id = data["send_quotation_id"]

    # Step 2: Try to get primary contact from core_customers
    customer = QueryBuilderService("core_customers")\
        .where("id", customer_id)\
        .select("primary_contact_id")\
        .first()

    contact = None
    email = None

    if customer and customer.get("primary_contact_id"):
        contact = QueryBuilderService("core_contacts")\
            .where("id", customer["primary_contact_id"])\
            .select("email", "contact_email")\
            .first()
        if contact:
            email = contact.get("email") or contact.get("contact_email")

    # If no email from primary_contact_id, try core_customer_contacts
    if not email:
        contact_group = QueryBuilderService("core_customer_contacts")\
            .where("customer_id", customer_id)\
            .select("contact_id")\
            .get()
        for cg in contact_group:
            c = QueryBuilderService("core_contacts")\
                .where("id", cg["contact_id"])\
                .select("email", "contact_email")\
                .first()
            if c:
                email = c.get("email") or c.get("contact_email")
                if email:
                    break

    if not email:
        return ResponseService.response("VALIDATION_ERROR", "No valid email found for this customer.", Error.VALIDATION_ERROR)

    recipient_emails = [email]
    print("recipient_emails",recipient_emails)

    # Step 3: Prepare links
    links = data.get("links", [])
    documents = data.get("documents", [])

    # Process documents array to get CDN URLs from doc field
    document_cdn_links = []
    if isinstance(documents, list):
        print(f"\n=== EMAIL DEBUG: Processing {len(documents)} documents ===")
        document_cdn_links = DocumentCDNService.process_documents_for_email(documents)
        print(f"Generated {len(document_cdn_links)} CDN links: {document_cdn_links}")
        
        # Also handle legacy document_link field for backward compatibility
        for doc in documents:
            if isinstance(doc, dict) and doc.get("document_link"):
                links.append(doc["document_link"])

    portal_id = (
        QueryBuilderService("core_customers")\
        .where("id", customer_id)\
        .select("portal_id")\
        .first()    
    )

    if not portal_id:
        return ResponseService.response("NOT_FOUND", {}, Error.VALIDATION_ERROR)


    # Build customer portal link separately
    customer_portal_link = f"{settings.CUSTOMER_FRONTEND_BASE_URL.rstrip('/')}/{portal_id['portal_id']}/a/my-quotations"
    
    # Keep document CDN links separate for attachments
    fixed_links = document_cdn_links.copy()

    # Validate links as URL
    link_errors = {}
    print(f"\n=== EMAIL DEBUG: Validating {len(fixed_links)} links ===")
    for idx, link in enumerate(fixed_links):
        print(f"Validating link {idx}: {link}")
        single_validation = ValidatorService.validate({"link": link}, {"link": "required|url"})
        if single_validation:
            print(f"Link {idx} validation failed: {single_validation}")
            link_errors[str(idx)] = single_validation["link"]
        else:
            print(f"Link {idx} validation passed")

    if link_errors:
        return ResponseService.response(
            "VALIDATION_ERROR",
            {"email_data": [{"links": link_errors}]},
            Error.VALIDATION_ERROR
        )

    # Step 4: Format email body with customer portal link
    # Replace the placeholder link in the body with the actual customer portal link
    formatted_body = body.replace(
        '<a href="#" rel="noopener noreferrer" target="_blank">Proceed to Enrolment</a>',
        f'<a href="{customer_portal_link}" rel="noopener noreferrer" target="_blank">Proceed to Enrolment</a>'
    )
    
    # Step 5: Download documents and create attachments
    attachments = []
    if fixed_links:
        try:
            import requests
            for link in fixed_links:
                try:
                    # Download the document
                    response = requests.get(link, timeout=30)
                    if response.status_code == 200:
                        # Extract filename from URL or use default
                        filename = os.path.basename(link.split('?')[0]) if isinstance(link, str) else "document.pdf"
                        if not filename or '.' not in filename:
                            filename = "document.pdf"
                        
                        # Determine content type based on file extension
                        content_type = "application/pdf"
                        if filename.lower().endswith(('.doc', '.docx')):
                            content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        elif filename.lower().endswith(('.xls', '.xlsx')):
                            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        elif filename.lower().endswith(('.jpg', '.jpeg')):
                            content_type = "image/jpeg"
                        elif filename.lower().endswith('.png'):
                            content_type = "image/png"
                        
                        attachments.append({
                            'filename': filename,
                            'content_type': content_type,
                            'data': response.content
                        })
                except Exception as e:
                    print(f"Failed to download document from {link}: {str(e)}")
                    continue
        except Exception as e:
            print(f"Error processing document attachments: {str(e)}")

    # Step 6: Fetch the system Gmail credential
    cred = None
    try:
        cred = GmailCredential.objects.first()
    except Exception:
        cred = None

    if not cred:
        return ResponseService.response("VALIDATION_ERROR", "No Gmail credential configured.", Error.VALIDATION_ERROR)

    send_result = None
    try:
        # Only one recipient is expected here
        primary_email = recipient_emails[0]
        send_result = gmail_send_email(cred, primary_email, subject, formatted_body, attachments=attachments)
    except Exception as e:
        return ResponseService.response("INTERNAL_SERVER_ERROR", str(e), Error.INTERNAL_SERVER_ERROR)

    # Step 7: Update quotation status to 'sent'
    QueryBuilderService("crmq_send_quotations")\
        .where("id", send_quotation_id)\
        .update({"status": "sent"})
    
    
    try:
            user = request.user
            print("User:", user.id)

            # Get send quotation details to fetch quotation and product information
            send_quotation = QueryBuilderService("crmq_send_quotations").where("id", send_quotation_id).first()
            
            # Get quotation details from entity_id
            quotation_details = None
            product_names = []
            recommendation_docs = []
            
            if send_quotation and send_quotation.get("entity_id"):
                quotation_details = QueryBuilderService("crmq_quotations")\
                    .select(
                        "crmq_quotations.id",
                        "crmq_quotations.code", 
                        "crmq_quotations.opportunity_type_id",
                        "crmq_quotations.opportunity_id",
                        "crmq_quotations.request_type",
                        "crmq_quotations.status",
                        "crmq_quotations.notes",
                        "crmq_quotations.email_data"
                    )\
                    .where("crmq_quotations.entity_id", send_quotation["entity_id"])\
                    .first()
                
                # Get product names from opportunity_type_id
                if quotation_details and quotation_details.get("opportunity_type_id"):
                    try:
                        opportunity_type_ids = quotation_details["opportunity_type_id"]
                        if isinstance(opportunity_type_ids, str):
                            import json
                            opportunity_type_ids = json.loads(opportunity_type_ids)
                        
                        if opportunity_type_ids and len(opportunity_type_ids) > 0:
                            opp_types = QueryBuilderService("crm_opportunity_types")\
                                .select("id", "title")\
                                .whereIn("id", opportunity_type_ids)\
                                .get()
                            product_names = [opp.get("title") for opp in opp_types if opp.get("title")]
                    except Exception as e:
                        print(f"Error fetching product names: {str(e)}")
                
                # Get recommendation documents from email_data
                if quotation_details and quotation_details.get("email_data"):
                    try:
                        email_data = quotation_details["email_data"]
                        if isinstance(email_data, str):
                            import json
                            email_data = json.loads(email_data)
                        
                        if isinstance(email_data, dict) and email_data.get("documents"):
                            recommendation_docs = email_data["documents"]
                    except Exception as e:
                        print(f"Error fetching recommendation documents: {str(e)}")

            # Build enhanced notification message
            notification_title = "Recommendation document for Quotation"
            notification_message = f"Recommendation documents are sent for Quotation with ID {send_quotation_id}"
            
            # Add product names to message
            if product_names:
                product_list = ", ".join(product_names)
                notification_message += f" for products: {product_list}"
            
            # Add detailed recommendation document information
            doc_count = 0
            if recommendation_docs:
                doc_count = len(recommendation_docs) if isinstance(recommendation_docs, list) else 1
            # Always include document count, even if zero
            # notification_message += f" with {doc_count} recommendation document(s)"
            
            # Add document names and details if documents exist
            if recommendation_docs and isinstance(recommendation_docs, list) and len(recommendation_docs) > 0:
                doc_names = []
                doc_types = []
                for doc in recommendation_docs:
                    if isinstance(doc, dict):
                        if doc.get("name"):
                            doc_names.append(doc["name"])
                        if doc.get("type"):
                            doc_types.append(doc["type"])
                
                if doc_names:
                    notification_message += f": {', '.join(doc_names)}"
                if doc_types and len(set(doc_types)) == 1:  # All same type
                    notification_message += f" ({doc_types[0]} files)"
                elif doc_types:  # Mixed types
                    notification_message += f" ({', '.join(set(doc_types))} files)"
            
            # Enhanced metadata with all required details
            meta_data = {
                "quotation_id": send_quotation_id,
                "recommendation_document": "sent",
                "id": send_quotation_id,
                "quotation_code": quotation_details.get("code") if quotation_details else None,
                "quotation_status": quotation_details.get("status") if quotation_details else None,
                "request_type": quotation_details.get("request_type") if quotation_details else None,
                "product_names": product_names,
                "recommendation_documents_count": len(recommendation_docs) if isinstance(recommendation_docs, list) else (1 if recommendation_docs else 0),
                "recommendation_documents": recommendation_docs if isinstance(recommendation_docs, list) else [],
                "document_names": [doc.get("name") for doc in recommendation_docs if isinstance(doc, dict) and doc.get("name")] if isinstance(recommendation_docs, list) else [],
                "document_types": [doc.get("type") for doc in recommendation_docs if isinstance(doc, dict) and doc.get("type")] if isinstance(recommendation_docs, list) else [],
                "opportunity_id": quotation_details.get("opportunity_id") if quotation_details else None,
                "notes": quotation_details.get("notes") if quotation_details else None
            }

            NotificationService.generate_notification(
                type_code="quotation",  #notification type code
                title=notification_title,
                meta_data=meta_data,
                message=notification_message,
                customer_id=customer_id,
                user_id=user.id if user else None
            )
    except Exception as notify_exc:
            print(f"NotificationService error: {notify_exc}")


    return ResponseService.response("SUCCESS", send_result, Message.EMAIL_SENT)


def analyze_document_with_external_api(document_url, document_type="policy"):
    """
    Analyze document using external data analyzer API.
    """
    try:
        DATA_ANALYZER = getattr(settings, 'DATA_ANALYZER', None)
        CDN_BASE_URL = getattr(settings, 'CDN_BASE_URL', None)
        
        if not DATA_ANALYZER or not CDN_BASE_URL:
            print("DATA_ANALYZER or CDN_BASE_URL not configured in settings")
            return None
        
        if not document_url:
            print("No document_url provided")
            return None
        
        # Check if document_url already includes the full URL or just the path
        if document_url.startswith('http://') or document_url.startswith('https://'):
            full_document_url = document_url
        else:
            full_document_url = CDN_BASE_URL + document_url
        
        payload = {
            "document_urls": [full_document_url], 
            "document_type": document_type
        }
        
        print(f"Calling external data analyzer with payload: {payload}")
        # Increase timeout for document analysis (60 seconds connect, 300 seconds read)
        response = requests.post(DATA_ANALYZER, json=payload, timeout=(60, 300))
        
        if response.status_code == 200:
            result = response.json()
            print(f"External API response received: {len(result) if isinstance(result, list) else 'not a list'}")
            return result
        else:
            print(f"External API request failed with status {response.status_code}: {response.text}")
            return None
            
    except requests.exceptions.Timeout:
        print("External API request timed out after 5 minutes")
        return None
    except requests.exceptions.RequestException as e:
        print(f"External API request error: {str(e)}")
        return None
    except Exception as e:
        print(f"Error analyzing document with external API: {str(e)}")
        return None


def parse_document_from_url(file_url, file_name, content_type):
    """
    Download and parse document from URL to extract data.
    Supports Excel (.xlsx, .xls) and PDF files.
    First tries external data analyzer API, then falls back to local parsing.
    Returns extracted data in structured format.
    """
    extracted_data = {}
    
    # First, try external data analyzer API
    try:
        # Determine document type from file name
        file_name_lower = file_name.lower()
        document_type = "policy"
        if "invoice" in file_name_lower:
            document_type = "invoice"
        elif "quotation" in file_name_lower:
            document_type = "quotation"
        
        external_analysis = analyze_document_with_external_api(file_url, document_type)
        
        if external_analysis and isinstance(external_analysis, list) and len(external_analysis) > 0:
            # Use external analysis result
            analysis_result = external_analysis[0]
            
            # Extract details from external analysis
            if isinstance(analysis_result, dict):
                details = analysis_result.get("details", {})
                policy_fields = details.get("policy_fields", {})
                endorsement_fields = details.get("endorsement_fields", {})
                
                extracted_data = {
                    "source": "external_api",
                    "raw_analysis": analysis_result,
                    "extracted_fields": {},
                    "policy_fields": policy_fields,
                    "endorsement_fields": endorsement_fields,
                }
                
                # Map policy fields
                if policy_fields:
                    extracted_data["extracted_fields"].update({
                        "insurer_policy_id": policy_fields.get("insurer_policy_id", ""),
                        "policy_issue_date": policy_fields.get("policy_issue_date", ""),
                        "start_date": policy_fields.get("start_date", ""),
                        "end_date": policy_fields.get("end_date", ""),
                        "sum_insured": policy_fields.get("sum_insured", ""),
                        "risk_type": policy_fields.get("risk_type", ""),
                        "payment_mode": policy_fields.get("payment_mode", ""),
                    })
                
                # Map invoice fields
                if endorsement_fields:
                    extracted_data["extracted_fields"].update({
                        "insurer_invoice_id": endorsement_fields.get("insurer_invoice_id", ""),
                        "insurer_invoice_number": endorsement_fields.get("insurer_invoice_number", ""),
                        "amount_or_cover_value": endorsement_fields.get("amount_or_cover_value", ""),
                    })
                
                return extracted_data
    except Exception as e:
        print(f"Error with external API, falling back to local parsing: {str(e)}")
        # Continue with local parsing
    
    # Fall back to local parsing if external API fails or is not available
    try:
        # Download the file
        response = requests.get(file_url, timeout=30, stream=True)
        response.raise_for_status()
        
        file_extension = os.path.splitext(file_name)[1].lower()
        file_content = response.content
        
        # Mark as local parsing
        extracted_data["source"] = "local_parsing"
        
        # Parse based on file type
        if file_extension in ['.xlsx', '.xls'] or 'spreadsheet' in content_type.lower():
            # Parse Excel file
            try:
                import openpyxl
                from openpyxl import load_workbook
                
                # Load workbook from bytes
                workbook = load_workbook(io.BytesIO(file_content))
                sheet = workbook.active
                
                # Extract data from Excel - look for common fields
                # This is a basic extraction - can be customized based on document structure
                data_dict = {}
                extracted_data["extracted_fields"] = {}
                
                # Read all rows and try to find key-value pairs
                for row in sheet.iter_rows(values_only=True):
                    if row and len(row) >= 2:
                        key = str(row[0]).strip() if row[0] else ""
                        value = str(row[1]).strip() if row[1] else ""
                        
                        if key and value:
                            # Normalize keys
                            key_lower = key.lower()
                            
                            # Map common field names
                            if 'policy' in key_lower and 'id' in key_lower:
                                data_dict['insurer_policy_id'] = value
                            elif 'policy' in key_lower and 'issue' in key_lower:
                                data_dict['policy_issue_date'] = value
                            elif 'start' in key_lower and 'date' in key_lower:
                                data_dict['start_date'] = value
                            elif 'end' in key_lower and 'date' in key_lower:
                                data_dict['end_date'] = value
                            elif 'sum' in key_lower and 'insured' in key_lower:
                                data_dict['sum_insured'] = value
                            elif 'risk' in key_lower and 'type' in key_lower:
                                data_dict['risk_type'] = value
                            elif 'payment' in key_lower and 'mode' in key_lower:
                                data_dict['payment_mode'] = value
                            elif 'invoice' in key_lower and 'id' in key_lower:
                                data_dict['insurer_invoice_id'] = value
                            elif 'invoice' in key_lower and 'number' in key_lower:
                                data_dict['insurer_invoice_number'] = value
                            elif 'amount' in key_lower or 'cover' in key_lower and 'value' in key_lower:
                                data_dict['amount_or_cover_value'] = value
                            else:
                                # Store other fields
                                data_dict[key] = value
                
                extracted_data["extracted_fields"] = data_dict
                
            except ImportError:
                print("openpyxl not installed. Install it with: pip install openpyxl")
                # Try with pandas as fallback
                try:
                    import pandas as pd
                    df = pd.read_excel(io.BytesIO(file_content))
                    # Convert to dictionary and extract fields
                    records = df.to_dict('records')
                    extracted_data["extracted_fields"] = records[0] if records else {}
                except ImportError:
                    print("Neither openpyxl nor pandas available for Excel parsing")
                    return None
            except Exception as e:
                print(f"Error parsing Excel file: {str(e)}")
                return None
                
        elif file_extension == '.pdf' or 'pdf' in content_type.lower():
            # Parse PDF file
            try:
                import PyPDF2
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                text_content = ""
                
                # Extract text from all pages
                for page in pdf_reader.pages:
                    text_content += page.extract_text() + "\n"
                
                # Try to extract structured data from PDF text
                # Look for common patterns like key-value pairs
                extracted_data = {
                    "raw_text": text_content,
                    "page_count": len(pdf_reader.pages),
                    "extracted_fields": {}
                }
                
                # Try to parse key-value pairs from the text
                lines = text_content.split('\n')
                for i, line in enumerate(lines):
                    line = line.strip()
                    if not line:
                        continue
                    
                    # Look for common patterns like "Key: Value" or "Key - Value"
                    if ':' in line:
                        parts = line.split(':', 1)
                        if len(parts) == 2:
                            key = parts[0].strip().lower()
                            value = parts[1].strip()
                            
                            # Map common field names
                            if 'policy' in key and 'id' in key:
                                extracted_data["extracted_fields"]['insurer_policy_id'] = value
                            elif 'policy' in key and ('issue' in key or 'date' in key):
                                extracted_data["extracted_fields"]['policy_issue_date'] = value
                            elif 'start' in key and 'date' in key:
                                extracted_data["extracted_fields"]['start_date'] = value
                            elif 'end' in key and 'date' in key:
                                extracted_data["extracted_fields"]['end_date'] = value
                            elif 'sum' in key and 'insured' in key:
                                extracted_data["extracted_fields"]['sum_insured'] = value
                            elif 'risk' in key and 'type' in key:
                                extracted_data["extracted_fields"]['risk_type'] = value
                            elif 'payment' in key and 'mode' in key:
                                extracted_data["extracted_fields"]['payment_mode'] = value
                            elif 'invoice' in key and 'id' in key:
                                extracted_data["extracted_fields"]['insurer_invoice_id'] = value
                            elif 'invoice' in key and 'number' in key:
                                extracted_data["extracted_fields"]['insurer_invoice_number'] = value
                            elif 'amount' in key or ('cover' in key and 'value' in key):
                                extracted_data["extracted_fields"]['amount_or_cover_value'] = value
                            elif ('received' in key and 'date' in key) or ('issued' in key and 'date' in key) or \
                                 ('from' in key and 'date' in key) or ('effective' in key and 'from' in key):
                                # Extract received_date using various keywords
                                if 'received' in key and 'date' in key:
                                    extracted_data["extracted_fields"]['received date'] = value
                                elif 'issued' in key and 'date' in key:
                                    extracted_data["extracted_fields"]['issued date'] = value
                                elif 'from' in key and 'date' in key:
                                    extracted_data["extracted_fields"]['from date'] = value
                                elif 'effective' in key and 'from' in key:
                                    extracted_data["extracted_fields"]['effective from'] = value
                            else:
                                # Store other key-value pairs
                                extracted_data["extracted_fields"][key] = value
                
                # Also try pdfplumber for better extraction if available
                try:
                    import pdfplumber
                    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                        tables_data = []
                        all_tables = []
                        for page in pdf.pages:
                            # First try default table extraction
                            try:
                                tables = page.extract_tables()
                            except Exception as e:
                                print(f"Error with default table extraction: {str(e)}")
                                tables = []
                            
                            if tables:
                                for table in tables:
                                    # Clean up table data - remove None values and empty rows
                                    cleaned_table = []
                                    for row in table:
                                        if row:
                                            cleaned_row = [cell.strip() if cell and isinstance(cell, str) else (cell if cell else "") for cell in row]
                                            # Only add row if it has at least one non-empty cell
                                            if any(cell for cell in cleaned_row):
                                                cleaned_table.append(cleaned_row)
                                    
                                    if cleaned_table:
                                        all_tables.append(cleaned_table)
                            
                            # Also try extracting tables with text strategy
                            try:
                                tables_text = page.extract_tables({
                                    "vertical_strategy": "text",
                                    "horizontal_strategy": "text",
                                })
                                
                                if tables_text:
                                    for table in tables_text:
                                        cleaned_table = []
                                        for row in table:
                                            if row:
                                                cleaned_row = [cell.strip() if cell and isinstance(cell, str) else (cell if cell else "") for cell in row]
                                                if any(cell for cell in cleaned_row):
                                                    cleaned_table.append(cleaned_row)
                                        
                                        if cleaned_table:
                                            # Check if this table is different from already extracted ones
                                            if cleaned_table not in all_tables:
                                                all_tables.append(cleaned_table)
                            except Exception as e:
                                print(f"Error with text strategy table extraction: {str(e)}")
                                # Continue with default extraction results
                            
                            # Extract text with better formatting
                            page_text = page.extract_text()
                            if page_text and not extracted_data.get("formatted_text"):
                                extracted_data["formatted_text"] = page_text
                        
                        # Also try to parse table data from formatted_text to supplement table extraction
                        if extracted_data.get("formatted_text"):
                            text_content = extracted_data["formatted_text"]
                            
                            # Look for table patterns in the text
                            # Pattern: Header row followed by data rows
                            lines = text_content.split('\n')
                            
                            # Find lines that look like table headers
                            header_keywords = ['quotation version', 'quotation value', 'received date', 'expiry date', 
                                             'version', 'value', 'date', 'received', 'expiry', 'issued date', 
                                             'from date', 'effective from', 'effective from date']
                            
                            for i, line in enumerate(lines):
                                line_lower = line.lower()
                                # Check if this line contains table header keywords
                                if any(keyword in line_lower for keyword in header_keywords):
                                    # This might be a header row
                                    # Try to split the line - might have header and data on same line
                                    header_parts = [p.strip() for p in line.split() if p.strip()]
                                    
                                    # Check if this line also contains data (numbers, dates)
                                    has_data_in_line = any(
                                        any(char.isdigit() for char in part) or 
                                        '/' in part or 
                                        (part.count('-') >= 2 and any(char.isdigit() for char in part))
                                        for part in header_parts
                                    )
                                    
                                    if len(header_parts) >= 2:
                                        table_rows = []
                                        
                                        # If line has both header and data, try to separate them
                                        if has_data_in_line and len(header_parts) >= 6:
                                            # Might be header + data on same line
                                            # Look for where header ends and data begins
                                            # Headers usually have words, data has numbers/dates
                                            header_end = 0
                                            for idx, part in enumerate(header_parts):
                                                # If we find a part that's mostly numbers/dates, that's likely data start
                                                if (any(char.isdigit() for char in part) and 
                                                    (part.count('-') >= 2 or part.count('/') >= 2 or 
                                                     part.replace('.', '').replace(',', '').isdigit())):
                                                    header_end = idx
                                                    break
                                            
                                            if header_end > 0:
                                                # Split into header and data
                                                header_row = header_parts[:header_end]
                                                data_row = header_parts[header_end:]
                                                
                                                # Try to align columns
                                                # Common pattern: "Quotation Version", "Quotation Value", "Received Date", "Expiry Date"
                                                # Data: "-", "201000", "2025-04-30", "2025-04-01"
                                                if len(data_row) >= len(header_row) - 1:
                                                    # Pad header if needed
                                                    while len(header_row) < len(data_row):
                                                        header_row.append("")
                                                    # Trim data if needed
                                                    if len(data_row) > len(header_row):
                                                        data_row = data_row[:len(header_row)]
                                                    
                                                    table_rows = [header_row, data_row]
                                        else:
                                            # Header only, look for data in next lines
                                            table_rows = [header_parts]
                                            
                                            # Check next few lines for data rows
                                            for j in range(i + 1, min(i + 10, len(lines))):
                                                data_line = lines[j].strip()
                                                if not data_line:
                                                    continue
                                                
                                                # Split by spaces to get columns
                                                data_parts = [p.strip() for p in data_line.split() if p.strip()]
                                                
                                                # Check if this looks like a data row
                                                if len(data_parts) >= len(header_parts) - 1:
                                                    # Check if it has data (numbers, dates, etc.)
                                                    has_data = any(
                                                        any(char.isdigit() for char in part) or 
                                                        '/' in part or 
                                                        '-' in part or
                                                        part.replace('.', '').replace(',', '').isdigit()
                                                        for part in data_parts
                                                    )
                                                    
                                                    if has_data:
                                                        # Try to align columns - pad if needed
                                                        while len(data_parts) < len(header_parts):
                                                            data_parts.append("")
                                                        # Trim if too many
                                                        if len(data_parts) > len(header_parts):
                                                            data_parts = data_parts[:len(header_parts)]
                                                        
                                                        table_rows.append(data_parts)
                                                    elif len(data_parts) == len(header_parts):
                                                        # Same number of columns, might be data
                                                        table_rows.append(data_parts)
                                                elif len(data_parts) < len(header_parts) - 2:
                                                    # Too few columns, probably not a data row
                                                    break
                                        
                                        # If we found a complete table (header + at least one data row)
                                        if len(table_rows) > 1:
                                            # Check if this table is better than what we already have
                                            found_better = False
                                            for idx, existing_table in enumerate(all_tables):
                                                # If existing table only has header, replace it
                                                if len(existing_table) == 1 and len(table_rows) > 1:
                                                    all_tables[idx] = table_rows
                                                    found_better = True
                                                    break
                                            
                                            if not found_better and table_rows not in all_tables:
                                                all_tables.append(table_rows)
                                            break  # Found a table, move on
                        
                        if all_tables:
                            extracted_data["tables"] = all_tables
                                        
                except ImportError:
                    # pdfplumber not available, continue with PyPDF2 results
                    pass
                except Exception as e:
                    print(f"Error using pdfplumber: {str(e)}")
                    import traceback
                    print(traceback.format_exc())
                    # Continue with PyPDF2 results
                
            except ImportError:
                print("PyPDF2 not installed. Install it with: pip install PyPDF2")
                # Try pdfplumber as alternative
                try:
                    import pdfplumber
                    with pdfplumber.open(io.BytesIO(file_content)) as pdf:
                        text_content = ""
                        tables_data = []
                        for page in pdf.pages:
                            text_content += page.extract_text() + "\n"
                            tables = page.extract_tables()
                            if tables:
                                tables_data.extend(tables)
                        
                        extracted_data = {
                            "raw_text": text_content,
                            "page_count": len(pdf.pages),
                            "extracted_fields": {},
                            "tables": tables_data if tables_data else None
                        }
                except ImportError:
                    print("Neither PyPDF2 nor pdfplumber installed. Install one with: pip install PyPDF2 or pip install pdfplumber")
                    return None
                except Exception as e:
                    print(f"Error parsing PDF file with pdfplumber: {str(e)}")
                    return None
            except Exception as e:
                print(f"Error parsing PDF file: {str(e)}")
                return None
        
        return extracted_data
        
    except requests.RequestException as e:
        print(f"Error downloading file from {file_url}: {str(e)}")
        return None
    except Exception as e:
        print(f"Error parsing document: {str(e)}")
        return None


def normalize_date(date_value):
    """
    Normalize date string to YYYY-MM-DD format.
    Handles various date formats like:
    - "24 november 2025"
    - "24-nov-2025"
    - "24/11/2025"
    - "2025-11-24"
    - "2025-04-30"
    - etc.
    
    Returns empty string if date_value is empty or cannot be parsed.
    """
    if not date_value or not isinstance(date_value, str):
        return ""
    
    date_value = date_value.strip()
    if not date_value:
        return ""
    
    # If already in YYYY-MM-DD format, return as is
    if re.match(r'^\d{4}-\d{2}-\d{2}$', date_value):
        return date_value
    
    # Try multiple date formats
    date_formats = [
        '%Y-%m-%d',           # 2025-04-30
        '%d-%m-%Y',           # 24-11-2025
        '%d/%m/%Y',           # 24/11/2025
        '%m/%d/%Y',           # 11/24/2025
        '%Y/%m/%d',           # 2025/11/24
        '%d %B %Y',           # 24 November 2025
        '%d %b %Y',           # 24 Nov 2025
        '%B %d, %Y',          # November 24, 2025
        '%b %d, %Y',          # Nov 24, 2025
        '%d-%b-%Y',           # 24-Nov-2025
        '%d-%B-%Y',           # 24-November-2025
        '%Y-%b-%d',           # 2025-Nov-24
        '%Y-%B-%d',           # 2025-November-24
        '%d.%m.%Y',           # 24.11.2025
        '%Y.%m.%d',           # 2025.11.24
    ]
    
    for date_format in date_formats:
        try:
            parsed_date = datetime.strptime(date_value, date_format)
            return parsed_date.strftime('%Y-%m-%d')
        except ValueError:
            continue
    
    # Try dateutil parser if available (more flexible)
    try:
        from dateutil import parser as date_parser
        parsed_date = date_parser.parse(date_value)
        return parsed_date.strftime('%Y-%m-%d')
    except (ImportError, ValueError, TypeError):
        pass
    
    # If all parsing fails, return empty string
    return ""


def extract_received_date(quotation_fields_from_doc, policy_fields=None, extracted_fields=None, details=None):
    """
    Extract received_date from document data by checking multiple possible field names.
    Checks in order:
    1. received_date
    2. quotation_received_date
    3. issued_date
    4. quotation_issued_date
    5. from_date
    6. effective_from
    7. effective_from_date
    
    Also checks in policy_fields, extracted_fields, and details if provided.
    Returns empty string if not found.
    """
    # Priority order of field names to check
    field_names = [
        "received date",
        "quotation received date",
        "issued date",
        "quotation issued date",
        "issue date",
        "quotation issue date",
        "from date",
        "effective from",
        "effective from date",
        "effective_from",
        "effective_from_date"
    ]
    
    # Check in quotation_fields_from_doc first
    if quotation_fields_from_doc:
        for field_name in field_names:
            value = quotation_fields_from_doc.get(field_name, "")
            if value and str(value).strip():
                return str(value).strip()
    
    # Check in policy_fields if provided
    if policy_fields:
        for field_name in field_names:
            value = policy_fields.get(field_name, "")
            if value and str(value).strip():
                return str(value).strip()
    
    # Check in extracted_fields if provided
    if extracted_fields:
        for field_name in field_names:
            value = extracted_fields.get(field_name, "")
            if value and str(value).strip():
                return str(value).strip()
    
    # Check in details if provided
    if details:
        for field_name in field_names:
            value = details.get(field_name, "")
            if value and str(value).strip():
                return str(value).strip()
    
    return ""


def clean_total_amount(total_amount_value):
    """
    Clean total_amount by removing currency symbols (LKR, RS, Rs, etc.) and other non-numeric characters.
    Preserves commas and decimal points.
    Returns only the numeric value with commas.
    
    Examples:
    - "LKR 3,950,000" -> "3,950,000"
    - "RS 1,000" -> "1,000"
    - "Rs. 500" -> "500"
    - "3,950,000" -> "3,950,000"
    - "3950000" -> "3950000"
    
    Returns empty string if total_amount_value is empty or invalid.
    """
    if not total_amount_value:
        return ""
    
    # Convert to string if not already
    if not isinstance(total_amount_value, str):
        total_amount_value = str(total_amount_value)
    
    # Remove whitespace
    cleaned = total_amount_value.strip()
    
    if not cleaned:
        return ""
    
    # Remove common currency symbols and prefixes (case-insensitive)
    currency_patterns = [
        r'^LKR\s*', r'^lkr\s*', r'^Lkr\s*',
        r'^RS\s*', r'^rs\s*', r'^Rs\s*',
        r'^USD\s*', r'^usd\s*',
        r'^EUR\s*', r'^eur\s*',
        r'^\$\s*', r'^€\s*', r'^£\s*', r'^₹\s*',
        r'\s*LKR$', r'\s*lkr$', r'\s*Lkr$',
        r'\s*RS$', r'\s*rs$', r'\s*Rs$',
        r'\s*USD$', r'\s*usd$',
        r'\s*EUR$', r'\s*eur$',
        r'\s*\$$', r'\s*€$', r'\s*£$', r'\s*₹$',
    ]
    
    for pattern in currency_patterns:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
    
    # Remove any remaining whitespace
    cleaned = cleaned.strip()
    
    # Remove any remaining non-numeric characters except commas and decimal point
    cleaned = re.sub(r'[^\d,.]', '', cleaned)
    
    # Return empty string if no digits found
    if not cleaned or not re.search(r'\d', cleaned):
        return ""
    
    return cleaned


def _filter_parsed_document_to_response(extracted_data, insurer_id_from_db=None, insurer_name_from_db=None):
    """
    Filter parsed document result (from parse_document_from_url) into extracted_data and quotation_fields.
    Handles dict format with source external_api or local_parsing. Returns (filtered_extracted_data, quotation_fields).
    """
    filtered_extracted_data = {}
    quotation_fields = {}
    if not extracted_data or not isinstance(extracted_data, dict):
        return filtered_extracted_data, quotation_fields

    # Handle external API format (dict with source="external_api")
    if extracted_data.get("source") == "external_api":
        policy_fields = extracted_data.get("policy_fields", {})
        extracted_fields = extracted_data.get("extracted_fields", {})
        raw_analysis = extracted_data.get("raw_analysis", {})
        details = raw_analysis.get("details", {}) if isinstance(raw_analysis, dict) else {}
        policy_fields_from_details = details.get("policy_fields", {})
        quotation_fields_from_doc = details.get("quotation_fields", {})

        all_fields = {}
        all_fields.update(extracted_fields)
        all_fields.update(policy_fields)
        all_fields.update(policy_fields_from_details)

        filtered_extracted_data = {
            "payment_mode": all_fields.get("payment_mode", ""),
            "policy_issue_date": all_fields.get("policy_issue_date", ""),
            "start_date": all_fields.get("start_date", ""),
            "end_date": all_fields.get("end_date", ""),
            "policy_period_from_date": all_fields.get("policy_period_from_date", ""),
            "policy_period_to_date": all_fields.get("policy_period_to_date", ""),
            "product_name": all_fields.get("product_name", ""),
        }
        if quotation_fields_from_doc:
            insurer_company_id_value = None
            if insurer_id_from_db:
                insurer_company_id_value = insurer_id_from_db
            elif quotation_fields_from_doc.get("insurer_company_id"):
                try:
                    insurer_company_id_value = int(quotation_fields_from_doc.get("insurer_company_id"))
                except (ValueError, TypeError):
                    insurer_company_id_value = None
            quotation_fields = {
                "insurer_company_name": insurer_name_from_db if insurer_name_from_db else quotation_fields_from_doc.get("insurer_company_name", ""),
                "insurer_company_id": insurer_company_id_value,
                "received_date": extract_received_date(quotation_fields_from_doc, policy_fields_from_details, extracted_fields, details),
                "expiry_date": quotation_fields_from_doc.get("expiry_date", ""),
                "total_amount": clean_total_amount(quotation_fields_from_doc.get("total_amount", "")),
                "revised": quotation_fields_from_doc.get("revised", ""),
                "uploaded_by": quotation_fields_from_doc.get("uploaded_by", ""),
            }
        elif insurer_id_from_db or insurer_name_from_db:
            quotation_fields = {
                "insurer_company_name": insurer_name_from_db if insurer_name_from_db else "",
                "insurer_company_id": insurer_id_from_db if insurer_id_from_db else None,
                "received_date": extract_received_date(None, policy_fields_from_details, extracted_fields, details),
                "expiry_date": "",
                "total_amount": "",
                "revised": "",
                "uploaded_by": "",
            }
        # Normalize dates
        for field in ["policy_issue_date", "start_date", "end_date", "policy_period_from_date", "policy_period_to_date"]:
            if field in filtered_extracted_data:
                filtered_extracted_data[field] = normalize_date(filtered_extracted_data[field])
        for field in ["received_date", "expiry_date"]:
            if field in quotation_fields:
                quotation_fields[field] = normalize_date(quotation_fields[field])
        return filtered_extracted_data, quotation_fields

    # Handle local parsing or other dict format
    extracted_fields = extracted_data.get("extracted_fields", {})
    policy_fields = extracted_data.get("policy_fields", {})
    raw_analysis = extracted_data.get("raw_analysis", {})
    details = raw_analysis.get("details", {}) if isinstance(raw_analysis, dict) else {}
    quotation_fields_from_doc = details.get("quotation_fields", {})

    all_fields = {}
    all_fields.update(extracted_fields)
    all_fields.update(policy_fields)

    filtered_extracted_data = {
        "payment_mode": all_fields.get("payment_mode", ""),
        "policy_issue_date": all_fields.get("policy_issue_date", ""),
        "start_date": all_fields.get("start_date", ""),
        "end_date": all_fields.get("end_date", ""),
        "policy_period_from_date": all_fields.get("policy_period_from_date", ""),
        "policy_period_to_date": all_fields.get("policy_period_to_date", ""),
        "product_name": all_fields.get("product_name", ""),
    }
    if quotation_fields_from_doc:
        insurer_company_id_value = None
        if insurer_id_from_db:
            insurer_company_id_value = insurer_id_from_db
        elif quotation_fields_from_doc.get("insurer_company_id"):
            try:
                insurer_company_id_value = int(quotation_fields_from_doc.get("insurer_company_id"))
            except (ValueError, TypeError):
                insurer_company_id_value = None
        quotation_fields = {
            "insurer_company_name": insurer_name_from_db if insurer_name_from_db else quotation_fields_from_doc.get("insurer_company_name", ""),
            "insurer_company_id": insurer_company_id_value,
            "received_date": extract_received_date(quotation_fields_from_doc, policy_fields, extracted_fields, details),
            "expiry_date": quotation_fields_from_doc.get("expiry_date", ""),
            "total_amount": clean_total_amount(quotation_fields_from_doc.get("total_amount", "")),
            "revised": quotation_fields_from_doc.get("revised", ""),
            "uploaded_by": quotation_fields_from_doc.get("uploaded_by", ""),
        }
    elif insurer_id_from_db or insurer_name_from_db:
        quotation_fields = {
            "insurer_company_name": insurer_name_from_db if insurer_name_from_db else "",
            "insurer_company_id": insurer_id_from_db if insurer_id_from_db else None,
            "received_date": extract_received_date(None, policy_fields, extracted_fields, details),
            "expiry_date": "",
            "total_amount": "",
            "revised": "",
            "uploaded_by": "",
        }
    for field in ["policy_issue_date", "start_date", "end_date", "policy_period_from_date", "policy_period_to_date"]:
        if field in filtered_extracted_data:
            filtered_extracted_data[field] = normalize_date(filtered_extracted_data[field])
    for field in ["received_date", "expiry_date"]:
        if field in quotation_fields:
            quotation_fields[field] = normalize_date(quotation_fields[field])
    return filtered_extracted_data, quotation_fields


def _extract_vendor_response_detailed_sections(parsed_result):
    """
    Extract the 9 detailed sections from a parsed document (for vendor-responses endpoint only).
    Works with raw_analysis from external API (if details contain these sections) or raw_text from local PDF.
    Returns dict: insurance_agency_details, customer_details, agent_prepared_by, policy_information,
    coverage_amounts, deductibles, discounts_mentioned, endorsements_addons, premium_cost_details.
    """
    empty_sections = {
        "insurance_agency_details": {"agency_name": "", "address": "", "phone": "", "website": ""},
        "customer_details": {"customer_name": "", "address": ""},
        "agent_prepared_by": {"agent_name": "", "phone": "", "email": ""},
        "policy_information": {"policy_type": "", "property_address": ""},
        "coverage_amounts": [],
        "deductibles": {"all_perils_deductible": "", "wind_hail_deductible": "", "named_storm_deductible": ""},
        "discounts_mentioned": [],
        "endorsements_addons": [],
        "premium_cost_details": {
            "premium": "", "flood_insurance_optional": "", "agency_fee": "",
            "total_without_flood": "", "total_with_flood": "",
        },
    }
    if not parsed_result or not isinstance(parsed_result, dict):
        return empty_sections

    # Prefer structure from external API if present
    raw_analysis = parsed_result.get("raw_analysis") or {}
    details = raw_analysis.get("details", {}) if isinstance(raw_analysis, dict) else {}
    if details.get("insurance_agency_details") or details.get("agency_details"):
        out = dict(empty_sections)
        agency = details.get("insurance_agency_details") or details.get("agency_details") or {}
        if isinstance(agency, dict):
            out["insurance_agency_details"] = {
                "agency_name": agency.get("agency_name", ""),
                "address": agency.get("address", ""),
                "phone": agency.get("phone", ""),
                "website": agency.get("website", ""),
            }
        cust = details.get("customer_details") or {}
        if isinstance(cust, dict):
            out["customer_details"] = {"customer_name": cust.get("customer_name", ""), "address": cust.get("address", "")}
        agent = details.get("agent_prepared_by") or details.get("agent_details") or {}
        if isinstance(agent, dict):
            out["agent_prepared_by"] = {
                "agent_name": agent.get("agent_name", ""),
                "phone": agent.get("phone", ""),
                "email": agent.get("email", ""),
            }
        policy = details.get("policy_information") or details.get("policy_info") or {}
        if isinstance(policy, dict):
            out["policy_information"] = {
                "policy_type": policy.get("policy_type", ""),
                "property_address": policy.get("property_address", ""),
            }
        out["coverage_amounts"] = details.get("coverage_amounts") or []
        ded = details.get("deductibles") or {}
        if isinstance(ded, dict):
            out["deductibles"] = {
                "all_perils_deductible": ded.get("all_perils_deductible", ""),
                "wind_hail_deductible": ded.get("wind_hail_deductible", ""),
                "named_storm_deductible": ded.get("named_storm_deductible", ""),
            }
        out["discounts_mentioned"] = details.get("discounts_mentioned") or []
        out["endorsements_addons"] = details.get("endorsements_addons") or details.get("endorsements") or []
        prem = details.get("premium_cost_details") or details.get("premium_details") or {}
        if isinstance(prem, dict):
            out["premium_cost_details"] = {
                "premium": prem.get("premium", ""),
                "flood_insurance_optional": prem.get("flood_insurance_optional", ""),
                "agency_fee": prem.get("agency_fee", ""),
                "total_without_flood": prem.get("total_without_flood", ""),
                "total_with_flood": prem.get("total_with_flood", ""),
            }
        return out

    # Parse from raw text (local PDF or full text from API)
    raw_text = parsed_result.get("raw_text") or ""
    if not raw_text and isinstance(raw_analysis, dict):
        raw_text = raw_analysis.get("raw_text") or details.get("raw_text") or ""
    if not raw_text or not isinstance(raw_text, str):
        return empty_sections

    out = dict(empty_sections)
    lines = [ln.strip() for ln in raw_text.split("\n") if ln.strip()]
    text_lower = raw_text.lower()

    # Helper: find value after "key:" in a given text block (e.g. section)
    def find_key_in_block(block, key_patterns):
        if not block:
            return ""
        for line in block.split("\n"):
            line = line.strip()
            for k in key_patterns:
                if k in line.lower() and ":" in line:
                    idx = line.find(":")
                    if idx >= 0:
                        return line[idx + 1 :].strip()
        return ""

    # Split into sections by numbered headers "1. ...", "2. ..." or unnumbered "Insurance Agency Details", "Customer Details", etc.
    section_starts = []
    for i, line in enumerate(lines):
        if re.match(r"^\d+\.\s+", line):
            section_starts.append((i, line))
    section_blocks = {}
    for idx, (i, header) in enumerate(section_starts):
        j = section_starts[idx + 1][0] if idx + 1 < len(section_starts) else len(lines)
        section_blocks[header.lower()[:50]] = "\n".join(lines[i:j])

    def get_section(*names):
        for n in names:
            for k, v in section_blocks.items():
                if n in k:
                    return v
        return ""

    # Full-text fallback: when there are no numbered sections, extract key: value from entire text (e.g. "Agency Name: XYZ")
    def find_key_in_full_text(key_aliases):
        for alias in key_aliases:
            m = re.search(r"{}\s*:\s*(.+?)(?=\n[A-Za-z]|\n\s*\n|$)".format(re.escape(alias)), raw_text, re.IGNORECASE | re.DOTALL)
            if m:
                return m.group(1).strip()
            m = re.search(r"{}\s*:\s*(.+?)(?=\n|$)".format(re.escape(alias)), raw_text, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""

    # Helper: extract chunk between two markers for full-text fallback when no numbered sections
    def chunk_between(marker_start, marker_end=None):
        start = raw_text.lower().find(marker_start.lower())
        if start < 0:
            return ""
        search_from = start + len(marker_start)
        end = len(raw_text)
        if marker_end:
            end_match = re.search(re.escape(marker_end), raw_text[search_from:], re.IGNORECASE)
            if end_match:
                end = search_from + end_match.start()
        return raw_text[start : min(start + 800, end)]

    # 1. Insurance Agency Details
    block1 = get_section("insurance agency details", "agency details")
    if not block1:
        block1 = chunk_between("Insurance Agency Details", "Customer Details") or chunk_between("Agency Details", "Customer")
    if block1:
        out["insurance_agency_details"] = {
            "agency_name": find_key_in_block(block1, ["agency name:"]) or find_key_in_full_text(["Agency Name"]),
            "address": find_key_in_block(block1, ["address:"]),
            "phone": find_key_in_block(block1, ["phone:"]),
            "website": find_key_in_block(block1, ["website:"]),
        }
        if not out["insurance_agency_details"]["address"]:
            out["insurance_agency_details"]["address"] = find_key_in_full_text(["Address"])
        if not out["insurance_agency_details"]["agency_name"]:
            out["insurance_agency_details"]["agency_name"] = find_key_in_full_text(["Agency Name"])

    # 2. Customer Details
    block2 = get_section("customer details")
    if not block2:
        block2 = chunk_between("Customer Details", "Agent") or chunk_between("Prepared Exclusively For", "Agent")
    if block2:
        out["customer_details"] = {
            "customer_name": find_key_in_block(block2, ["customer name:"]) or find_key_in_full_text(["Customer Name"]),
            "address": find_key_in_block(block2, ["address:"]),
        }
        if not out["customer_details"]["customer_name"]:
            out["customer_details"]["customer_name"] = find_key_in_full_text(["Customer Name"])

    # 3. Agent / Prepared By
    block3 = get_section("agent", "prepared by")
    if not block3:
        block3 = chunk_between("Agent Name", "Policy") or chunk_between("Prepared By", "Policy")
    if block3:
        out["agent_prepared_by"] = {
            "agent_name": find_key_in_block(block3, ["agent name:"]) or find_key_in_full_text(["Agent Name"]),
            "phone": find_key_in_block(block3, ["phone:"]),
            "email": find_key_in_block(block3, ["email:"]) or find_key_in_full_text(["Email"]),
        }

    # 4. Policy Information
    block4 = get_section("policy information", "policy type")
    if not block4:
        block4 = chunk_between("HOB Home Policy", "Coverage") or chunk_between("Policy Type", "Coverage")
    if block4:
        out["policy_information"] = {
            "policy_type": find_key_in_block(block4, ["policy type:"]) or find_key_in_full_text(["Policy Type"]),
            "property_address": find_key_in_block(block4, ["property address:"]) or find_key_in_full_text(["Property Address"]),
        }
        if not out["policy_information"]["policy_type"]:
            m = re.search(r"(HOB\s+Home\s+Policy|Policy\s+Type\s*:\s*[^\n]+)", raw_text, re.IGNORECASE)
            if m:
                out["policy_information"]["policy_type"] = re.sub(r"^Policy\s*Type\s*:\s*", "", m.group(1), flags=re.IGNORECASE).strip()

    # 5. Coverage Amounts – look for "Coverage A", "Dwelling", amounts like $285,000
    coverage_pattern = re.compile(
        r"(Coverage\s+[A-Z]\s*[–\-]\s*[^$\d]+?)\s*\$?([\d,]+(?:\.\d{2})?)",
        re.IGNORECASE,
    )
    for m in coverage_pattern.finditer(raw_text):
        cov_type = m.group(1).strip().strip("–\-")
        amount = m.group(2)
        if cov_type and amount:
            out["coverage_amounts"].append({"coverage_type": cov_type, "amount": amount})
    if not out["coverage_amounts"]:
        # Fallback: line containing "Coverage" and $ number
        for line in lines:
            if "coverage" in line.lower() and "$" in line:
                parts = re.split(r"\$", line, 1)
                if len(parts) == 2:
                    left = parts[0].strip().strip("–\-:")
                    right = re.search(r"[\d,]+(?:\.\d{2})?", parts[1])
                    if right:
                        out["coverage_amounts"].append({"coverage_type": left, "amount": right.group()})

    # 6. Deductibles
    ded_patterns = [
        ("all perils", "all_perils_deductible"),
        ("wind/hail", "wind_hail_deductible"),
        ("named storm", "named_storm_deductible"),
    ]
    for label, key in ded_patterns:
        for line in lines:
            if label in line.lower() and ("deductible" in line.lower() or "%" in line or "$" in line):
                val = re.search(r"\$?[\d,]+(?:\.\d{2})?|[\d.]+%", line)
                if val:
                    out["deductibles"][key] = val.group().strip()
                    break

    # 7. Discounts Mentioned – from section 7 block or chunk between Discounts / Endorsements
    block7 = get_section("discounts mentioned", "discount")
    if not block7:
        block7 = chunk_between("Discounts", "Endorsements") or chunk_between("Discounts Mentioned", "Endorsements")
    if block7:
        for line in block7.split("\n"):
            line = line.strip()
            if not line or "discounts mentioned" in line.lower() or line.lower() == "discounts":
                continue
            if "endorsement" in line.lower() or line.strip().startswith("8."):
                break
            if line and not line.lower().startswith(("premium", "9.")):
                out["discounts_mentioned"].append(line.strip(".–\- "))

    # 8. Endorsements / Add-ons – from section 8 block or chunk between Endorsements / Premium
    block8 = get_section("endorsement", "add-on", "add-ons")
    if not block8:
        block8 = chunk_between("Endorsements", "Premium") or chunk_between("Add-ons", "Premium")
    if block8:
        for line in block8.split("\n"):
            line = line.strip()
            if not line:
                continue
            if ("endorsements" in line.lower() or "add-on" in line.lower()) and len(line) < 40:
                continue  # skip section header
            if "premium" in line.lower() or line.strip().startswith("9.") or "cost detail" in line.lower():
                break
            out["endorsements_addons"].append(line.strip(".–\- "))

    # 9. Premium / Cost Details
    prem_keys = [
        ("premium", "premium"),
        ("flood insurance", "flood_insurance_optional"),
        ("agency fee", "agency_fee"),
        ("total without flood", "total_without_flood"),
        ("total with flood", "total_with_flood"),
    ]
    for line in lines:
        line_lower = line.lower()
        for label, key in prem_keys:
            if label in line_lower and ("$" in line or re.search(r"[\d,]+\.\d{2}", line)):
                val = re.search(r"\$?[\d,]+\.\d{2}", line)
                if val:
                    out["premium_cost_details"][key] = val.group().replace("$", "").strip()
                    break

    return out


def _fetch_pdf_raw_text(file_url):
    """
    Fetch PDF from URL and extract raw text using PyPDF2 or pdfplumber.
    Returns extracted text string or empty string on failure.
    """
    if not file_url:
        return ""
    try:
        resp = requests.get(file_url, timeout=30, stream=True)
        resp.raise_for_status()
        content = resp.content
    except Exception as e:
        print(f"[_fetch_pdf_raw_text] Failed to fetch {file_url[:80]}: {e}")
        return ""
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(io.BytesIO(content))
        text = ""
        for page in reader.pages:
            text += page.extract_text() or ""
            text += "\n"
        return text.strip()
    except Exception as e:
        print(f"[_fetch_pdf_raw_text] PyPDF2 failed: {e}")
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
            return text.strip()
    except Exception as e:
        print(f"[_fetch_pdf_raw_text] pdfplumber failed: {e}")
    return ""


def _is_document_extracted_details_empty(details):
    """Return True if document_extracted_details has no meaningful content."""
    if not details or not isinstance(details, dict):
        return True
    if (details.get("insurance_agency_details") or {}).get("agency_name"):
        return False
    if (details.get("customer_details") or {}).get("customer_name"):
        return False
    if (details.get("coverage_amounts") or []):
        return False
    if (details.get("premium_cost_details") or {}).get("premium"):
        return False
    return True


def extract_document_data_from_coverage_path(coverage_path, file_name, content_type, insurer_id=None, insurer_name=None):
    """
    Extract document data from a coverage file path (e.g. from vendor response coverage_details).
    Builds full URL from CDN_BASE_URL, parses the document, and returns extracted_data and quotation_fields.
    Returns dict with keys: extracted_data, quotation_fields. On failure or unsupported type returns empty dicts.
    """
    _empty_details = lambda: _extract_vendor_response_detailed_sections({})
    if not coverage_path or not file_name:
        return {"extracted_data": {}, "quotation_fields": {}, "document_extracted_details": _empty_details()}
    file_name_lower = (file_name or "").lower()
    content_type_lower = (content_type or "").lower()
    supported = (
        file_name_lower.endswith((".pdf", ".xlsx", ".xls"))
        or "pdf" in content_type_lower
        or "spreadsheet" in content_type_lower
    )
    if not supported:
        return {"extracted_data": {}, "quotation_fields": {}, "document_extracted_details": _empty_details()}
    CDN_BASE_URL = getattr(settings, "CDN_BASE_URL", None)
    if not CDN_BASE_URL:
        return {"extracted_data": {}, "quotation_fields": {}, "document_extracted_details": _empty_details()}
    CDN_BASE_URL = CDN_BASE_URL.rstrip("/")
    file_url = CDN_BASE_URL + "/" + coverage_path.lstrip("/")
    parsed = None
    try:
        parsed = parse_document_from_url(file_url, file_name, content_type or "")
    except Exception as e:
        print(f"[extract_document_data_from_coverage_path] Error parsing {coverage_path}: {e}")
    if not parsed:
        return {"extracted_data": {}, "quotation_fields": {}, "document_extracted_details": _empty_details()}
    filtered_extracted_data, quotation_fields = _filter_parsed_document_to_response(
        parsed, insurer_id_from_db=insurer_id, insurer_name_from_db=insurer_name
    )
    document_extracted_details = _extract_vendor_response_detailed_sections(parsed)
    # When external API was used, parsed often has no raw_text so details are empty; fetch PDF and extract text for detailed sections
    if _is_document_extracted_details_empty(document_extracted_details) and (
        file_name_lower.endswith(".pdf") or "pdf" in (content_type or "").lower()
    ):
        raw_text = _fetch_pdf_raw_text(file_url)
        if raw_text:
            document_extracted_details = _extract_vendor_response_detailed_sections({"raw_text": raw_text})
    return {
        "extracted_data": filtered_extracted_data,
        "quotation_fields": quotation_fields,
        "document_extracted_details": document_extracted_details,
    }


@csrf_exempt
@api_view(["GET"])
def extract_document_data(request, document_id):
    """
    Extract data from a document by document_id (attachment_id from core_email_attachments table).
    """
    try:
        # Get attachment by ID from core_email_attachments table
        attachment = QueryBuilderService("core_email_attachments")\
            .where("id", document_id)\
            .select("core_email_attachments.*")\
            .first()
        
        if not attachment:
            return ResponseService.response("NOT_FOUND", None, "Document attachment not found")
        
        file_name = attachment.get("file_name", "")
        file_url = attachment.get("file_url", "")
        content_type = attachment.get("content_type", "")
        data_analysis = attachment.get("data_analysis")
        
        # Extract file_key from URL by removing CDN base URL
        file_key = None
        if file_url:
            CDN_BASE_URL = getattr(settings, 'CDN_BASE_URL', None)
            if CDN_BASE_URL:
                CDN_BASE_URL = CDN_BASE_URL.rstrip('/')
                if file_url.startswith(CDN_BASE_URL):
                    file_key = file_url.replace(CDN_BASE_URL, "").lstrip('/')
                else:
                    # If URL doesn't start with CDN base URL, try to extract path after domain
                    parsed_url = urlparse(file_url)
                    file_key = parsed_url.path.lstrip('/')
            else:
                # Fallback: extract path from URL
                parsed_url = urlparse(file_url)
                file_key = parsed_url.path.lstrip('/')
        
        # Get insurer_id and insurer_name from database by tracing through tables
        insurer_id_from_db = None
        insurer_name_from_db = None
        try:
            # Step 1: Get email_message_id from attachment
            email_message_id = attachment.get("email_message_id")
            
            if email_message_id:
                # Step 2: Get email message to find conversation_id
                email_message = QueryBuilderService("core_email_messages")\
                    .where("id", email_message_id)\
                    .select("conversation_id")\
                    .first()
                
                if email_message:
                    conversation_id = email_message.get("conversation_id")
                    
                    if conversation_id:
                        # Step 3: Get conversation to find insurer_id
                        conversation = QueryBuilderService("core_chat_conversations")\
                            .where("id", conversation_id)\
                            .select("insurer_id")\
                            .first()
                        
                        if conversation:
                            insurer_id_from_db = conversation.get("insurer_id")
                            
                            # Step 4: Get insurer name from core_service_providers
                            if insurer_id_from_db:
                                service_provider = QueryBuilderService("core_service_providers")\
                                    .where("id", insurer_id_from_db)\
                                    .select("name")\
                                    .first()
                                
                                if service_provider:
                                    insurer_name_from_db = service_provider.get("name", "")
        except Exception as e:
            print(f"Error getting insurer details from database: {str(e)}")
            # Continue without insurer details from DB
        
        # Prioritize data_analysis from database - skip expensive document parsing if available
        extracted_data = None
        
        # First, try to use data_analysis from database (fast path)
        if data_analysis:
            try:
                # Parse JSON if it's a string
                if isinstance(data_analysis, str):
                    extracted_data = json.loads(data_analysis)
                else:
                    extracted_data = data_analysis
                
                # If we successfully got data from data_analysis, skip document parsing
                if extracted_data:
                    # Quick validation - check if it has meaningful data
                    has_data = False
                    if isinstance(extracted_data, list) and len(extracted_data) > 0:
                        has_data = True
                    elif isinstance(extracted_data, dict) and (extracted_data.get("policy_fields") or 
                                                              extracted_data.get("extracted_fields") or 
                                                              extracted_data.get("raw_analysis")):
                        has_data = True
                    
                    if has_data:
                        # We have valid data_analysis, skip expensive parsing
                        pass  # Continue with extracted_data
                    else:
                        # data_analysis exists but is empty/invalid, fall through to parsing
                        extracted_data = None
                        
            except json.JSONDecodeError as e:
                print(f"Error parsing JSON for attachment {document_id}: {str(e)}")
                extracted_data = None  # Will fall through to parsing
            except Exception as e:
                print(f"Error extracting data from data_analysis for attachment {document_id}: {str(e)}")
                extracted_data = None  # Will fall through to parsing
        
        # Only parse document if data_analysis is not available or invalid (slow path)
        if not extracted_data:
            file_name_lower = file_name.lower()
            parsed_document_data = None
            
            # Try to parse the document if it's a supported format
            if file_url and (file_name_lower.endswith(('.xlsx', '.xls', '.pdf')) or 
                           'spreadsheet' in content_type.lower() or 'pdf' in content_type.lower()):
                try:
                    parsed_document_data = parse_document_from_url(file_url, file_name, content_type)
                except Exception as e:
                    print(f"Error parsing document {file_name}: {str(e)}")
            
            extracted_data = parsed_document_data
        
        # Filter extracted_data to only return specific fields
        # Optimize by handling most common format first (list from data_analysis)
        filtered_extracted_data = {}
        quotation_fields = {}
        insurer_id = None
        
        if extracted_data:
            # Handle list format first (most common from data_analysis - fastest path)
            if isinstance(extracted_data, list) and len(extracted_data) > 0:
                first_item = extracted_data[0]
                if isinstance(first_item, dict):
                    details = first_item.get("details", {})
                    policy_fields = details.get("policy_fields", {})
                    quotation_fields_from_doc = details.get("quotation_fields", {})
                    
                    filtered_extracted_data = {
                        "payment_mode": policy_fields.get("payment_mode", ""),
                        "policy_issue_date": policy_fields.get("policy_issue_date", ""),
                        "start_date": policy_fields.get("start_date", ""),
                        "end_date": policy_fields.get("end_date", ""),
                        "policy_period_from_date": policy_fields.get("policy_period_from_date", ""),
                        "policy_period_to_date": policy_fields.get("policy_period_to_date", ""),
                        "product_name": policy_fields.get("product_name", "")
                    }
                    
                    # Extract quotation fields from document
                    if quotation_fields_from_doc:
                        # Convert insurer_company_id to int if it exists in document, otherwise use DB value
                        insurer_company_id_value = None
                        if insurer_id_from_db:
                            insurer_company_id_value = insurer_id_from_db
                        elif quotation_fields_from_doc.get("insurer_company_id"):
                            try:
                                insurer_company_id_value = int(quotation_fields_from_doc.get("insurer_company_id"))
                            except (ValueError, TypeError):
                                insurer_company_id_value = None
                        
                        quotation_fields = {
                            "insurer_company_name": insurer_name_from_db if insurer_name_from_db else quotation_fields_from_doc.get("insurer_company_name", ""),
                            "insurer_company_id": insurer_company_id_value,
                            "received_date": extract_received_date(quotation_fields_from_doc, policy_fields, None, details),
                            "expiry_date": quotation_fields_from_doc.get("expiry_date", ""),
                            "total_amount": clean_total_amount(quotation_fields_from_doc.get("total_amount", "")),
                            "revised": quotation_fields_from_doc.get("revised", ""),
                            "uploaded_by": quotation_fields_from_doc.get("uploaded_by", "")
                        }
                    elif insurer_id_from_db or insurer_name_from_db:
                        # If no quotation_fields in document but we have insurer from DB, create quotation_fields
                        quotation_fields = {
                            "insurer_company_name": insurer_name_from_db if insurer_name_from_db else "",
                            "insurer_company_id": insurer_id_from_db if insurer_id_from_db else None,
                            "received_date": extract_received_date(None, policy_fields, None, details),
                            "expiry_date": "",
                            "total_amount": "",
                            "revised": "",
                            "uploaded_by": ""
                        }
                    
                    # Extract insurer_id from various possible locations (use DB value as priority)
                    insurer_id = insurer_id_from_db or (
                        quotation_fields_from_doc.get("insurer_id") if quotation_fields_from_doc else None
                    ) or policy_fields.get("insurer_id") or details.get("insurer_id") or first_item.get("insurer_id")
            
            # Handle external API format (dict with source="external_api")
            elif isinstance(extracted_data, dict) and extracted_data.get("source") == "external_api":
                # Extract from policy_fields, extracted_fields, or raw_analysis
                policy_fields = extracted_data.get("policy_fields", {})
                extracted_fields = extracted_data.get("extracted_fields", {})
                raw_analysis = extracted_data.get("raw_analysis", {})
                
                # Get details from raw_analysis if available
                details = raw_analysis.get("details", {}) if isinstance(raw_analysis, dict) else {}
                policy_fields_from_details = details.get("policy_fields", {})
                quotation_fields_from_doc = details.get("quotation_fields", {})
                
                # Merge all sources (priority: policy_fields_from_details > policy_fields > extracted_fields)
                all_fields = {}
                all_fields.update(extracted_fields)
                all_fields.update(policy_fields)
                all_fields.update(policy_fields_from_details)
                
                # Extract only the required fields
                filtered_extracted_data = {
                    "payment_mode": all_fields.get("payment_mode", ""),
                    "policy_issue_date": all_fields.get("policy_issue_date", ""),
                    "start_date": all_fields.get("start_date", ""),
                    "end_date": all_fields.get("end_date", ""),
                    "policy_period_from_date": all_fields.get("policy_period_from_date", ""),
                    "policy_period_to_date": all_fields.get("policy_period_to_date", ""),
                    "product_name": all_fields.get("product_name", "")
                }
                
                # Extract quotation fields from document
                if quotation_fields_from_doc:
                    # Convert insurer_company_id to int if it exists in document, otherwise use DB value
                    insurer_company_id_value = None
                    if insurer_id_from_db:
                        insurer_company_id_value = insurer_id_from_db
                    elif quotation_fields_from_doc.get("insurer_company_id"):
                        try:
                            insurer_company_id_value = int(quotation_fields_from_doc.get("insurer_company_id"))
                        except (ValueError, TypeError):
                            insurer_company_id_value = None
                    
                    quotation_fields = {
                        "insurer_company_name": insurer_name_from_db if insurer_name_from_db else quotation_fields_from_doc.get("insurer_company_name", ""),
                        "insurer_company_id": insurer_company_id_value,
                        "received_date": extract_received_date(quotation_fields_from_doc, policy_fields_from_details, extracted_fields, details),
                        "expiry_date": quotation_fields_from_doc.get("expiry_date", ""),
                        "total_amount": clean_total_amount(quotation_fields_from_doc.get("total_amount", "")),
                        "revised": quotation_fields_from_doc.get("revised", ""),
                        "uploaded_by": quotation_fields_from_doc.get("uploaded_by", "")
                    }
                elif insurer_id_from_db or insurer_name_from_db:
                    # If no quotation_fields in document but we have insurer from DB, create quotation_fields
                    quotation_fields = {
                        "insurer_company_name": insurer_name_from_db if insurer_name_from_db else "",
                        "insurer_company_id": insurer_id_from_db if insurer_id_from_db else None,
                        "received_date": extract_received_date(None, policy_fields_from_details, extracted_fields, details),
                        "expiry_date": "",
                        "total_amount": "",
                        "revised": "",
                        "uploaded_by": ""
                    }
                
                # Extract insurer_id from various possible locations (use DB value as priority)
                insurer_id = insurer_id_from_db or (
                    quotation_fields_from_doc.get("insurer_id") if quotation_fields_from_doc else None
                ) or policy_fields_from_details.get("insurer_id") or policy_fields.get("insurer_id") or extracted_fields.get("insurer_id") or details.get("insurer_id") or raw_analysis.get("insurer_id") or extracted_data.get("insurer_id")
            
            # Handle local parsing format or other dict formats
            elif isinstance(extracted_data, dict):
                extracted_fields = extracted_data.get("extracted_fields", {})
                policy_fields = extracted_data.get("policy_fields", {})
                raw_analysis = extracted_data.get("raw_analysis", {})
                
                # Get details from raw_analysis if available
                details = raw_analysis.get("details", {}) if isinstance(raw_analysis, dict) else {}
                quotation_fields_from_doc = details.get("quotation_fields", {})
                
                # Merge sources
                all_fields = {}
                all_fields.update(extracted_fields)
                all_fields.update(policy_fields)
                
                filtered_extracted_data = {
                    "payment_mode": all_fields.get("payment_mode", ""),
                    "policy_issue_date": all_fields.get("policy_issue_date", ""),
                    "start_date": all_fields.get("start_date", ""),
                    "end_date": all_fields.get("end_date", ""),
                    "policy_period_from_date": all_fields.get("policy_period_from_date", ""),
                    "policy_period_to_date": all_fields.get("policy_period_to_date", ""),
                    "product_name": all_fields.get("product_name", "")
                }
                
                # Extract quotation fields from document
                if quotation_fields_from_doc:
                    # Convert insurer_company_id to int if it exists in document, otherwise use DB value
                    insurer_company_id_value = None
                    if insurer_id_from_db:
                        insurer_company_id_value = insurer_id_from_db
                    elif quotation_fields_from_doc.get("insurer_company_id"):
                        try:
                            insurer_company_id_value = int(quotation_fields_from_doc.get("insurer_company_id"))
                        except (ValueError, TypeError):
                            insurer_company_id_value = None
                    
                    quotation_fields = {
                        "insurer_company_name": insurer_name_from_db if insurer_name_from_db else quotation_fields_from_doc.get("insurer_company_name", ""),
                        "insurer_company_id": insurer_company_id_value,
                        "received_date": extract_received_date(quotation_fields_from_doc, policy_fields, extracted_fields, details),
                        "expiry_date": quotation_fields_from_doc.get("expiry_date", ""),
                        "total_amount": clean_total_amount(quotation_fields_from_doc.get("total_amount", "")),
                        "revised": quotation_fields_from_doc.get("revised", ""),
                        "uploaded_by": quotation_fields_from_doc.get("uploaded_by", "")
                    }
                elif insurer_id_from_db or insurer_name_from_db:
                    # If no quotation_fields in document but we have insurer from DB, create quotation_fields
                    quotation_fields = {
                        "insurer_company_name": insurer_name_from_db if insurer_name_from_db else "",
                        "insurer_company_id": insurer_id_from_db if insurer_id_from_db else None,
                        "received_date": extract_received_date(None, policy_fields, extracted_fields, details),
                        "expiry_date": "",
                        "total_amount": "",
                        "revised": "",
                        "uploaded_by": ""
                    }
                
                # Extract insurer_id from various possible locations (use DB value as priority)
                insurer_id = insurer_id_from_db or (
                    quotation_fields_from_doc.get("insurer_id") if quotation_fields_from_doc else None
                ) or policy_fields.get("insurer_id") or extracted_fields.get("insurer_id") or details.get("insurer_id") or raw_analysis.get("insurer_id") or extracted_data.get("insurer_id")
        
        # Normalize all date fields to YYYY-MM-DD format
        if filtered_extracted_data:
            date_fields_in_extracted = [
                "policy_issue_date",
                "start_date",
                "end_date",
                "policy_period_from_date",
                "policy_period_to_date"
            ]
            for field in date_fields_in_extracted:
                if field in filtered_extracted_data:
                    filtered_extracted_data[field] = normalize_date(filtered_extracted_data[field])
        
        if quotation_fields:
            date_fields_in_quotation = [
                "received_date",
                "expiry_date"
            ]
            for field in date_fields_in_quotation:
                if field in quotation_fields:
                    quotation_fields[field] = normalize_date(quotation_fields[field])
        
        # Build response with document metadata, extracted data, and quotation fields
        response_data = {
            "document_name": file_name,
            "document_url": file_url,
            "document_type": content_type,
            "file_key": file_key,
            "extracted_data": filtered_extracted_data,
            "quotation_fields": quotation_fields if quotation_fields else {}
        }
        
        return ResponseService.response("SUCCESS", response_data, "Document data extracted successfully")
    
    except Exception as e:
        print(f"Error in extract_document_data: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return ResponseService.response("INTERNAL_SERVER_ERROR", None, f"Internal server error: {str(e)}")